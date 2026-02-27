import time
from datetime import datetime, timedelta, date
from xml.dom import ValidationErr
import pandas as pd
from django.db import transaction
from django.contrib.auth.models import Permission, Group
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
# from django.utils.translation import ugettext_lazy as _
from io import BytesIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse
# Create your views here.
from hospital.forms import DetailsBillsIngredientForm, DetailsInventoryForm, DetailsPatientAccountForm, DetailsStock_movementForm, HospitalFormRule, PatientAccountForm, CateringInfoForm, EventInfoForm,DeliveryInfoForm, CityForm, DistrictForm, InsuranceForm, RegionForm, Stock_movementForm, Storage_depotsForm, Type_patientForm, UserForm, UserFormUpdate, HospitalForm, \
    PatientForm, Expenses_natureForm, \
    CashForm, Cash_movementForm, CategoryForm,\
    SuppliesForm, SuppliersForm, DetailsSuppliesForm, BillsForm, DetailsBillsForm,PatientSettlementForm, \
    InventoryForm
from hospital.models import CategoryTranslation, ComposeIngredient, DetailsBillsIngredient, DetailsStock_movement, DishPreparation, DishTranslation, ExtendedPermission, DetailsPatientAccount, ExtendedGroup, Ingredient, MovementStock, PatientAccount, RecipeIngredient, District, Insurance, Stock, Storage_depots, Type_patient, User, Hospital, Patient, \
    Expenses_nature,  Cash, Cash_movement, Category, Supplies, Suppliers, \
    DetailsSupplies, Bills, DetailsBills,PatientSettlement, Stock_movement, \
    Inventory, DetailsInventory, City, Region, \
    Module, Archive, BackupFile, DeliveryInfo, EventInfo, CateringInfo, WarehouseTranslation
from hospital.serializers import DetailsBillsIngredientSerializer, DetailsPatientAccountSerializer, DetailsStock_movementSerializer, ExtendedGroupSerializer, IngredientSerializer, MovementStockSerializer,PatientAccountSerializer,DeliveryInfoSerializer, CateringInfoSerializer,EventInfoSerializer, DistrictSerializer, InsuranceSerializer, \
    BillsSerializerAnalysis, Stock_movementSerializer, StockSerializer, Storage_depotsSerializer, Type_patientSerializer,UserSerializer, MyTokenObtainPairSerializer, \
    TokenRefreshSerializer, \
    ChangePasswordSerializer, HospitalSerializer, \
    PatientSerializer, \
     Expenses_natureSerializer, CashSerializer, \
    Cash_movementSerializer, CategorySerializer, SuppliesSerializer, \
    SuppliersSerializer, DetailsSuppliesSerializer, BillsSerializer, DetailsBillsSerializer, \
    PatientSettlementSerializer,  \
    InventorySerializer, DetailsInventorySerializer, CitySerializer, RegionSerializer, \
    PermissionsSerializer, GroupSerializer, ContentTypeSerializer, \
    ModuleSerializer, ArchiveSerializer, UserMeSerializer, BackupFileSerializer, HistorySerializer
from django_filters import rest_framework as filters
from rest_framework import viewsets
from rest_framework.permissions import AllowAny, IsAuthenticated, DjangoModelPermissions
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from globals.pagination import CustomPagination
from rest_framework import status
from .filters import DetailsBillsIngredientFilter, DetailsPatientAccountFilter, DetailsStock_movementFilter, MovementStockFilter, PatientAccountFilter,EventInfoFilter, CateringInfoFilter, DeliveryInfoFilter, DistrictFilter, InsuranceFilter, Stock_movementFilter, StockFilter, Storage_depotsFilter, Type_patientFilter, UserFilter, HospitalFilter, \
    PatientFilter, Expenses_natureFilter, \
    CashFilter, Cash_movementFilter, CategoryFilter,  SuppliesFilter, \
    SuppliersFilter, DetailsSuppliesFilter, BillsFilter, DetailsBillsFilter, PatientSettlementFilter, \
    InventoryFilter, DetailsInventoryFilter, \
     CityFilter, \
    RegionFilter, ModuleFilter, \
    ArchiveFilter, BackupFileFilter
from rest_framework.decorators import action, permission_classes, api_view
from django.shortcuts import render
from collections import OrderedDict
from itertools import chain
from django.db.models import F, Sum,  Count
from track_actions.models import History
from django.shortcuts import get_object_or_404

import shutil
from django.core.management import call_command
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from decimal import Decimal, ROUND_HALF_UP


from .helpers import apply_promotions, apply_reduction, checkContent, checkContentPhone, destocker, destocker_compose, get_applicable_reduction, get_future_remise_notification, get_prepaid_account_detail, link_callback, archive, get_last_date_of_month, get_first_date_of_month, get_back, get_archive, \
    delete_archive, normalize_rules, restore, backup_all, setup_hospital_permissions, split_entry_exit


def home_view(request):
    return render(request, 'index.html')


class HospitalViewSet(viewsets.ModelViewSet):
    queryset = Hospital.objects.filter(deleted=False)
    serializer_class = HospitalSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = HospitalFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        """
        This view should return a list of all the enterprise
        for the currently authenticated user.
        """
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Hospital.objects.filter(id=user_hospital.id, deleted=False)
        return Hospital.objects.filter(deleted=False)

    def create(self, request, *args, **kwargs):
        if request.data['rules_reduction'] != '[]':
            hospital_form = HospitalFormRule(request.data)
            if hospital_form.is_valid():
                hospital = hospital_form.save()
                hospital.save()

                setup_hospital_permissions(hospital)
                # for perm in Permission.objects.all():
                #     ExtendedPermission.objects.get_or_create(
                #         permission=perm,
                #         hospital=hospital,
                #         defaults={'is_shared': False}
                #     )
                # for perm in Permission.objects.all():
                #     for hospital in Hospital.objects.all():
                #         ExtendedPermission.objects.get_or_create(
                #             permission=perm,
                #             hospital=hospital,
                #             defaults={'is_shared': False}
                #         )
                serializer = HospitalSerializer(hospital, many=False)
                return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            errors = {**hospital_form.errors}
        else:
            hospital_form = HospitalForm(request.data)
            if hospital_form.is_valid():
                hospital = hospital_form.save()
                hospital.save()
                setup_hospital_permissions(hospital)
                # for perm in Permission.objects.all():
                #     ExtendedPermission.objects.get_or_create(
                #         permission=perm,
                #         hospital=hospital,
                #         defaults={'is_shared': False}
                #     )
                # for perm in Permission.objects.all():
                #     for hospital in Hospital.objects.all():
                #         ExtendedPermission.objects.get_or_create(
                #             permission=perm,
                #             hospital=hospital,
                #             defaults={'is_shared': False}
                #         )
                serializer = HospitalSerializer(hospital, many=False)
                return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            errors = {**hospital_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        hospital = self.get_object()
        if request.data['rules_reduction'] != '[]':
            hospital_form = HospitalFormRule(request.data, instance=hospital)
            if hospital_form.is_valid():
                hospital = hospital_form.save()
                if request.FILES:
                    if "logo" in request.FILES:
                        for logos in request.FILES.getlist("logo"):
                            hospital.logo = logos

                hospital.save()
                for perm in Permission.objects.all():
                    if ExtendedPermission.objects.filter(permission=perm, hospital=hospital).last():
                        pass
                    else:
                        ExtendedPermission.objects.get_or_create(
                            permission=perm,
                            hospital=hospital,
                            defaults={'is_shared': False}
                        )
                serializer = HospitalSerializer(hospital, many=False)
                return Response(data=serializer.data, status=status.HTTP_200_OK)
            errors = {**hospital_form.errors}
        else:
            hospital_form = HospitalForm(request.data, instance=hospital)
            if hospital_form.is_valid():
                hospital = hospital_form.save()
                if request.FILES:
                    if "logo" in request.FILES:
                        for logos in request.FILES.getlist("logo"):
                            hospital.logo = logos

                hospital.save()
                for perm in Permission.objects.all():
                    if ExtendedPermission.objects.filter(permission=perm, hospital=hospital).last():
                        pass
                    else:
                        ExtendedPermission.objects.get_or_create(
                            permission=perm,
                            hospital=hospital,
                            defaults={'is_shared': False}
                        )
                serializer = HospitalSerializer(hospital, many=False)
                return Response(data=serializer.data, status=status.HTTP_200_OK)
            errors = {**hospital_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        hospital = self.get_object()
        hospital.deleted = True
        hospital.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='statistiques')
    def statistiques(self, request, *args, **kwargs):
        user_hospital = self.request.user.hospital
        if user_hospital:
            get_user = User.objects.filter(hospital=user_hospital).count()
            get_expensive_nature = Expenses_nature.objects.filter(hospital=user_hospital).count()
            get_group = ExtendedGroup.objects.filter(hospital=user_hospital).count()
        else:
            get_user = User.objects.count()
            get_expensive_nature = Expenses_nature.objects.count()
            get_group = ExtendedGroup.objects.count()
        content = {'content': {
            'user': get_user,
            'role': get_group,
            'expenses_nature': get_expensive_nature
        }}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='name/exists')
    def exists(self, request):
        data = request.data
        errors = {"name": ["This field is required."]}
        if 'name' in data:
            association = Hospital.objects.filter(id = self.request.user.hospital.id,name__icontains=data['name'])
            if association:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['PATCH'], url_path='isInventory')
    def isInventory(self, request):
        hospital = Hospital.objects.last()
        hospital.is_inventory = request.data['is_inventory']
        hospital.save()
        return Response(status=status.HTTP_200_OK)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(deleted=False)
    serializer_class = UserSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = UserFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        
        user_hospital = self.request.user.hospital
        if user_hospital:
            return User.objects.filter(deleted=False, hospital=user_hospital)
        return User.objects.filter(deleted=False)
        

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        user_form = UserForm(request.data)
        if user_form.is_valid():
            user = user_form.save()
            if self.request.user.is_superuser:
                pass
            else:
                user.hospital = self.request.user.hospital
            user.set_password(user_form.cleaned_data.get('password'))
            my_group = Group.objects.get(id=request.data['groups'])
            my_group.user_set.add(user)
            user.role = my_group.name
            user.save()
            ExtendedGroup.objects.create(group=my_group, hospital=user.hospital)
            
            serializer = UserSerializer(user, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**user_form.errors, }
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    # def create(self, request, *args, **kwargs):
    #     user_form = UserForm(request.data)
    #     if user_form.is_valid():
    #         get_user = User.objects.first()
    #         if get_user:
    #             regex = re.compile(r'[\d]+')
    #             find = re.findall(regex, get_user.code)
    #             code = str('%04d' % (int(find[0]) + 1))
    #             user = user_form.save()
    #             user.set_password(user_form.cleaned_data.get('password'))
    #             user.code = get_user.code[0:3] + code
    #             user.save()
    #             serializer = self.get_serializer(user, many=False)
    #             return Response(data=serializer.data, status=status.HTTP_201_CREATED)
    #         else:
    #             user = user_form.save()
    #             user.set_password(user_form.cleaned_data.get('password'))
    #             user.save()
    #             serializer = UserSerializer(user, many=False)
    #             return Response(data=serializer.data, status=status.HTTP_201_CREATED)
    #     errors = {**user_form.errors, }
    #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        user = self.get_object()
        user_form = UserFormUpdate(request.data, instance=user)
        if user_form.is_valid():
            user = user_form.save()
            user.groups.clear()
            my_group = Group.objects.get(id=request.data['groups'])
            my_group.user_set.add(user)
            user.role = my_group.name
            user.save()
            serializer = UserSerializer(user, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**user_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['GET'], url_path='history')
    def getHistory(self, request, *args, **kwargs):
        history = History.objects.all()
        serializer = HistorySerializer(history, many=True)
        
        content = {'content': serializer.data}
        return Response(data=content,
                        status=status.HTTP_200_OK
                        )

    @action(detail=False, methods=['post'], url_path='username/exists', permission_classes=[AllowAny])
    def check_username(self, request, *args, **kwargs):
        data = request.data
        errors = {"username": ["This field already exists."]}
        if 'username' in data:
            users = User.objects.filter(hospital = self.request.user.hospital,username=data['username'])
            if users:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='email/exists', permission_classes=[AllowAny])
    def check_email(self, request, *args, **kwargs):
        data = request.data
        errors = {"email": ["This field already exists."]}
        if 'email' in data:
            users = User.objects.filter(hospital = self.request.user.hospital,email=data['email'])
            if users:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='phone/exists', permission_classes=[AllowAny])
    def check_phone(self, request, *args, **kwargs):
        data = request.data
        errors = {"phone": ["This field already exists."]}
        if 'phone' in data:
            users = User.objects.filter(hospital = self.request.user.hospital,phone=data['phone'])
            if users:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post', 'get'], url_path='roles/exists', permission_classes=[AllowAny])
    def check_phone(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            users = ExtendedGroup.objects.filter(hospital = self.request.user.hospital, group__name__icontains=data['name'])
            if users:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET'], url_path='me')
    def get(self, request):
        # Hospital.objects.filter().update(address_ip = request.META['REMOTE_ADDR'])
        groupPerm = list()
        serializer = UserMeSerializer(request.user).data
        get_hospital = self.request.user.hospital
        if get_hospital:
            type_enterprise = get_hospital.type_enterprise
        else:
            type_enterprise=''
        if serializer['groups']:
            for perm in serializer['groups'][0]['permissions']:
                groupPerm.append(dict(perm)['codename'])
            serializer['permissions'] = groupPerm
            serializer['server_ip'] = request.META['REMOTE_ADDR']
            serializer['server_port'] = request.META['SERVER_PORT']
            serializer['type_enterprise'] = type_enterprise
            return Response(serializer)
        return Response(serializer)

    @action(detail=False, methods=['PUT'], url_path='change_password')
    def change_password(self, request):
        user = request.user
        serializer = ChangePasswordSerializer(data=request.data)
        message = {"password": ["Password updated successfully."]}
        error = {"error": ["Your old password was entered incorrectly. Please enter it again."]}
        if serializer.is_valid():
            if user.check_password(serializer.data.get("old_password")):
                user.set_password(serializer.data.get("new_password"))
                user.save()
                # print(user)
                return Response(data=message, status=status.HTTP_200_OK)
            return Response(data=error, status=status.HTTP_400_BAD_REQUEST)

        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST', 'GET', 'DELETE'], url_path='roles')
    def roles(self, request):
        if request.method == 'DELETE':
            group_id = request.query_params.get("id")
            if not group_id:
                return Response({"error": "Group ID is required."}, status=status.HTTP_400_BAD_REQUEST)

            group = get_object_or_404(
                ExtendedGroup,
                id=group_id,
                hospital=request.user.hospital
            )

            group.group.permissions.clear()  # clear Django Group permissions
            group.group.delete()             # delete the underlying Group object
            group.delete()

            return Response(status=status.HTTP_200_OK)
        else:
            if self.request.user.hospital:
                group = ExtendedGroup.objects.filter(hospital = self.request.user.hospital)
            else:
                group = ExtendedGroup.objects.filter(hospital__isnull=True)
            serializer = ExtendedGroupSerializer(group, many=True)
            content = {'content': serializer.data}
            return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST', 'GET'], url_path='contentTypes/permissions')
    def contentTypePermission(self, request):
        hospital = self.request.user.hospital
        if hospital:

            permissions = Permission.objects.filter(extended_permissions__hospital=hospital)
        else:
            permissions = Permission.objects.all()
        serializer_perm = PermissionsSerializer(permissions, many=True)
        # get_module=Module.objects.all()
        # serial_module=ModuleSerializer(get_module,many=True)
        # for perm in serial_module.data:
        #     update=dict(perm)
        #     get_perm=Permission.objects.filter(module_id=update['id'])
        #     update['contentType']=get_perm.content_type
        #     print(update)
        content = {'content': serializer_perm.data}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST', 'GET'], url_path='contentTypes')
    def contentType(self, request):

        contentType = ContentType.objects.filter(hospital = self.request.user.hospital,)
        serializer = ContentTypeSerializer(contentType, many=True)
        content = {'content': serializer.data}
        # remove_model=['group', 'city', 'region', 'logentry', 'manufacturer', 'provider', 'price_level']
        # content_type=[]
        # for model in serializer.data:
        #     if model['model'] in remove_model:
        #         print(model)
        #         print(model.index(model['model']))
        #     else:
        #         content_type.append(model)
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST', 'GET', 'PUT'], url_path='permissions')
    def permissions(self, request):
        try:
            if request.method == 'POST':
                checkbox = request.data.get('checkbox', False)
                group_name = request.data.get('name')

                # Vérifier que le nom existe
                if not group_name:
                    return Response({"error": "Group name required"}, status=400)

                # Récupérer ou créer le groupe
                group, created = Group.objects.get_or_create(name=group_name)

                # ExtendedGroup
                if self.request.user.is_superuser:
                    ExtendedGroup.objects.get_or_create(
                        group=group,
                        defaults={"hospital": None}
                    )
                else:
                    ExtendedGroup.objects.get_or_create(
                        group=group,
                        defaults={"hospital": self.request.user.hospital}
                    )

                # Attribution permissions
                if checkbox:
                    if self.request.user.is_superuser:
                        permissions = Permission.objects.all()
                    else:
                        permissions = Permission.objects.filter(
                            extended_permissions__hospital=self.request.user.hospital,
                            extended_permissions__is_active=True
                        )

                    group.permissions.set(permissions)

                else:
                    permission_ids = request.data.get('permission', [])
                    permissions = Permission.objects.filter(id__in=permission_ids)
                    group.permissions.set(permissions)

                return Response(status=status.HTTP_201_CREATED)
            elif request.method == 'GET':
                # Get all groups linked to the user's hospital
                
                hospital = request.user.hospital  # or any hospital instance
                if hospital:

                    permissions = ExtendedPermission.objects.filter(
                            hospital=hospital,
                            is_active=True
                        )
                else:

                    permissions = Permission.objects.filter(content_type=self.request.query_params.get("contentType"))
                # permission = Permission.objects.filter(content_type=self.request.query_params.get("contentType"), hospital=self.request.user.hospital)
                serializer = PermissionsSerializer(permissions, many=True)
                content = {'content': serializer.data}
                return Response(data=content, status=status.HTTP_200_OK)
            else:
                hospital=self.request.user.hospital
                if hospital:
                    if request.data['checkbox']:
                        
                        group = ExtendedGroup.objects.get(hospital=hospital, id=self.request.query_params.get("id"))
                        group.group.permissions.clear()
                        permissions_list = Permission.objects.filter(extension__hospital=hospital)
                        for permission_index in permissions_list:
                            permission = Permission.objects.get(id=permission_index)
                            group.group.permissions.add(permission)
                        
                        # group.group.permissions.set(permissions_list)
                    else:
                        group = ExtendedGroup.objects.get(hospital=hospital,id=self.request.query_params.get("id"))
                        group.group.permissions.clear()
                        # permissions = request.data['permission']
                        # permissions_list_id = []
                        #
                        # for perm in permissions:
                        #     p = Permission.objects.get(id=perm)
                        #     print(p)
                        #     permissions_list_id.append(p.id)

                        # new_group.permissions.set(permissions_list_id)
                        #
                        try:
                            permissions = request.data['permission']
                            group.group.permissions.set(permissions)
                            # for permission_index in permissions:
                            #     permission = Permission.objects.get(id=permission_index)
                            #     group.permissions.add(permission)
                        except Exception as e:
                            print("Error in creating")
                    return Response(status=status.HTTP_201_CREATED)
                else:
                    
                    if request.data['checkbox']:
                        extended_group = ExtendedGroup.objects.filter(
                            id=self.request.query_params.get("id"),
                            hospital__isnull=True
                        ).select_related('group').last()
                        if self.request.user.hospital:
                            permissions_list = Permission.objects.filter(extension__hospital=self.request.user.hospital)
                        else:
                            permissions_list = Permission.objects.all()
                        print(extended_group.group.id, permissions_list)
                        if extended_group:
                            # On travaille sur le vrai Group Django
                            try:
                                for permission_index in permissions_list:
                                    permission = Permission.objects.get(id=permission_index)
                                    extended_group.group.permissions.add(permission)
                            except Exception as e:
                                print("Error in creating")
                    else:
                        group = ExtendedGroup.objects.get(id=self.request.query_params.get("id"))
                        group.permissions.clear()
                        # permissions = request.data['permission']
                        # permissions_list_id = []
                        #
                        # for perm in permissions:
                        #     p = Permission.objects.get(id=perm)
                        #     print(p)
                        #     permissions_list_id.append(p.id)

                        # new_group.permissions.set(permissions_list_id)
                        #
                        try:
                            permissions = request.data['permission']
                            group.group.permissions.set(permissions)
                            # for permission_index in permissions:
                            #     permission = Permission.objects.get(id=permission_index)
                            #     group.permissions.add(permission)
                        except Exception as e:
                            print("Error in creating")
                    return Response(status=status.HTTP_201_CREATED)
        except Exception as e:
            print("Error in creating", e)
            return Response(data=e, status=status.HTTP_404_NOT_FOUND)

class StockViewSet(viewsets.ModelViewSet):
    queryset = Stock.objects.all()
    serializer_class = StockSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    filterset_class = StockFilter

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'], url_path='all')
    def get_all_stock(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='peremption')
    def get_all_stock_peremption(self, request, *args, **kwargs):
        
        get_hospital = Hospital.objects.last()
        get_product = Ingredient.objects.all()
        get_storage_depot = Storage_depots.objects.all()
        queryset=[]
        if get_product:
            for product in get_product:
                for depot in get_storage_depot:
                    querysets = Stock.objects.filter(ingredient_id=product.id,storage_depots_id=depot.id,quantity__lte=0).last()
                    if querysets:
                        queryset.append(querysets)
        serializer = self.get_serializer(queryset, many=True, fields=('id','ingredient', 'quantity', 'storage_depots')).data
        # today = datetime.today().strftime("%Y-%m-%d")
        if get_hospital:
            # end_date = datetime.now() + relativedelta(days=int(get_hospital.days_before_expiry_date))
            # queryset_expired = Ingredient.objects.filter(expiry_date__lte=end_date).all()
            # serializer_product = IngredientSerializer(queryset_expired, many=True ,fields=('id','name', 'code')).data
            if 'export' in request.query_params:
                html_render = get_template('peremption.html')
                if request.query_params.get('export')=='qty':
                    # logo = settings.MEDIA_ROOT + '/logo.png'
                    html_content = html_render.render(
                        {'products': queryset, 'export': 'qty',
                            'hospital': get_hospital})
                else:
                    # logo = settings.MEDIA_ROOT + '/logo.png'
                    html_content = html_render.render(
                        {'products': queryset_expired, 'export': 'date',
                            'hospital': get_hospital})
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
                if not pdf.err:
                    response = HttpResponse(content_type='application/pdf')
                    filename = 'Liste des produits' + '.pdf'
                    response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                    response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                    response.write(result.getvalue())
                    return response
                else:
                    return None
            else:
                content = {'content': {'qty':serializer}}
                return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='specify')
    def get_all_stock_specify(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True,
                                         fields=('id', 'ingredient', 'quantity', 'quantity_two')).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='stock_available')
    def stock_availables(self, request):
        # queryset = self.filter_queryset(self.get_queryset())
        if 'date' not in self.request.query_params and 'product' not in self.request.query_params:
            #     queryset = DetailsStock.objects.all()
            queryset = self.filter_queryset(self.get_queryset())
            serialzer = self.get_serializer(queryset, many=True)
            content = {'content': serialzer.data}

            return Response(data=content, status=status.HTTP_200_OK)
        # elif 'date' not in self.request.query_params and 'product' in self.request.query_params:
        #     product = Product.objects.filter(deleted=False, id=request.query_params.get("product")).last()
        #     serializer = ProductSerializer(product, many=False)
        #     prod = [serializer.data]
        #     prod_update = []
        #     for con in prod:
        #         update_prod = dict(con)
        #         stock = DetailsSupplies.objects.filter(product_id=dict(con)['id'],
        #                                                createdAt=request.query_params.get("date")).exclude(
        #             supplies=None).aggregate(
        #             Sum('quantity'))
        #
        #         if stock['quantity__sum'] is None:
        #             sum_stock = 0
        #         else:
        #             sum_stock = stock['quantity__sum']
        #         stock_movement_exit = DetailsStock_movement.objects.filter(createdAt=request.query_params.get("date"),
        #                                                                    details_stock__product_id=dict(con)['id'],
        #                                                                    type_movement='EXIT',
        #                                                                    storage_depots_id=request.query_params.get(
        #                                                                        "id")).exclude(
        #             stock_movement=None).aggregate(
        #             Sum('quantity'))
        #         stock_movement_entry = DetailsStock_movement.objects.filter(createdAt=request.query_params.get("date"),
        #                                                                     details_stock__product_id=dict(con)['id'],
        #                                                                     type_movement='ENTRY',
        #                                                                     storage_depots_id=request.query_params.get(
        #                                                                         "id")).exclude(
        #             stock_movement=None).aggregate(
        #             Sum('quantity'))
        #         if stock_movement_exit['quantity__sum'] is None:
        #             sum_stock_movement_exit = 0
        #         else:
        #             sum_stock_movement_exit = stock_movement_exit['quantity__sum']
        #         if stock_movement_entry['quantity__sum'] is None:
        #             sum_stock_movement_entry = 0
        #         else:
        #             sum_stock_movement_entry = stock_movement_entry['quantity__sum']
        #         bills = DetailsBills.objects.filter(createdAt=request.query_params.get("date"), deleted=False,
        #                                             details_stock__product_id=dict(con)['id'],
        #                                             storage_depots_id=request.query_params.get("id")).exclude(
        #             bills=None).aggregate(
        #             Sum('quantity_served'))
        #         if bills['quantity_served__sum'] is None:
        #             sum_bills = 0
        #         else:
        #             sum_bills = bills['quantity_served__sum']
        #
        #         update_prod['qte_stock'] = sum_stock - sum_stock_movement_exit - sum_bills + sum_stock_movement_entry
        #         stock_cmup = DetailsSupplies.objects.filter(product_id=dict(con)['id'],
        #                                                     createdAt=request.query_params.get("date")).exclude(
        #             supplies=None).order_by('-id').first()
        #         if stock_cmup is None:
        #             update_prod['cmup'] = 0
        #
        #         else:
        #             update_prod['cmup'] = stock_cmup.cmup
        #         prod_update.append(OrderedDict(update_prod))
        #     newlist = sorted(prod_update, key=lambda d: d['id'])
        #     content = {'content': newlist}
        #
        #     return Response(data=content, status=status.HTTP_200_OK)

        else:
            if 'date' not in self.request.query_params:
                queryset = self.filter_queryset(self.get_queryset())
                serialzer = self.get_serializer(queryset, many=True)
                content = {'content': serialzer.data}
                return Response(data=content, status=status.HTTP_200_OK)
            else:
                queryset = self.filter_queryset(self.get_queryset())
                product_list = []
                for prod in queryset:
                    product = Ingredient.objects.filter(deleted=False, id=prod.ingredient_id).last()
                    serializer = IngredientSerializer(product, many=False)
                    prod = serializer.data
                    product_list.append(prod)
                prod_update = []
                for con in product_list:
                    update_prod = dict(con)
                    stock = DetailsSupplies.objects.filter(product_id=dict(con)['id'],
                                                           createdAt=request.query_params.get("date")).exclude(
                        supplies=None).aggregate(
                        Sum('quantity'))

                    if stock['quantity__sum'] is None:
                        sum_stock = 0
                    else:
                        sum_stock = stock['quantity__sum']
                    stock_movement_exit = DetailsStock_movement.objects.filter(
                        createdAt=request.query_params.get("date"),
                        details_stock__product_id=dict(con)['id'],
                        type_movement='EXIT',
                        storage_depots_id=request.query_params.get(
                            "id")).exclude(
                        stock_movement=None).aggregate(
                        Sum('quantity'))
                    stock_movement_entry = DetailsStock_movement.objects.filter(
                        createdAt=request.query_params.get("date"),
                        details_stock__product_id=dict(con)['id'],
                        type_movement='ENTRY',
                        storage_depots_id=request.query_params.get(
                            "id")).exclude(
                        stock_movement=None).aggregate(
                        Sum('quantity'))
                    if stock_movement_exit['quantity__sum'] is None:
                        sum_stock_movement_exit = 0
                    else:
                        sum_stock_movement_exit = stock_movement_exit['quantity__sum']
                    if stock_movement_entry['quantity__sum'] is None:
                        sum_stock_movement_entry = 0
                    else:
                        sum_stock_movement_entry = stock_movement_entry['quantity__sum']
                    bills = DetailsBills.objects.filter(createdAt=request.query_params.get("date"), deleted=False,
                                                        details_stock__product_id=dict(con)['id'],
                                                        storage_depots_id=request.query_params.get("id")).exclude(
                        bills=None).aggregate(
                        Sum('quantity_served'))
                    if bills['quantity_served__sum'] is None:
                        sum_bills = 0
                    else:
                        sum_bills = bills['quantity_served__sum']

                    update_prod[
                        'qte_stock'] = sum_stock - sum_stock_movement_exit - sum_bills + sum_stock_movement_entry
                    stock_cmup = DetailsSupplies.objects.filter(product_id=dict(con)['id'],
                                                                createdAt=request.query_params.get("date")).exclude(
                        supplies=None).order_by('-id').first()
                    if stock_cmup is None:
                        update_prod['cmup'] = 0

                    else:
                        update_prod['cmup'] = stock_cmup.cmup
                    prod_update.append(update_prod)
                content = {'content': prod_update}

                return Response(data=content, status=status.HTTP_200_OK)
        # if str(timezone.now().date()) == request.query_params.get("date"):


class Expenses_natureViewSet(viewsets.ModelViewSet):
    queryset = Expenses_nature.objects.filter(deleted=False)
    serializer_class = Expenses_natureSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = Expenses_natureFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Expenses_nature.objects.filter(deleted=False, hospital=user_hospital)
        return Expenses_nature.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        expenses_nature_form = Expenses_natureForm(request.data)
        if expenses_nature_form.is_valid():
            expenses_nature = expenses_nature_form.save()
            expenses_nature.hospital = self.request.user.hospital
            expenses_nature.save()
            serializer = self.get_serializer(expenses_nature, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**expenses_nature_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        expenses_nature = self.get_object()
        expenses_nature_form = Expenses_natureForm(request.data, instance=expenses_nature)
        if expenses_nature_form.is_valid():
            expenses_nature = expenses_nature_form.save()
            expenses_nature.save()
            serializer = self.get_serializer(expenses_nature, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**expenses_nature_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        expenses_nature = self.get_object()
        expenses_nature.deleted = True
        expenses_nature.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_exp(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_depart(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            departments = Departments.objects.filter(hospital=self.request.user.hospital, name__icontains=data['name'])
            if departments:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    # @action(detail=False, methods=['get'], url_path='all')
    # def get_all_expense(self, request, *args, **kwargs):
    #     queryset = self.filter_queryset(self.get_queryset())
    #     expence_update = []
    #     for expence_id in Expenses_natureSerializer(queryset, many=True).data:
    #         get_expence = Cash_movement.objects.filter(expenses_nature_id=dict(expence_id)['id'], deleted=False).all()
    #         update_expence = dict(expence_id)
    #         update_expence['total'] = get_expence.count()
    #         expence_update.append(update_expence)

    #     content = {'content': expence_update}
    #     # patient_id = self.request.query_params.get("patient_id")
    #     # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
    #     # content = {'content': {'solde_patient': get_bills}}
    #     return Response(data=content, status=status.HTTP_200_OK)


class CashViewSet(viewsets.ModelViewSet):
    queryset = Cash.objects.all()
    serializer_class = CashSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = CashFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital=self.request.user.hospital
        if self.request.user.role == 'CASHIER':
            if user_hospital:
                return Cash.objects.filter(user_id=self.request.user.id, deleted=False, hospital=user_hospital).all()
            return Cash.objects.filter(user_id=self.request.user.id, deleted=False).all()
        else:
            if user_hospital:
                return Cash.objects.filter(deleted=False, hospital=user_hospital).all()
            return Cash.objects.filter(deleted=False).all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        user = self.request.user
        cash_form = CashForm(request.data)
        if Cash.objects.filter(user=user,is_active=True, type_cash='CASH_COUNTERS').exists():

            raise Exception("Session déjà ouverte")
        get_cash = Cash.objects.filter(hospital=self.request.user.hospital, user_id=user.id, is_active=False, type_cash='CASH_COUNTERS', deleted=False).last()
        if user.check_password(request.data['password']):
            if cash_form.is_valid():
                cash = cash_form.save()
                cash.hospital = self.request.user.hospital
                cash.user_id = user.id
                if get_cash:
                    cash.balance = int(request.data['cash_fund']) + int(get_cash.balance)
                    cash.cash_fund = int(request.data['cash_fund']) + int(get_cash.balance)
                else:
                    cash.balance = request.data['cash_fund']
                    cash.cash_fund = request.data['cash_fund']
                cash.save()
                serializer = self.get_serializer(cash, many=False)
                return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            errors = {**cash_form.errors}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            errors = {"password": ["Username/Password incorrect."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        cash = self.get_object()
        cash_form = CashForm(request.data, instance=cash)
        if cash_form.is_valid():
            cash = cash_form.save()
            if request.data['is_active'] == False:
                cash.close_date = timezone.now()
                cash.save()
            cash.save()
            serializer = self.get_serializer(cash, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**cash_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        cash = self.get_object()
        cash.deleted = True
        cash.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='current')
    def get_current(self, request, *args, **kwargs):
        user = self.request.user
        get_cash = Cash.objects.filter(hospital = user.hospital, user_id=user.id, is_active=True, type_cash = 'CASH_COUNTERS').last()
        serializer_cash = CashSerializer(get_cash, many=False,  fields=('id', 'code', 'cash_fund', 'is_active', 'user', 'open_date', 'balance'))
        content = {'content': {'cash':serializer_cash.data}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='close/(?P<pk>.+)')
    def close(self, request, *args, **kwargs):
        get_user = User.objects.filter(username=request.data['cashier']).last()
        if get_user.check_password(request.data['password']):
            get_cash = Cash.objects.filter(id=request.data['id'], is_active=True, type_cash='CASH_COUNTERS', hospital=self.request.user.hospital).last()
            get_cash.close_date = timezone.now()
            get_cash.is_active = False
            get_cash.save()
            return Response(status=status.HTTP_200_OK)
        else:
            errors = {"user": ["No permission allowed."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post', 'get'], url_path='days_state')
    def get_days_state(self, request):
        year_month = {}
        today = date.today()

        startdate = get_first_date_of_month(year=int(today.year), month=int(today.month))
        enddate = get_last_date_of_month(year=int(today.year), month=int(today.month))

        bills = DetailsBills.objects.filter(createdAt__range=[startdate, enddate], deleted=False,
                                            cash_id=self.request.query_params.get("id")).exclude(
            ingredient_id=None).values(product=F('details_stock__product_name')).annotate(quantity_stock=F('details_stock__qte_stock'),
            quantity_served=Sum('quantity_served'))
        # stat_category = {'category': 'C.A'}
        # for bill in bills:
        #     stat_category[bill['category']] = bill['turnover']
        get_cash = Cash.objects.filter(id=request.query_params.get("id")).last()
        html_render = get_template('export_days_state.html')
        html_content = html_render.render(
            {'bills': bills, 'cash': get_cash, 'hospital':self.request.user.hospital})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)


    @action(detail=False, methods=['post'], url_path='open/(?P<pk>.+)')
    def open(self, request, *args, **kwargs):

        get_user = User.objects.filter(username=request.data['cashier']).last()
        if get_user.check_password(request.data['password']):
            get_cash = Cash.objects.filter(id=request.data['id'], is_active=False).last()
            get_cash.is_active = True
            get_cash.save()
            return Response(status=status.HTTP_200_OK)

        else:
            errors = {"user": ["No permission allowed."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='sessions_analysis')
    def sessions_analysis(self, request, *args, **kwargs):
        get_cash_movement = Cash_movement.objects.filter(cash__id=self.request.query_params.get("id"),
                                                         cash__is_active=False).all()
        serializer = Cash_movementSerializer(get_cash_movement, many=True)
        content = {'content': serializer.data}
        return Response(data=content, status=status.HTTP_200_OK)

    # @action(detail=False, methods=['get'], url_path='allCashs')
    # def getAllCashs(self, request, *args, **kwargs):
    #     get_cash = User.objects.filter(groups__name='CAISSIER').all()
    #     serializer = UserSerializer(get_cash, many=True)
    #     content = {'content': serializer.data}
    #     return Response(data=content, status=status.HTTP_200_OK)
    #
    @action(detail=False, methods=['get'], url_path='allMyCashs')
    def getAllMyCashs(self, request, *args, **kwargs):
        get_cash = Cash.objects.filter(user_id=self.request.query_params.get("user")).all()
        serializer = CashSerializer(get_cash, many=True)
        content = {'content': serializer.data}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='isOpen')
    def is_open(self, request, *args, **kwargs):
        user = self.request.user
        bills = Bills.objects.filter(cash__user=user.id).all()
        for bill in bills:
            Bills.objects.filter(id=bill.id).update(cash_code=bill.cash.code, patient_name=bill.patient.name, cashier_name=bill.cash.user.username)

        if DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None).all():
            DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None).all().delete()
        get_cash = Cash.objects.filter(hospital=self.request.user.hospital,user_id=user.id, is_active=True, type_cash='CASH_COUNTERS').last()
        get_hospital = self.request.user.hospital
        if get_cash:
            content = {'content': {'is_active': True, 'is_inventory': False}}
            return Response(data=content, status=status.HTTP_200_OK)
        else:

            content = {'content': {'is_active': False, 'is_inventory': False}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='details_sessions_analysis')
    def details_sessions_analysis(self, request, *args, **kwargs):
        # get_cash = Cash.objects.filter(id=request.query_params.get("id")).last()
        # if get_cash:

        #use select_related tojoin related tables in one query
        get_cash_movment = Cash_movement.objects.selected_related('cash').filter(cash_id=request.query_params.get("id")).all()

        serializer = Cash_movementSerializer(get_cash_movment, many=True, fields=('id', 'code', 'cash', 'motive','expenses_nature', 'type','amount_movement', 'createdAt', 'timeAt'))
        get_settlement = PatientSettlement.objects.selected_related('cash').filter(cash_id=request.query_params.get("id"), deleted=False)
        serializer_settle = PatientSettlementSerializer(get_settlement, many=True)
        content = {'content': {'cash': get_cash, 'cash_movement': serializer.data,
                                'settlement': serializer_settle.data}}
        return Response(data=content, status=status.HTTP_200_OK)
        # else:
        #     content = {"content": []}
        #     return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='movements_analysis')
    def movements_analysis(self, request, *args, **kwargs):
        startdate = request.query_params.get("start_date")
        enddate = request.query_params.get("end_date")
        get_cash_movement = Cash_movement.objects.filter(createdAt__range=[startdate, enddate])
        serializer = Cash_movementSerializer(get_cash_movement, many=True, fields=('id', 'code', 'cash', 'motive','expenses_nature', 'type','amount_movement', 'createdAt', 'timeAt'))
        get_settlement = PatientSettlement.objects.filter(createdAt__range=[startdate, enddate], deleted=False)
        serializer_settle = PatientSettlementSerializer(get_settlement, many=True)
        content = {'content': {'cash_movement': serializer.data, 'settlement': serializer_settle.data}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export_movements_analysis')
    def export_movements_analysis(self, request, *args, **kwargs):
        startdate = request.query_params.get("start_date")
        enddate = request.query_params.get("end_date")
        get_cash = Cash.objects.filter(id=request.query_params.get(
            "id")).last()
        get_bills = Bills.objects.filter(cash_id=request.query_params.get(
            "id"), deleted=False).last()
        serializer_bills = BillsSerializer(get_bills, many=True)
        bills = serializer_bills.data
        html_render = get_template('export_item_movement.html')
        html_content = html_render.render(
            {'bills': bills, 'cash_code': get_cash.code, 'cashier': get_cash.user.username, 'hospital':self.request.user.hospital,
             'start_date': startdate, 'end_date': enddate,
             'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)

    @action(detail=False, methods=['get'], url_path='export_bills_analysis')
    def export_bills_analysis(self, request, *args, **kwargs):
        startdate = request.query_params.get("start_date")
        enddate = request.query_params.get("end_date")
        get_cash = Cash.objects.filter(id=request.query_params.get(
            "id")).last()
        get_bills = Bills.objects.filter(cash_id=request.query_params.get(
            "id"), deleted=False).last()
        serializer_bills = BillsSerializer(get_bills, many=True)
        bills = serializer_bills.data
        html_render = get_template('export_item_movement.html')
        html_content = html_render.render(
            {'bills': bills, 'cash_code': get_cash.code, 'cashier': get_cash.user.username, 'hospital': self.request.user.hospital,
             'start_date': startdate, 'end_date': enddate,
             'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_cash(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)


def inventory_update(user, inventory, request):
    get_details_inventory = DetailsInventory.objects.filter(hospital = user.hospital,inventory=None, user_id=user.id).all()
    for inventories in get_details_inventory:
        get_product_stock = Stock.objects.filter(hospital = user.hospital,id=inventories.ingredient_id).last()
        if inventories.quantity_adjusted > get_product_stock.qte_stock:
            result = int(inventories.quantity_adjusted) - int(get_product_stock.qte_stock)
            total_amount = result * get_product_stock.cmup
            DetailsStock_movement.objects.create(hospital = user.hospital,storage_depots_id=inventories.storage_depots_id,
                                                 ingredient_id=inventories.ingredient_id,
                                                 total_amount=total_amount, quantity=result,
                                                 unit_price=get_product_stock.cmup, type_movement='ENTRY',
                                                 user_id=user.id)
            get_product_stock.qte_stock = inventories.quantity_adjusted
            get_product_stock.save()
        else:
            result = int(get_product_stock.qte_stock) - int(inventories.quantity_adjusted)
            total_amount = result * get_product_stock.cmup
            DetailsStock_movement.objects.create(hospital = user.hospital,storage_depots_id=inventories.storage_depots_id,
                                                 ingredient_id=inventories.ingredient_id,
                                                 total_amount=total_amount, quantity=result,
                                                 unit_price=get_product_stock.cmup, type_movement='EXIT',
                                                 user_id=user.id)
            get_product_stock.qte_stock = inventories.quantity_adjusted
            get_product_stock.save()
        inventories.inventory_id = inventory
        inventories.save()
    get_stock_movement = DetailsStock_movement.objects.filter(hospital = user.hospital,type_movement='ENTRY', user_id=user.id,
                                                              stock_movement=None).all()
    if get_stock_movement:
        sum = get_stock_movement.aggregate(Sum('total_amount'))['total_amount__sum']

        save_stock_movement = Stock_movement.objects.create(hospital = user.hospital,storage_depots_id=request.data[
            'storage_depots'],
                                                            type_movement='ENTRY',
                                                            movement_value=sum,
                                                            reason_movement='inventory',
                                                            createdAt=request.data[
                                                                'createdAt'])
        for movement in get_stock_movement:
            movement.stock_movement_id = save_stock_movement.id
            movement.save()
    else:
        get_stock_movement = DetailsStock_movement.objects.filter(hospital = user.hospital,type_movement='EXIT',
                                                                  user_id=user.id, stock_movement=None).all()
        if get_stock_movement:
            sum = get_stock_movement.aggregate(Sum('total_amount'))['total_amount__sum']

            save_stock_movement = Stock_movement.objects.create(hospital = user.hospital,storage_depots_id=request.data[
                'storage_depots'],
                                                                type_movement='EXIT',
                                                                movement_value=sum,
                                                                reason_movement='inventory',
                                                                createdAt=request.data[
                                                                    'createdAt'])
            for movement in get_stock_movement:
                movement.stock_movement_id = save_stock_movement
                movement.save()


class InventoryViewSet(viewsets.ModelViewSet):
    queryset = Inventory.objects.filter(deleted=False)
    serializer_class = InventorySerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = InventoryFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Inventory.objects.filter(deleted=False, hospital=user_hospital)
        return Inventory.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        inventory_form = InventoryForm(request.data)
        if inventory_form.is_valid():
            inventory = inventory_form.save()
            inventory.hospital = self.request.user.hospital
            # inventory_update(user=user, inventory=inventory, request=request)
            inventory.save()
            get_details_inv = DetailsInventory.objects.filter(hospital = self.request.user.hospital,
                storage_depots_id=request.data[
                    'storage_depots'], inventory=None).all()

            for inv in get_details_inv:
                inv.inventory_id = inventory
                inv.save()
                get_details_stock = Stock.objects.filter(hospital = self.request.user.hospital,ingredient_id=inv.stock.ingredient_id,
                                                                storage_depots_id=request.data[
                                                                    'storage_depots']).last()
                get_details_stock.qte_stock = inv.quantity_adjusted
                get_details_stock.save()
            get_hospital = Hospital.objects.filter(id=self.request.user.hospital.id).last()
            get_hospital.is_inventory = False
            get_hospital.save()
            serializer = self.get_serializer(inventory, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**inventory_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        inventory = self.get_object()
        inventory_form = InventoryForm(request.data, instance=inventory)
        if inventory_form.is_valid():
            inventory = inventory_form.save()
            inventory.save()
            serializer = self.get_serializer(inventory, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**inventory_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        inventory = self.get_object()
        inventory.deleted = True
        inventory.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='close')
    def close(self, request, *args, **kwargs):
        user = self.request.user
        get_cash = Cash.objects.filter(hospital = self.request.user.hospital,user_id=user.id, is_active=True).last()
        get_cash.is_active = False
        get_cash.save()
        return Response(status=status.HTTP_200_OK)


class Cash_movementViewSet(viewsets.ModelViewSet):
    queryset = Cash_movement.objects.filter(deleted=False)
    serializer_class = Cash_movementSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = Cash_movementFilter

    def get_queryset(self):
        user_hospital = self.request.user
        if user_hospital.hospital:
            if 'type' in self.request.query_params:
                return Cash_movement.objects.filter(deleted=False, hospital=user_hospital.hospital)
            return Cash_movement.objects.filter(cash__user_id=user_hospital.id,deleted=False, hospital=user_hospital.hospital)
        else:
            if 'type' in self.request.query_params:
                return Cash_movement.objects.filter(deleted=False)
            return Cash_movement.objects.filter(cash__user_id=user_hospital.id,deleted=False)
    
    # filter_backends = (filters.DjangoFilterBackend)

    # def get_queryset(self):
    #     user = self.request.user
    #     return Cash_movement.objects.filter(cash__user_id=user.id, cash__is_active=True, deleted=False).all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        cash_movement_form = Cash_movementForm(request.data)
        if cash_movement_form.is_valid():
            user = self.request.user
            with transaction.atomic():
                if request.data['type'] == 'TRANSFER':
                    get_cash = Cash.objects.filter(id=request.data['cash_destination'], hospital = user.hospital).last()
                    
                    # get_cash_movement = Cash_movement.objects.filter(hospital=user.hospital,type = 'TRANSFER',createdAt = timezone.now().date(), cash_destination_id=get_cash.id).last()
                    # if get_cash_movement and get_cash.type == 'CASH_MAIN':
                    #     errors = {"transfer": ["The daily discount has already been made today."]}
                    #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                    # else:
                    cash_movement = cash_movement_form.save()
                    cash_movement.hospital = user.hospital
                    get_cash_origin = Cash.objects.filter(id=request.data['cash_origin']).last()
                    if get_cash_origin.balance == 0 :
                        errors = {"transfer": ["Montant insuffisant pour la transfert."]}
                        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                    get_cash_dest = Cash.objects.filter(id=request.data['cash_destination']).last()
                    balance_after = int(get_cash_dest.balance) + int(request.data['physical_amount'])
                    balance_after_origin = int(get_cash_origin.balance) - int(request.data['physical_amount'])
                    cash_movement.balance_after = balance_after 
                    cash_movement.balance_before = get_cash_dest.balance
                    get_cash_dest.balance = balance_after
                    get_cash_dest.save()
                    get_cash_origin.balance = balance_after_origin
                    get_cash_origin.save()
                    
                    cash_movement.save() 
                    serializer = self.get_serializer(cash_movement, many=False)
                    return Response(data=serializer.data, status=status.HTTP_201_CREATED)  
                else:
                    get_cash = Cash.objects.filter(user_id=user.id, is_active=True, hospital = user.hospital).last()
                    if get_cash:
                        cash_movement = cash_movement_form.save()
                        cash_movement.hospital = self.request.user.hospital
                        cash_movement.cash_id = get_cash.id
                        
                        if request.data['type'] == 'EXIT':
                            cash_movement.balance_before = get_cash.balance
                            balance = int(get_cash.balance) - int(request.data['amount_movement'])
                            cash_movement.balance_after = balance
                            get_cash.balance = balance
                            get_cash.save()
                        if request.data['type'] == 'ENTER':
                            cash_movement.balance_before = get_cash.balance
                            balance = int(get_cash.balance) + int(request.data['amount_movement'])
                            cash_movement.balance_after = balance
                            get_cash.balance = balance
                            get_cash.save()
                        cash_movement.save()
                        serializer = self.get_serializer(cash_movement, many=False)
                        if 'print' in request.query_params:
                            html_render = get_template('print_cash_movement.html')
                            # logo = settings.MEDIA_ROOT + '/logo.png'
                            html_content = html_render.render(
                                {'products': serializer.data,
                                'hospital': user.hospital,
                                'url': request.build_absolute_uri('/'),
                                'Cashier': get_cash.user.username})
                            result = BytesIO()
                            pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-16")), result,
                                                    link_callback=link_callback)
                            if not pdf.err:                       
                                response = HttpResponse(content_type='application/pdf')
                                filename = 'Mouvement de caisse_' + str(cash_movement.code) + '.pdf'
                                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                                response.write(result.getvalue())
                                return response
                            else:
                                errors = {"pdf": ["Error to generate PDF."]}
                                return Response(data=errors, status=status.HTTP_500)
                        return Response(data=serializer.data, status=status.HTTP_201_CREATED)
                    else:
                        errors = {"cash": ["No Cash open."]}
                        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        errors = {**cash_movement_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        user=self.request.user
        cash_movement = self.get_object()
        cash_movement_form = Cash_movementForm(request.data, instance=cash_movement)
        if cash_movement_form.is_valid():
            cash_movement = cash_movement_form.save()
            cash_movement.hospital = user.hospital
            cash_movement.save()
            serializer = self.get_serializer(cash_movement, many=False)
            if 'print' in request.query_params:
                html_render = get_template('print_cash_movement.html')
                
                # logo = settings.MEDIA_ROOT + '/logo.png'
                html_content = html_render.render(
                    {'products': serializer.data,
                        'hospital': self.request.user.hospital,
                        'url': request.build_absolute_uri('/'),
                        'Cashier': cash_movement.cash.user.username})
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-16")), result,
                                        link_callback=link_callback)
                if not pdf.err:                       
                    response = HttpResponse(content_type='application/pdf')
                    filename = 'Mouvement de caisse_' + str(cash_movement.code) + '.pdf'
                    response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                    response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                    response.write(result.getvalue())
                    return response
                else:
                    errors = {"pdf": ["Error to generate PDF."]}
                    return Response(data=errors, status=status.HTTP_500)

            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**cash_movement_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        cash_movement = self.get_object()
        with transaction.atomic():
            get_cash = Cash.objects.get(cash_id=cash_movement.cash.id, hospital=self.request.user.hospital)
            if cash_movement.type == 'EXIT':
                balance = int(get_cash.balance) + int(cash_movement.amount_movement)
                get_cash.balance = balance
                get_cash.save()
            if cash_movement.type == 'ENTER':
                balance = int(get_cash.balance) - int(cash_movement.amount_movement)
                get_cash.balance = balance
                get_cash.save()
        cash_movement.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'], url_path='specific')
    def specific(self, request, *args, **kwargs):
        queryset = Cash_movement.objects.filter(hospital=self.request.user.hospital, expenses_nature_id=self.request.query_params.get('expenses_nature')).all()
        serializer = self.get_serializer(queryset, many=True,fields=('id', 'code', 'cash', 'motive', 'amount_movement', 'createdAt', 'timeAt'))
        return Response(data={'content':serializer.data}, status=status.HTTP_200_OK)    
    
    @action(detail=False, methods=['get'], url_path='all')
    def get_all_mvt(self, request, *args, **kwargs):
        user = self.request.user
        if 'id' in self.request.query_params:
            get_cash = Cash_movement.objects.filter(hospital=user.hospital,id=self.request.query_params.get('id')).last()
        else:
            get_cash = Cash_movement.objects.filter(hospital=user.hospital).last()
        # if 'type' in self.request.query_params:
        #     queryset = Cash_movement.objects.filter(type__icontains=self.request.query_params.get('type'), cash_id=get_cash.id).all().order_by('-createdAt')
        # else:
        #     queryset = Cash_movement.objects.filter(cash_id=get_cash.id).all().order_by('-createdAt')
        if user.hospital:
            queryset = self.filter_queryset(self.get_queryset()).filter(hospital=user.hospital).all().order_by('-createdAt')
        else:
            queryset = self.filter_queryset(self.get_queryset()).all().order_by('-createdAt')
        # if 'type' in self.request.query_params:
        #     queryset = self.filter_queryset(self.get_queryset()).all().order_by('-createdAt')
        # else:
        #     queryset = self.filter_queryset(self.get_queryset()).filter(cash_id=get_cash.id).all().order_by('-createdAt')
        serializer = self.get_serializer(queryset, many=True)
        if self.request.query_params.get("export") == 'pdf':
            sum_total = queryset.aggregate(Sum('amount_movement'))
            html_render = get_template('export_movement_cash.html')
            html_content = html_render.render(
                {'movements': serializer.data, 'hospital': self.request.user.hospital, 'sum_total': sum_total['amount_movement__sum']})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("utf-16")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            content = {'content': serializer.data}
            return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_depart(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            departments = Cash_movement.objects.filter(hospital=self.request.user.hospital,name__icontains=data['name'])
            if departments:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = CategoryFilter
    filter_backends = (filters.DjangoFilterBackend,)
    
    # def get_queryset(self):
    #     # if self.request.user.hospital:
    #     #     user_hospital = self.request.user.hospital
    #     #     return Category.objects.filter(deleted=False, hospital=user_hospital)
    #     # else:
    #     return Category.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):

        category_form = CategoryForm(request.data)
        if category_form.is_valid():
            obj = category_form.save()
            # obj.hospital = self.request.user.hospital
            obj.save()
            for translate in self.request.data['name_language']:
                get_translate = CategoryTranslation.objects.filter(category_id=obj.id, language=translate['language']).last()
                if get_translate:
                    get_translate.name = translate['name']
                    get_translate.save()
                else:
                    CategoryTranslation.objects.create(user=self.request.user, category_id=obj.id, language=translate['language'], name = translate['name'])
            
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**category_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        category = self.get_object()
        category_form = CategoryForm(request.data, instance=category)
        if category_form.is_valid():
            obj = category_form.save()
            obj.save()
            for translate in self.request.data['name_language']:
                get_translate = CategoryTranslation.objects.filter(category_id=translate['category'], language=translate['language']).last()
                if get_translate:
                    get_translate.name = request.data['name']
                    get_translate.save()
                else:
                    CategoryTranslation.objects.create(user=self.request.user, category_id=obj.id, language=translate['language'], name = translate['name'])
            
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**category_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.deleted = True
        obj.save()
        return Response(status=status.HTTP_204_NO_CONTENT)
 
    
    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                data = {}
                for dbframe in dbframe.itertuples():
                    get_obj = Category.objects.filter(translations__name__icontains=dbframe.Nom_fr, hospital=self.request.user.hospital).last()
                    print(get_obj)
                    if get_obj:
                        pass
                    else:
                        get_obj = Category.objects.filter(translations__name__icontains=dbframe.Nom_en, hospital=self.request.user.hospital).last()
                    
                    langue = [{"name": checkContent(content=dbframe.Nom_fr), "saved": False, "language": "fr"}, {"name": checkContent(content=dbframe.Nom_en), "saved": False, "language": "en"}]

                    if get_obj is None:
                        obj = Category.objects.create(hospital = self.request.user.hospital, name_language=langue)

                        for translate in langue:
                            get_translate = CategoryTranslation.objects.filter(hospital = self.request.user.hospital, category_id=obj.id, language=translate['language']).last()
                            if get_translate:
                                get_translate.name = translate['name']
                                get_translate.save()
                            else:
                                CategoryTranslation.objects.create(hospital = self.request.user.hospital, user=self.request.user, category_id=obj.id, language=translate['language'], name = translate['name'])

                    else:
                        for translate in langue:
                            get_translate = CategoryTranslation.objects.filter(hospital = self.request.user.hospital, category_id=get_obj.id, language=translate['language']).last()
                            if get_translate:
                                get_translate.name = translate['name']
                                get_translate.save()
                            else:
                                CategoryTranslation.objects.create(hospital = self.request.user.hospital, user=self.request.user, category_id=get_obj.id, language=translate['language'], name = translate['name'])


        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_exp(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_category(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            category = Category.objects.filter(hospital=self.request.user.hospital,translations__name__icontains=data['name_language'])
            if category:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'], url_path='translate/exists', permission_classes=[AllowAny])
    def check_translate(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'title' in data:
            obj = CategoryTranslation.objects.filter(name__icontains=data['name'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)



class ModuleViewSet(viewsets.ModelViewSet):
    queryset = Module.objects.filter(deleted=False)
    serializer_class = ModuleSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated)
    pagination_class = CustomPagination
    filterset_class = ModuleFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        if self.request.user.hospital:
            user_hospital = self.request.user.hospital
            return Module.objects.filter(deleted=False, hospital=user_hospital)
        else:
            return Module.objects.filter(deleted=False, hospital=user_hospital)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        module_form = ModuleForm(request.data)
        if module_form.is_valid():
            module = module_form.save()
            module.save()
            for content in request.data['contentType']:
                get_permission = Permission.objects.filter(content_type_id=content).all()
                for permission in get_permission:
                    get_perm = Permission.objects.filter(id=permission.id).last()
                    get_perm.module_id = module.id
                    get_perm.save()
            serializer = self.get_serializer(module, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**module_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        module = self.get_object()
        module_form = ModuleForm(request.data, instance=module)
        if module_form.is_valid():
            module = module_form.save()
            module.save()
            for content in request.data['contentType']:
                get_permission = Permission.objects.filter(content_type_id=content).all()
                for permission in get_permission:
                    get_perm = Permission.objects.filter(id=permission.id).last()
                    get_perm.module_id = module.id
                    get_perm.save()
            serializer = self.get_serializer(module, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**module_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        module = self.get_object()
        module.deleted = True
        module.save()
        get_permission = Permission.objects.filter(module_id=module.id).all()
        for permission in get_permission:
            get_perm = Permission.objects.filter(id=permission.id).last()
            get_perm.module_id = 'null'
            get_perm.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_module(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            module = Module.objects.filter(name__icontains=data['name'])
            if module:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


class ArchiveViewSet(viewsets.ModelViewSet):
    queryset = Archive.objects.filter(deleted=False)
    serializer_class = ArchiveSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated)
    pagination_class = CustomPagination
    filterset_class = ArchiveFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        if self.request.user.hospital:
            user_hospital = self.request.user.hospital
            return Archive.objects.filter(deleted=False, hospital=user_hospital)
        else:
            return Archive.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        archives = archive(date_month=request.data['date_month'], user=self.request.user)
        serializer = self.get_serializer(archives, many=False)
        return Response(data=serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        module = self.get_object()
        module_form = ModuleForm(request.data, instance=module)
        if module_form.is_valid():
            module = module_form.save()
            module.save()
            for content in request.data['contentType']:
                get_permission = Permission.objects.filter(content_type_id=content).all()
                for permission in get_permission:
                    get_perm = Permission.objects.filter(id=permission.id).last()
                    get_perm.module_id = module.id
                    get_perm.save()
            serializer = self.get_serializer(module, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**module_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        delete_archive(obj.year, obj.month, obj.file)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)



class BackupFileViewSet(viewsets.ModelViewSet):
    queryset = BackupFile.objects.filter(deleted=False)
    serializer_class = BackupFileSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated)
    pagination_class = CustomPagination
    filterset_class = BackupFileFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return BackupFile.objects.filter(deleted=False, hospital=user_hospital)
        return BackupFile.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        try:
            call_command('dbbackup')
            # backup_path = os.path.join(settings.DBBACKUP_STORAGE_FOLDER, '*.sql')
            # #
            # # Get a list of all backup files in the directory
            # backup_files = glob.glob(backup_path)

            # # Check if there are any backup files
            # # Get the last modified backup file
            # last_backup_file = max(backup_files, key=os.path.getmtime)
            # shutil.copy(last_backup_file, settings.BACKUP_STORAGE_FOLDER)

            # backup = BackupFile.objects.create(hospital = self.request.user.hospital,user_id=self.request.user.id, file=last_backup_file)
            # file_delete = f"{backup.file}"
            # os.remove(file_delete.replace('\\', '/'))
            backup_all()
            serializer = self.get_serializer(backup, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
                    
        except Exception as e:
            print("Error in creating", e)

    def update(self, request, *args, **kwargs):
        backup = self.get_object()

        file_copy = f"{backup.file}"
        shutil.copy(file_copy.replace('\\', '/').replace('backup_first', 'backup'), settings.DBBACKUP_STORAGE_FOLDER)
        call_command('dbrestore', '--noinput')
        file_delete = f"{backup.file}"

        return Response(status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        delete_archive(obj.year, obj.month, obj.file)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CityViewSet(viewsets.ModelViewSet):
    queryset = City.objects.all()
    serializer_class = CitySerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = CityFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return City.objects.filter(deleted=False, hospital=user_hospital)
        return City.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    
    def create(self, request, *args, **kwargs):
        obj_form = CityForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = CityForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_city(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                data = {}
                for dbframe in dbframe.itertuples():
                    get_obj = City.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.Nom).last()
                    if get_obj:
                        get_obj.name = dbframe.Nom
                        get_obj.save()
                    else:
                        if str(dbframe.region) == 'nan':
                            get_region = Region.objects.filter(hospital=self.request.user.hospital,name__icontains='default').last()
                            if get_region:
                                City.objects.create(hospital = self.request.user.hospital,region_id=get_region.id,name=dbframe.Nom)
                            else:
                                create_region = Region.objects.create(hospital = self.request.user.hospital,name='default')
                                City.objects.create(hospital = self.request.user.hospital,region_id=create_region.id,name=dbframe.Nom)
                        else:

                            get_region = Region.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.region).last()
                            if get_region:
                                City.objects.create(hospital = self.request.user.hospital,region_id=get_region.id,name=dbframe.Nom)
                            else:
                                create_region = Region.objects.create(hospital = self.request.user.hospital,name=dbframe.region)
                                City.objects.create(hospital = self.request.user.hospital,region_id=create_region.id,name=dbframe.Nom)
                            

        return Response(status=status.HTTP_200_OK)


    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_category(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            obj = City.objects.filter(hospital=self.request.user.hospital, name__icontains=data['name'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


class DistrictViewSet(viewsets.ModelViewSet):
    queryset = District.objects.all()
    serializer_class = DistrictSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DistrictFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return District.objects.filter(deleted=False, hospital=user_hospital)
        return District.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    
    def create(self, request, *args, **kwargs):
        obj_form = DistrictForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = DistrictForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                data = {}
                for dbframe in dbframe.itertuples():
                    get_obj = District.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.Nom).last()
                    if get_obj:
                        get_obj.name = dbframe.Nom
                        get_obj.save()
                    else:
                        if str(dbframe.city) == 'nan':
                            get_region = City.objects.filter(hospital=self.request.user.hospital,name__icontains='default').last()
                            if get_region:
                                District.objects.create(hospital = self.request.user.hospital,region_id=get_region.id,name=dbframe.Nom)
                            else:
                                create_region = City.objects.create(hospital = self.request.user.hospital,name='default')
                                District.objects.create(hospital = self.request.user.hospital,region_id=create_region.id,name=dbframe.Nom)
                        else:

                            get_region = City.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.region).last()
                            if get_region:
                                District.objects.create(hospital = self.request.user.hospital,region_id=get_region.id,name=dbframe.Nom)
                            else:
                                create_region = City.objects.create(hospital = self.request.user.hospital,name=dbframe.region)
                                District.objects.create(hospital = self.request.user.hospital,region_id=create_region.id,name=dbframe.Nom)
                            

        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_exp(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_category(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            category = Category.objects.filter(name__icontains=data['name'])
            if category:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = RegionFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Region.objects.filter(deleted=False, hospital=user_hospital)
        return Region.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    
    def create(self, request, *args, **kwargs):
        obj_form = RegionForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = RegionForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='all')
    def get_all_region(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                data = {}
                for dbframe in dbframe.itertuples():
                    get_obj = Region.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.Nom).last()
                    if get_obj:
                        get_obj.name = dbframe.Nom
                        get_obj.save()
                    else:
                        Region.objects.create(hospital = self.request.user.hospital,name=dbframe.Nom)
        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_Region(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            obj = Region.objects.filter(hospital=self.request.user.hospital,name__icontains=data['name'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


class SuppliesViewSet(viewsets.ModelViewSet):
    queryset = Supplies.objects.filter(deleted=False)
    serializer_class = SuppliesSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = SuppliesFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Supplies.objects.filter(deleted=False, hospital=user_hospital)
        return Supplies.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):

        supplies_form = SuppliesForm(request.data)
        if supplies_form.is_valid():
            user = self.request.user
            supplies = Supplies.objects.filter(id=request.data['supplies'], hospital = user.hospital, user=user).last()

            # supplies = supplies_form.save()
            # supplies.hospital = user.hospital
            # supplies.timeAt = time.strftime("%H:%M:%S", time.localtime())
            # supplies.save()
            get_details_supplies = DetailsSupplies.objects.filter(hospital=user.hospital,
                user_id=user.id, supplies=request.data['supplies']
            ).all()
            bulk_list = list()
            for supplie in get_details_supplies:
                # supplie.supplies_id = supplies
                # supplie.save()

                get_stock = Stock.objects.filter(hospital=user.hospital, ingredient_id=supplie.ingredient_id,
                    storage_depots_id=request.data['storage_depots']
                ).last()
                get_ingredient = Ingredient.objects.filter(
                        id=supplie.ingredient_id
                    ).last()
                if get_stock:
                    get_stock.stock_quantity += supplie.quantity
                    if supplie.quantity_two:
                        get_stock.stock_quantity_two += supplie.quantity_two
                    get_stock.save()
                    
                    unit_cost = supplie.arrival_price / supplie.quantity
                    if get_stock.stock_quantity > 0:
                        old_value = Decimal(get_stock.stock_quantity) * Decimal(get_ingredient.price_per_unit)
                        new_value = Decimal(supplie.quantity) * Decimal(unit_cost)

                        total_qty = Decimal(get_stock.stock_quantity) + Decimal(supplie.quantity)
                        price = (old_value + new_value) / total_qty

                        get_ingredient.cmup = price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    else:
                        pass

                else:
                    Stock.objects.create(hospital=user.hospital,ingredient_id=supplie.ingredient_id, quantity = supplie.quantity,quantity_two = supplie.quantity_two, storage_depots_id=request.data['storage_depots'])
                        
                get_ingredient.last_paid_price=supplie.total_amount
                get_ingredient.save()
                supplie.stock_initial = supplie.quantity
                supplie.save()
                MovementStock.objects.create(
                    ingredient=supplie.ingredient,
                    type="ENTRY",
                    quantity=supplie.quantity,
                    source="SUPPLIE",
                    reference_id=request.data['supplies'],
                    hospital = user.hospital
                )
                supplies.supply_amount=request.data['supply_amount']
                supplies.storage_depots_id=request.data['storage_depots']
                supplies.supply_amount=request.data['supply_amount']
                supplies.save()
                
            serializer = self.get_serializer(supplies, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**supplies_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        supplies = self.get_object()
        user = self.request.user
        supplies_form = SuppliesForm(request.data, instance=supplies)
        if supplies_form.is_valid():
            if request.data["is_accounted"] == True:
                pass
                # supplies = supplies_form.save()
                # supplies.save()
                # get_details_supplies = DetailsSupplies.objects.filter(user_id=user.id,
                #                                                       supplies_id=supplies).all()
                # bulk_list = list()
                # for supplie in get_details_supplies:
                #     get_details_stock = DetailsStock.objects.filter(product_id=supplie.product_id,
                #                                                     storage_depots_id=supplie.storage_depots_id).last()
                #     if not get_details_stock:
                #         bulk_list.append(DetailsStock(product_id=supplie.product.id,
                #                                       storage_depots_id=supplie.storage_depots_id,
                #                                       product_name=supplie.product.name,
                #                                       qte_stock=supplie.quantity,
                #                                       cmup=supplie.arrival_price))
                #         get_product = Product.objects.filter(id=supplie.product_id).last()
                #         get_product.purchase_price = supplie.arrival_price
                #         get_product.save()
                #
                #     else:
                #         get_details_stock.qte_stock = get_details_stock.qte_stock + supplie.quantity
                #         get_details_stock.cmup = supplie.cmup
                #         get_details_stock.save()
                #         get_product = Product.objects.filter(id=supplie.product_id).last()
                #         get_product.purchase_price = supplie.arrival_price
                #         get_product.save()
                #
                # DetailsStock.objects.bulk_create(bulk_list)
                # serializer = self.get_serializer(supplies, many=False)
                # return Response(data=serializer.data, status=status.HTTP_200_OK)
            else:
                supplies = supplies_form.save()
                supplies.timeAt = time.strftime("%H:%M:%S", time.localtime())
                supplies.save()
                path = request.path
                end_path = path.rsplit("/", 1)[-1]
                get_details_supplies = DetailsSupplies.objects.filter(
                    hospital=self.request.user.hospital,
                    user_id=user.id, supplies=end_path
                ).all()
                # for supplie in get_details_supplies:
                    # supplie.save()
                    # get_details_stock = DetailsStock.objects.filter(product_id=supplie.product_id,
                    #                                                 storage_depots_id=supplies.storage_depots_id).last()

                    # get_details_stock.qte_stock = get_details_stock.qte_stock + supplie.quantity
                    # get_details_stock.cmup = supplie.cmup
                    # get_details_stock.save()
                    # get_product.purchase_price = supplie.arrival_price
                    # get_product.save()
                serializer = self.get_serializer(supplies, many=False)
                return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**supplies_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        supplies = self.get_object()
        supplies.deleted = True
        supplies.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_category(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            product = Supplies.objects.filter(hospital=self.request.user.hospital,name__icontains=data['name'])
            if product:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['DELETE'], url_path='delete-empty')
    def delete_empty(self, request, pk=None):
        supply = self.get_object()

        has_lines = DetailsSupplies.objects.filter(supplies=supply, hospital=self.request.user.hospital).exists()

        if has_lines:
            return Response(
                {"detail": "Cannot delete, lines exist."},
                status=status.HTTP_400_BAD_REQUEST
            )

        supply.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='create-empty')
    def create_empty(self, request):
        lines = Supplies.objects.annotate(
            line_count=Count('supplies'),
        ).filter(line_count=0, hospital=self.request.user.hospital,user_id=self.request.user.id)
        lines.delete()

        supply = Supplies.objects.create(
            hospital=self.request.user.hospital,
            suppliers=None,
            supply_amount=0,
            user_id = request.user.id
        )
        return Response(data={"id": supply.id}, status=status.HTTP_200_OK)

    # @action(detail=False, methods=['get'], url_path='stock_available', permission_classes=[AllowAny])
    # def stock_available(self, request, *args, **kwargs):
    #     product={}
    #     get_stock_available = Supplies.objects.filter(storage_depot_id=request.query_params.get("id")).last()
    #     supplies = SuppliesSerializer(get_stock_available, many=False)
    #     product['supplie'] = supplies.data
    #     # print(supplies.data)
    #     get_product = Product.objects.filter(id=get_stock_available.id).all()
    #     serializer = ProductSerializer(get_product, many=True)
    #     product['product'] = serializer.data
    #     return Response(data=product, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get', 'post'], url_path='details', permission_classes=[AllowAny])
    def details(self, request, *args, **kwargs):
        if request.method == 'GET':
            get_product_supplies = DetailsSupplies.objects.filter(
                hospital=self.request.user.hospital,
                supplies_id=request.query_params.get("id_supplies")).all()
            details = DetailsSuppliesSerializer(get_product_supplies, many=True)
            return Response(data=details.data, status=status.HTTP_200_OK)
        else:
            get_details = DetailsSupplies.objects.create(hospital = self.request.user.hospital,supplies_id=request.data['id_supplies'],
                                                         product_id=request.data['product']['id'],
                                                         quantity=request.data['quantity'],
                                                         total_amount=request.data['total_amount'],
                                                         arrival_price=request.data['arrival_price'],
                                                         product_code=request.data['product_code'],
                                                         product_name=request.data['product_name'])
            details = DetailsSuppliesSerializer(get_details, many=False)
            return Response(data=details.data, status=status.HTTP_200_OK)

class DetailsBillsViewSet(viewsets.ModelViewSet):
    queryset = DetailsBills.objects.filter(deleted=False)
    serializer_class = DetailsBillsSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DetailsBillsFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user = self.request.user
        
        if user.hospital:
            if self.request.GET.get("bills"):
                
                if self.request.GET.get("bills") == 'undefined' or self.request.GET.get("bills") == 'null':
                    
                    return DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None, deleted=False, hospital=user.hospital).all()
                else:
                    if self.request.GET.get("bills") == "reset":
                        DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None,
                                                            deleted=False, hospital=user.hospital).all().delete()
                        return DetailsBills.objects.none()
                    else:
                        return DetailsBills.objects.filter(bills_id=self.request.GET.get("bills"), deleted=False, hospital=user.hospital).all()
            elif self.request.GET.get("patient"):
                patient_id = self.request.GET.get("patient")
                # Return all details for the given hospitalization
                return DetailsBills.objects.filter(
                    patient_id=patient_id,
                    deleted=False, hospital=user.hospital
                ).all()
            else:
                
                return DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None, deleted=False, hospital=user.hospital).all()
        else:
            if self.request.GET.get("bills"):
                if self.request.GET.get("bills") == 'undefined' or self.request.GET.get("bills") == 'null':
                    
                    return DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None, deleted=False).all()
                else:
                    if self.request.GET.get("bills") == "reset":
                        DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None,
                                                            deleted=False).all().delete()
                        return DetailsBills.objects.none()
                    else:
                        return DetailsBills.objects.filter(bills_id=self.request.GET.get("bills"), deleted=False).all()
            elif self.request.GET.get("patient"):
                patient_id = self.request.GET.get("patient")
                # Return all details for the given hospitalization
                return DetailsBills.objects.filter(
                    patient_id=patient_id,
                    deleted=False).all()
            else:
                return DetailsBills.objects.filter(cash__user_id=user.id, bills_id=None, deleted=False).all()


    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Prix de vente HT – Coût d’achat HT = Marge commerciale"""
        user = self.request.user
        if user.hospital.is_inventory == False:
            detailsBills_form = DetailsBillsForm(data=request.data)
            if detailsBills_form.is_valid():
                detailsBills = detailsBills_form.save()
                detailsBills.bill_type = request.data['bill_type']
                detailsBills.user_id = user.id
                detailsBills.hospital = user.hospital
                detailsBills.storage_depots_id = request.data['storage_depots']
                bills = Bills.objects.filter(hospital = user.hospital, id=request.data['bills']).last()
                bills.storage_depots_id = request.data['storage_depots']
                bills.save()
                
                if request.data['cash']:
                    get_cash = Cash.objects.filter(hospital=user.hospital,id=request.data['cash'], user_id=user.id, is_active=True).last()
                else:
                    get_cash = Cash.objects.filter(hospital=user.hospital, user_id=user.id, is_active=True).last()
                if get_cash:
                    reduction = 0
                    if 'type' in request.data and request.data['type']=='free':
                        result = DetailsBills.objects.filter(
                            hospital=user.hospital,
                            patient=request.data['patient'],
                            dish__is_delivery=True,
                            deleted=False
                        ).aggregate(total_dishes=Sum('quantity_served'))

                        if result['total_dishes'] is not None:
                            total_dishes =  result['total_dishes']
                        else:
                            total_dishes =  0
                        count_dish = total_dishes + int(detailsBills.quantity_served)
                        pun_initial = float(detailsBills.pub)
                        quantity = detailsBills.quantity_served
                        
                        result, cummulative = apply_promotions(detailsBills)
                        if result == True and cummulative == True:
                                
                            if user.hospital.rules_reduction and detailsBills.dish.is_delivery == True and  user.hospital.use_delivery == True:
                                
                                rules = normalize_rules(user.hospital.rules_reduction)

                                future_remise = get_future_remise_notification(count_dish, rules)
                                current_remise = get_applicable_reduction(count_dish, rules)

                                if future_remise["should_notify"]:
                                    reduction = future_remise

                                elif current_remise["should_apply"]:
                                    pun_reduit, total = apply_reduction(
                                        pun_initial=pun_initial,
                                        quantity=quantity,
                                        reduction_value=current_remise['reduction']
                                    )
                                

                                    detailsBills.delivery = current_remise['reduction']
                                    detailsBills.pun = pun_reduit
                                    detailsBills.amount_net = total
                                    reduction = current_remise

                                else:
                                    reduction = 0
                            
                        elif result == True and cummulative == False:
                            detailsBills.cash_id = get_cash.id
                            detailsBills.timeAt = time.strftime("%H:%M:%S", time.localtime())
                            detailsBills.save()
                            
                            recipes = RecipeIngredient.objects.filter(hospital = user.hospital, dish_id = detailsBills.dish.id).all()
                            if recipes:
                                for recipe in recipes:
                                    if recipe.ingredient:
                                        if recipe.ingredient.price_per_unit:
                                            price = recipe.ingredient.price_per_unit
                                        else:
                                            price=0

                                        DetailsBillsIngredient.objects.create(hospital = user.hospital, ingredient_id=recipe.ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)
                                    else:
                                        
                                        price=0
                                        DetailsBillsIngredient.objects.create(hospital = user.hospital, compose_ingredient_id=recipe.compose_ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)

                                total = DetailsBillsIngredient.objects.filter(details_bills_id=detailsBills.id,hospital=user.hospital).aggregate(Sum('total_amount'))
                                
                                detailsBills.cost_production = total['total_amount__sum']
                                detailsBills.margin = float(detailsBills.amount_net) - float(total['total_amount__sum'])
                                detailsBills.save()
                                get_bills = Bills.objects.filter(id = request.data['bills'], hospital=user.hospital).last()
                                get_bills.bill_type = request.data['bill_type']
                                get_bills.cash_id = get_cash.id
                                get_bills.patient_id = request.data['patient']
                                get_bills.district_id = request.data['district']
                                get_bills.save()
                                serializer = self.get_serializer(detailsBills, many=False)
                                content = {
                                    'data': serializer.data,
                                    'reduction': 0
                                }
                                return Response(data=content, status=status.HTTP_201_CREATED)
                        else:
                                 
                            if user.hospital.rules_reduction and detailsBills.dish.is_delivery == True and  user.hospital.use_delivery == True:
                                
                                rules = normalize_rules(user.hospital.rules_reduction)

                                future_remise = get_future_remise_notification(count_dish, rules)
                                current_remise = get_applicable_reduction(count_dish, rules)

                                if future_remise["should_notify"]:
                                    reduction = future_remise

                                elif current_remise["should_apply"]:
                                    pun_reduit, total = apply_reduction(
                                        pun_initial=pun_initial,
                                        quantity=quantity,
                                        reduction_value=current_remise['reduction']
                                    )
                        

                                    detailsBills.delivery = current_remise['reduction']
                                    detailsBills.pun = pun_reduit
                                    detailsBills.amount_net = total
                                    reduction = current_remise

                                else:
                                    reduction = 0
                                

                        detailsBills.cash_id = get_cash.id
                        detailsBills.timeAt = time.strftime("%H:%M:%S", time.localtime())
                        detailsBills.save()
                        
                        recipes = RecipeIngredient.objects.filter(dish_id = detailsBills.dish.id).all()
                        if recipes:
                            for recipe in recipes:
                                if recipe.ingredient:
                                    if recipe.ingredient.price_per_unit:
                                        price = recipe.ingredient.price_per_unit
                                    else:
                                        price=0

                                    DetailsBillsIngredient.objects.create(hospital = user.hospital, ingredient_id=recipe.ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)
                                else:
                                    
                                    price=0
                                    DetailsBillsIngredient.objects.create(hospital = user.hospital, compose_ingredient_id=recipe.compose_ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)

                            total = DetailsBillsIngredient.objects.filter(details_bills_id=detailsBills.id,hospital=user.hospital).aggregate(Sum('total_amount'))
                            
                            detailsBills.cost_production = total['total_amount__sum']
                            detailsBills.margin = float(detailsBills.amount_net) - float(total['total_amount__sum'])
                            detailsBills.save()
                            get_bills = Bills.objects.filter(id = request.data['bills'], hospital=user.hospital).last()
                            get_bills.bill_type = request.data['bill_type']
                            get_bills.cash_id = get_cash.id
                            get_bills.patient_id = request.data['patient']
                            get_bills.save()
                            serializer = self.get_serializer(detailsBills, many=False)
                            content = {
                                'data': serializer.data,
                                'reduction': reduction
                            }
                            return Response(data=content, status=status.HTTP_201_CREATED)
                        else:
                            errors = {"ingredient": ["No ingredient related to the dish."]}
                            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        get_details = DetailsBills.objects.filter(hospital=user.hospital,dish_id=request.data['dish'],
                                                                cash_id=get_cash.id, bills=request.data['bills'],
                                                                deleted=False).last()

                        if get_details:
                            total = DetailsBillsIngredient.objects.filter(details_bills_id=get_details.id,hospital=user.hospital).aggregate(Sum('total_amount'))
                            get_details.quantity_served = request.data['quantity_served']
                            get_details.pub = request.data['pub']
                            get_details.pun = request.data['pun']
                            get_details.cost_production = total['total_amount__sum']
                            get_details.amount_net = request.data['amount_net']
                            get_details.margin = detailsBills.amount_net - total['total_amount__sum']
                            get_details.amount_gross = request.data['amount_gross']
                            get_details.save()
                            
                            return Response(status=status.HTTP_201_CREATED)

                        else:
                            # Total des plats déjà servis (historique)
                            result = DetailsBills.objects.filter(
                                hospital=user.hospital,
                                patient=request.data['patient'],
                                dish__is_delivery=True,
                                deleted=False
                            ).aggregate(total_dishes=Sum('quantity_served'))

                            if result['total_dishes'] is not None:
                                total_dishes =  result['total_dishes']
                            else:
                                total_dishes =  0
                            count_dish = total_dishes + int(detailsBills.quantity_served)
                           
                            pun_initial = float(detailsBills.pub)
                            quantity = detailsBills.quantity_served
                            result, cummulative = apply_promotions(detailsBills)
                            
                            if result == True and cummulative == True:
                                if user.hospital.rules_reduction and detailsBills.dish.is_delivery == True and  user.hospital.use_delivery == True:
                                    
                                    rules = normalize_rules(user.hospital.rules_reduction)

                                    future_remise = get_future_remise_notification(count_dish, rules)
                                    current_remise = get_applicable_reduction(count_dish, rules)

                                    if future_remise["should_notify"]:
                                        reduction = future_remise
                                    elif quantity > 1 and future_remise["should_notify"]:
                                        if future_remise["should_notify"] == 50:
                                            pun_reduit, total = apply_reduction(
                                                pun_initial=pun_initial,
                                                quantity=quantity,
                                                reduction_value=current_remise['reduction']
                                            )
                                            detailsBills.delivery = current_remise['reduction']
                                            detailsBills.pun = pun_reduit
                                            detailsBills.amount_net = total
                                            reduction = current_remise
                                        else:
                                            pass
                                    elif current_remise["should_apply"]:
                                        pun_reduit, total = apply_reduction(
                                            pun_initial=pun_initial,
                                            quantity=quantity,
                                            reduction_value=current_remise['reduction']
                                        )
                                        detailsBills.delivery = current_remise['reduction']
                                        detailsBills.pun = pun_reduit
                                        detailsBills.amount_net = total
                                        reduction = current_remise

                                    else:
                                        reduction = 0
                                        

                                detailsBills.cash_id = get_cash.id
                                detailsBills.timeAt = time.strftime("%H:%M:%S", time.localtime())
                                detailsBills.save()
                                
                                recipes = RecipeIngredient.objects.filter(dish_id = detailsBills.dish.id).all()
                                if recipes:
                                    for recipe in recipes:
                                        if recipe.ingredient:
                                            if recipe.ingredient.price_per_unit:
                                                price = recipe.ingredient.price_per_unit
                                            else:
                                                price=0

                                            DetailsBillsIngredient.objects.create(hospital = user.hospital, ingredient_id=recipe.ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)
                                        else:
                                            
                                            price=0
                                            DetailsBillsIngredient.objects.create(hospital = user.hospital, compose_ingredient_id=recipe.compose_ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)

                                    total = DetailsBillsIngredient.objects.filter(details_bills_id=detailsBills.id,hospital=user.hospital).aggregate(Sum('total_amount'))
                                    
                                    detailsBills.cost_production = total['total_amount__sum']
                                    detailsBills.margin = float(detailsBills.amount_net) - float(total['total_amount__sum'])
                                    detailsBills.save()
                                    get_bills = Bills.objects.filter(id = request.data['bills'], hospital=user.hospital).last()
                                    get_bills.bill_type = request.data['bill_type']
                                    get_bills.cash_id = get_cash.id
                                    get_bills.patient_id = request.data['patient']
                                    get_bills.district_id = request.data['district']
                                    get_bills.save()
                                    serializer = self.get_serializer(detailsBills, many=False)
                                    content = {
                                        'data': serializer.data,
                                        'reduction': reduction
                                    }
                                    return Response(data=content, status=status.HTTP_201_CREATED)
                                else:
                                    errors = {"ingredient": ["No ingredient related to the dish."]}
                                    return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                            elif result == True and cummulative == False:
                                detailsBills.cash_id = get_cash.id
                                detailsBills.timeAt = time.strftime("%H:%M:%S", time.localtime())
                                detailsBills.save()
                                
                                recipes = RecipeIngredient.objects.filter(dish_id = detailsBills.dish.id).all()
                                if recipes:
                                    for recipe in recipes:
                                        if recipe.ingredient:
                                            if recipe.ingredient.price_per_unit:
                                                price = recipe.ingredient.price_per_unit
                                            else:
                                                price=0

                                            DetailsBillsIngredient.objects.create(hospital = user.hospital, ingredient_id=recipe.ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)
                                        else:
                                            
                                            price=0
                                            DetailsBillsIngredient.objects.create(hospital = user.hospital, compose_ingredient_id=recipe.compose_ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)

                                    total = DetailsBillsIngredient.objects.filter(details_bills_id=detailsBills.id,hospital=user.hospital).aggregate(Sum('total_amount'))
                                    
                                    detailsBills.cost_production = total['total_amount__sum']
                                    detailsBills.margin = float(detailsBills.amount_net) - float(total['total_amount__sum'])
                                    detailsBills.save()
                                    get_bills = Bills.objects.filter(id = request.data['bills'], hospital=user.hospital).last()
                                    get_bills.bill_type = request.data['bill_type']
                                    get_bills.cash_id = get_cash.id
                                    get_bills.patient_id = request.data['patient']
                                    get_bills.district_id = request.data['district']
                                    get_bills.save()
                                    serializer = self.get_serializer(detailsBills, many=False)
                                    content = {
                                        'data': serializer.data,
                                        'reduction': 0
                                    }
                                    return Response(data=content, status=status.HTTP_201_CREATED)
                                else:
                                    errors = {"ingredient": ["No ingredient related to the dish."]}
                                    return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                            else:
                                if user.hospital.rules_reduction and detailsBills.dish.is_delivery == True and user.hospital.use_delivery == True:
                                    rules = normalize_rules(user.hospital.rules_reduction)

                                    future_remise = get_future_remise_notification(count_dish, rules)
                                    current_remise = get_applicable_reduction(count_dish, rules)

                                    if future_remise["should_notify"]:
                                        reduction = future_remise
                                    elif quantity > 1 and future_remise["should_notify"]:
                                        pun_reduit, total = apply_reduction(
                                            pun_initial=pun_initial,
                                            quantity=quantity,
                                            reduction_value=current_remise['reduction']
                                        )
                                        detailsBills.delivery = current_remise['reduction']
                                        detailsBills.pun = pun_reduit
                                        detailsBills.amount_net = total
                                        reduction = current_remise

                                    elif current_remise["should_apply"]:
                                        pun_reduit, total = apply_reduction(
                                            pun_initial=pun_initial,
                                            quantity=quantity,
                                            reduction_value=current_remise['reduction']
                                        )
                                        detailsBills.delivery = current_remise['reduction']
                                        detailsBills.pun = pun_reduit
                                        detailsBills.amount_net = total
                                        reduction = current_remise

                                    else:
                                        reduction = 0
                                        

                                detailsBills.cash_id = get_cash.id
                                detailsBills.timeAt = time.strftime("%H:%M:%S", time.localtime())
                                detailsBills.save()
                                
                                recipes = RecipeIngredient.objects.filter(dish_id = detailsBills.dish.id).all()
                                if recipes:
                                    for recipe in recipes:
                                        if recipe.ingredient:
                                            if recipe.ingredient.price_per_unit:
                                                price = recipe.ingredient.price_per_unit
                                            else:
                                                price=0

                                            DetailsBillsIngredient.objects.create(hospital = user.hospital, ingredient_id=recipe.ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)
                                        else:
                                            
                                            price=0
                                            DetailsBillsIngredient.objects.create(hospital = user.hospital, compose_ingredient_id=recipe.compose_ingredient.id, quantity=recipe.quantity,impact_price=price, details_bills_id = detailsBills.id, total_amount=Decimal(recipe.quantity) * price)

                                    total = DetailsBillsIngredient.objects.filter(details_bills_id=detailsBills.id, hospital=user.hospital).aggregate(Sum('total_amount'))
                                    detailsBills.cost_production = total['total_amount__sum']
                                    detailsBills.margin = float(detailsBills.amount_net) - float(total['total_amount__sum'])
                                    detailsBills.save()
                                    get_bills = Bills.objects.filter(id = request.data['bills'], hospital=user.hospital).last()
                                    get_bills.bill_type = request.data['bill_type']
                                    get_bills.cash_id = get_cash.id
                                    get_bills.patient_id = request.data['patient']
                                    get_bills.district_id = request.data['district']
                                    get_bills.save()
                                    serializer = self.get_serializer(detailsBills, many=False)
                                    content = {
                                        'data': serializer.data,
                                        'reduction': reduction
                                    }
                                    return Response(data=content, status=status.HTTP_201_CREATED)
                                else:
                                    errors = {"ingredient": ["No ingredient related to the dish."]}
                                    return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                                
            
                else:
                    errors = {"cash": ["No open checkout."]}
                    return Response(errors, status=status.HTTP_400_BAD_REQUEST)
            errors = {**detailsBills_form.errors}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            errors = {"is_inventory": ["Current inventory."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        user = self.request.user
        path = request.path
        end_path = path.rsplit('/', 1)[-1]
        # get_bills = Bills.objects.filter(id=end_path, deleted=False).last()
        # if get_bills.amount_paid == 0:
            # if request.data['details_stock'] != '':
            #     get_details = DetailsBills.objects.filter(bills_id=end_path,
            #                                               details_stock=request.data['details_stock'],
            #                                               deleted=False).last()
            #     if get_details:
            #         get_details_stock = DetailsStock.objects.filter(id=request.data['details_stock'],
            #                                                         storage_depots_id=request.data[
            #                                                             'storage_depots']).last()
            #         margin = int(request.data['amount_net']) - (int(
            #             get_details_stock.cmup * int(request.data['quantity_served'])))
            #         get_details.quantity_served = get_details.quantity_served + int(request.data['quantity_served'])
            #         get_details.quantity_added = request.data['quantity_served']
            #         get_details.amount_net = get_details.amount_net + int(request.data['amount_net'])
            #         get_details.amount_net = get_details.amount_net + int(request.data['amount_net'])
            #         get_details.amount_gross = get_details.amount_net + int(request.data['amount_gross'])
            #         get_details.margin = get_details.margin + margin
            #         get_details.save()
            #         return Response(status=status.HTTP_201_CREATED)
            #     else:
            #         detailsBills_form = DetailsBillsForm(data=request.data)
            #         if detailsBills_form.is_valid():
            #             user = self.request.user
            #             get_cash = Cash.objects.filter(user_id=user.id, is_active=True).last()
            #             if get_cash:
            #                 get_details_stock = DetailsStock.objects.filter(id=request.data['details_stock'],
            #                                                                 storage_depots_id=request.data[
            #                                                                     'storage_depots']).last()
            #                 margin = int(request.data['amount_net']) - (int(
            #                     get_details_stock.cmup * int(request.data['quantity_served'])))
            #                 
            #                 detailsBills.cash_id = get_cash.id
            #                 detailsBills.bills_id = end_path
            #                 detailsBills.margin = margin
            #                 detailsBills.save()
            #                 return Response(status=status.HTTP_201_CREATED)
            #             else:
            #                 errors = {"cash": ["No open checkout."]}
            #                 return Response(errors, status=status.HTTP_400_BAD_REQUEST)
            #         errors = {**detailsBills_form.errors}
            #         return Response(errors, status=status.HTTP_400_BAD_REQUEST)
            # else:
        get_details = DetailsBills.objects.filter(hospital=user.hospital,id=request.data['id'],
            deleted=False).last()

        # if get_details:
        get_details.quantity_served = request.data['quantity_served']
        get_details.pub = request.data['pub']
        get_details.pun = request.data['pun']
        get_details.delivery = request.data['delivery']
        get_details.amount_net = request.data['amount_net']
        get_details.amount_gross = request.data['amount_gross']
        get_details.save()
        return Response(status=status.HTTP_201_CREATED)
                # else:
        # obj = self.get_object()
        # detailsBills_form = DetailsBillsForm(data=request.data, instance=obj)
        # get_details = DetailsBills.objects.filter(id=request.data['details_stock']).last()
        # if detailsBills_form.is_valid():
        #     
        #     # detailsBills.cash_id = get_cash.id
        #     # detailsBills.bills_id = end_path
        #     detailsBills.save()
        #     return Response(status=status.HTTP_201_CREATED)
            # user = self.request.user
            # get_cash = Cash.objects.filter(user_id=user.id, is_active=True).last()
            # if get_cash:
            #     
            #     detailsBills.cash_id = get_cash.id
            #     detailsBills.bills_id = end_path
            #     detailsBills.save()
            #     return Response(status=status.HTTP_201_CREATED)
            # else:
            #     errors = {"cash": ["No open checkout."]}
            #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        # errors = {**detailsBills_form.errors}
        # return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        # else:
        #     errors = {"errors": ["Not able to update"]}
        #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        user = self.request.user
        get_details = DetailsBills.objects.filter(hospital=user.hospital,id=kwargs['pk'], deleted=False).last()
        if get_details:
            get_details.delete()
            return Response(status=status.HTTP_200_OK)
        else:
            # get_user = User.objects.filter(hospital=user.hospital,=request.data['username'], deleted=False).first()
            # if get_user.check_password(request.data['password']):
            if user.has_perm('hospital.delete_bills'):
                get_details = DetailsBills.objects.filter(hospital=user.hospital,id=kwargs['pk'], deleted=False).last()
                if get_details:
                    get_stock = Ingredient.objects.filter(hospital=user.hospital, id=get_details.dish.id).last()
                    get_stock.stock_quantity = get_stock.stock_quantity + get_details.quantity_served
                    get_stock.save()
                    message= f'Effectue lors de la suppresion du produit {get_stock.product_name} dans la facture {get_details.bills.code}'
                    save_mvt_entry = Supplies.objects.create(hospital = user.hospital, additional_info=message)
                    DetailsSupplies.objects.create(hospital = user.hospital,supplies_id=save_mvt_entry.id, dish_id=get_details.dish.id, quantity=get_details.quantity_served)
                    get_details.delete()
                else:
                    get_details.delete()
                return Response(status=status.HTTP_200_OK)
            else:
                errors = {"permission": ["permission not allowed."]}
                return Response(errors, status=status.HTTP_400_BAD_REQUEST)
            # else:
            #     errors = {"credential": ["Username or password incorrect."]}
            #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)
            

    @action(detail=False, methods=['get'], url_path='favorite-dish')
    def get_favorite(self, request, *args, **kwargs):
        # Activer la langue
        patient_id = request.query_params.get("patient_id")
        if not patient_id:
            return Response({"error": "patient_id is required"}, status=400)

        qs = DetailsBills.objects.filter(patient_id=patient_id)

        result = []
        for item in qs:
            dish_translation = item.dish.translations.first()
            category_translation = item.dish.category.translations.first() if item.dish.category else None

            result.append({
                'dish_id': item.dish.id,
                'name': dish_translation.name if dish_translation else None,
                'price': item.dish.price,
                'category': category_translation.name if category_translation else None,
                'total_qty': item.quantity_served,  # ou utiliser Sum si tu agrèges
            })

        # Si tu veux grouper par plat et sommer les quantités
        from collections import defaultdict
        agg = defaultdict(lambda: {'total_qty': 0})
        for r in result:
            key = r['dish_id']
            if key not in agg:
                agg[key].update(r)
            agg[key]['total_qty'] += r['total_qty']

        final_result = list(agg.values())
        final_result.sort(key=lambda x: x['total_qty'], reverse=True)
        content = {"content": final_result}
        # queryset = self.filter_queryset(self.get_queryset())
        # serializer = self.get_serializer(queryset, many=True, fields=('id', 'patient', 'dish')).data
        # content = {'content': serializer}

        # # patient_id = self.request.query_params.get("patient_id")
        # # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['PATCH'], url_path='reload')
    def reload_bills(self, request, *args, **kwargs):
        get_details = DetailsBillsIngredient.objects.filter(details_bills_id=self.request.query_params.get('details_bills'), is_treated=False)
        get_details_bills = DetailsBills.objects.filter(id=self.request.query_params.get('details_bills')).last()
        
        for details in get_details:
            if details.action == 'ADD':
                get_details_bills.amount_net = get_details_bills.amount_net + details.total_amount
                details.is_treated = True
                details.save()
            else:
                get_details_bills.amount_net = get_details_bills.amount_net - details.total_amount
                details.is_treated = True
                details.save()
            get_details_bills.save()
        return Response(status=status.HTTP_200_OK)

    # @action(detail=False, methods=['get'], url_path='stock_available')
    # def stock_available(self, request):
    #     get_product = DetailsSupplies.objects.filter(supplies__storage_depot=request.query_params.get("id")).order_by(
    #         'product_id').distinct("product_id")
    #     productList = []
    #     for product in get_product:
    #         get_prod = Product.objects.filter(id=product.product_id).last()
    #         productList.append(get_prod)
    #     serializer = DetailsSuppliesSerializer(get_product, many=True)
    #     return Response(data=serializer.data, status=status.HTTP_200_OK)
    #
    # @action(detail=False, methods=['get'], url_path='get_items')
    # def get_items(self, request):
    #     startdate = request.query_params.get("start_date")
    #     enddate = request.query_params.get("end_date")
    #     get_product = DetailsBills.objects.filter(storage_depots_id=request.query_params.get("id"),
    #                                               details_stock__product__id=request.query_params.get("product"),
    #                                               createdAt__range=[startdate,enddate])
    #     serializer = DetailsBillsSerializer(get_product, many=True)
    #     content = {'content': serializer.data}
    #     return Response(data=content, status=status.HTTP_200_OK)

class BillViewSet(viewsets.ModelViewSet):
    queryset = Bills.objects.filter(deleted=False)
    serializer_class = BillsSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = BillsFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Bills.objects.filter(cash__user_id=self.request.user.id, cash__is_active=True,deleted=False, hospital=user_hospital)
        return Bills.objects.filter(cash__user_id=self.request.user.id, cash__is_active=True,deleted=False)

    # def get_queryset(self):
    #     return Bills.objects.filter(cash__user_id=self.request.user.id, cash__is_active=True, deleted=False).all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        
        user = self.request.user
        bills_form = BillsForm(request.data)
        if bills_form.is_valid():
            # bills = bills_form.save()
            # bills.hospital = user.hospital
            
            get_cash = Cash.objects.filter(hospital = user.hospital,user_id=request.data['cashier'], is_active=True).last()
            bills = Bills.objects.filter(hospital = user.hospital, id=request.data['bills']).last()
            if get_cash:
                bills.cash_id = get_cash.id
                bills.cash_code = get_cash.code
                bills.cashier_name = get_cash.user.username
                bills.timeAt = time.strftime("%H:%M:%S", time.localtime())
                get_details_bills = DetailsBills.objects.filter(hospital = user.hospital,cash__user_id=request.data['cashier'],bills=bills, deleted=False).all()
                for bill in get_details_bills:
                    # if 'PREPAID' in request.data['payment_method'] and request.data['amount_prepaid'] > 0:
                    #     get_prepaid_account = PatientAccount.objects.filter(patient=request.data['patient'], type_account='PREPAID').last()
                    #     DetailsPatientAccount.objects.create(type_operation = 'DEBIT', hospital = user.hospital,patient_account_id=get_account.id, amount = int(request.data['net_payable']), balance_before = get_account.balance, balance_after = get_account.balance + int(request.data['net_payable']),unpaid = int(request.data['balance']), unpaid_before = get_account.unpaid, unpaid_after = get_account.unpaid + int(request.data['balance']), user_id = user.id)
                    #     get_prepaid_account.balance=get_prepaid_account.balance - int(request.data['amount_prepaid'])
                    #     get_prepaid_account.unpaid=get_account.unpaid + int(request.data['balance'])
                    #     get_account.save()
                    
                    total_cost = Decimal("0")
                    # ingrédients standards du plat
                    # recette = RecipeIngredient.objects.filter(hospital = user.hospital, dish=bill.dish)
                    # ingrédients retirés
                    # removed_ids = DetailsBillsIngredient.objects.filter(hospital = user.hospital, details_bills_id = bill.id, action="REMOVE").values_list("ingredient_id", flat=True)
                    # ingrédients ajoutés
                    # get_dish_preparation = DishPreparation.objects.filter(dish=bill.dish, deleted=False).last()
                    # if get_dish_preparation:
                    #     get_dish_preparation.quantity -= bill.quantity_served
                    #     get_dish_preparation.save()
                    #     bill.save()
                    # else:
                    added = DetailsBillsIngredient.objects.filter(hospital = user.hospital, details_bills_id = bill.id)
                    # 1. Ingrédients normaux
                    # for item in recette.exclude(ingredient_id__in=removed_ids):
                    #     qty_out = item.quantity * bill.quantity_served
                    #     total_cost += Decimal(qty_out) * item.ingredient.price_per_unit
                    #     destocker(item.ingredient, qty_out, bills.id, hospital = user.hospital)
                    # 2. Ingrédients ajoutés
                    if added:
                        for opt in added:
                            qty_out = Decimal(opt.quantity) * Decimal(bill.quantity_served)
                            if opt.ingredient:
                                if opt.ingredient.price_per_unit:

                                    total_cost += qty_out * Decimal(opt.ingredient.price_per_unit)
                                else:
                                    total_cost += qty_out
                                destocker(user ,opt.ingredient, qty_out, bills.id, request.data['storage_depots'], user.hospital, source="BILL")
                            else:
                                destocker_compose(user, opt.compose_ingredient, qty_out, bills.id,  request.data['storage_depots'], user.hospital, source="BILL")
                        bill.save()
                    else:
                        errors = {"Recipe": ["No recipe for this dishe."]}
                        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                    # bill.margin = price_sell_ligne(bill)
                    
                bills.save()
            
                if request.data['amount_paid'] != '0':
                    if PatientSettlement.objects.filter(hospital = user.hospital,patient_id=request.data['patient'], deleted=False,bills_id=bills.id):
                        pass
                    else:
                        bills.amount_paid=request.data['amount_paid']
                        bills.amount_received=request.data['amount_received']
                        bills.net_payable=request.data['net_payable']
                        bills.bills_amount=request.data['bills_amount']
                        if int(request.data['amount_paid']) < int(request.data['net_payable']):
                            bills.status = 'PARTIAL'
                        else:
                            bills.status = 'PAID'
                        if 'overpayment_action' in request.data:
                            bills.overpayment_action=request.data['overpayment_action']
                            if request.data['overpayment_action'] == 'PREPAID':
                                get_account_prepaid=PatientAccount.objects.filter(patient_id=request.data['patient'], type_account='PREPAID').last()
                                get_account_prepaid.balance += float(request.data['refund'])
                                get_account_prepaid.save()
                                DetailsPatientAccount.objects.create(type_operation = 'CREDIT', hospital = user.hospital,patient_account_id=get_account_prepaid.id,source='BILL',reference_id=bills.id, balance = int(request.data['refund']), balance_before = get_account_prepaid.balance - int(request.data['refund']), balance_after = get_account_prepaid.balance, user_id = user.id)
                                

                        bills.refund=request.data['refund']
                        bills.balance=request.data['balance']
                        bills.ttc=request.data['ttc']
                        bills.tva=request.data['tva']
                        bills.delivery=request.data['delivery']
                        get_account=PatientAccount.objects.filter(patient_id=request.data['patient'], type_account='PRIVATE').last()
                        get_account.unpaid += float(request.data['balance'])
                        get_account.save()
                        
                        if int(request.data['amount_cash']) > 0:
                            bills.amount_cash=request.data['amount_cash']
                            
                            DetailsPatientAccount.objects.create(source='BILL',reference_id=bills.id,type_operation = 'CREDIT',hospital = user.hospital,patient_account_id=get_account.id, balance = int(request.data['amount_cash']), balance_before = get_account.balance, balance_after = get_account.balance + int(request.data['amount_cash']),unpaid = int(request.data['balance']), unpaid_before = get_account.unpaid, unpaid_after = get_account.unpaid + int(request.data['balance']), user_id = user.id)
                            get_account.balance+= int(request.data['amount_cash'])
                            get_account.save()
                            
                        if int(request.data['amount_momo']) > 0 :
                            bills.amount_momo=request.data['amount_momo']
                            
                            DetailsPatientAccount.objects.create(source='BILL',reference_id=bills.id,type_operation = 'CREDIT',hospital = user.hospital,patient_account_id=get_account.id, balance = int(request.data['amount_momo']), balance_before = get_account.balance, balance_after = get_account.balance + int(request.data['amount_momo']),unpaid = int(request.data['balance']), unpaid_before = get_account.unpaid, unpaid_after = get_account.unpaid + int(request.data['balance']), user_id = user.id)
                            get_account.balance+= int(request.data['amount_momo'])
                            get_account.save()
                            
                        if int(request.data['amount_om']) > 0:
                            bills.amount_om=request.data['amount_om']
                            
                            DetailsPatientAccount.objects.create(source='BILL',reference_id=bills.id,type_operation = 'CREDIT',hospital = user.hospital,patient_account_id=get_account.id, balance = int(request.data['amount_om']), balance_before = get_account.balance, balance_after = get_account.balance + int(request.data['amount_om']),unpaid = int(request.data['balance']), unpaid_before = get_account.unpaid, unpaid_after = get_account.unpaid + int(request.data['balance']), user_id = user.id)
                            get_account.balance+= int(request.data['amount_om'])
                            get_account.save()
                            
                        if int(request.data['amount_bank_card']) > 0:
                            bills.amount_bank_card=request.data['amount_bank_card']
                            
                            DetailsPatientAccount.objects.create(source='BILL',reference_id=bills.id,type_operation = 'CREDIT',hospital = user.hospital,patient_account_id=get_account.id, balance = int(request.data['amount_bank_card']), balance_before = get_account.balance, balance_after = get_account.balance + int(request.data['amount_bank_card']),unpaid = int(request.data['balance']), unpaid_before = get_account.unpaid, unpaid_after = get_account.unpaid + int(request.data['balance']), user_id = user.id)
                            get_account.balance+= int(request.data['amount_bank_card'])
                            get_account.save()
                            
                        if int(request.data['amount_prepaid']) > 0:
                            bills.amount_prepaid=request.data['amount_prepaid']
                            get_account=PatientAccount.objects.filter(hospital = user.hospital,patient_id=request.data['patient'], type_account='PREPAID').last()
                            if get_account:
                                DetailsPatientAccount.objects.create(source='BILL',reference_id=bills.id,type_operation = 'DEBIT', hospital = user.hospital,patient_account_id=get_account.id, balance = int(request.data['amount_prepaid']), balance_before = get_account.balance, balance_after = get_account.balance - int(request.data['amount_prepaid']),unpaid = int(request.data['balance']), unpaid_before = get_account.unpaid, unpaid_after = get_account.unpaid + int(request.data['balance']), user_id = user.id)
                                get_account.balance-= int(request.data['amount_prepaid'])
                                get_account.unpaid+= int(request.data['balance'])
                                get_account.save()
                                    # errors = {"amount": ["Not enough money in th prepaid account."]}
                                    # return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                            else:
                                errors = {"Prepaid account": ["The patient do not have an prepaid account."]}
                        bills.save()
                        get_bills = Bills.objects.filter(hospital = user.hospital,patient_id=request.data['patient'], deleted=False).all()
                        dette = get_bills.aggregate(Sum('balance'))['balance__sum']
                        if int(request.data['amount_cash']) > 0:
                            amount_cash=request.data['amount_cash']
                        else:
                            amount_cash = 0
                        if int(request.data['amount_momo']) > 0 :
                            amount_momo=request.data['amount_momo']
                        else:
                            amount_momo = 0
                        if int(request.data['amount_om']) > 0:
                            amount_om=request.data['amount_om']
                        else:
                            amount_om = 0
                        if int(request.data['amount_bank_card']) > 0:
                            amount_bank_card=request.data['amount_bank_card']
                        else:
                            amount_bank_card = 0
                        if int(request.data['amount_prepaid']) > 0:
                            amount_prepaid=request.data['amount_prepaid']
                        else:
                            amount_prepaid = 0
                        PatientSettlement.objects.create(hospital = user.hospital,patient_id=request.data['patient'], deleted=False,current_balance = int(dette)-int(bills.balance),new_balance = dette,
                                                        bills_id=bills.id,
                                                        amount_paid=request.data['amount_paid'],amount_bank_card=amount_bank_card,amount_prepaid=amount_prepaid,amount_om=amount_om,amount_momo=amount_momo,amount_cash=amount_cash,
                                                        amount_received=request.data['amount_received'], cash_id=get_cash.id,
                                                        wordings='')
                    get_cash.balance = get_cash.balance + request.data['amount_paid']
                    get_cash.save()
                    # subprocess.call([r'D:\runserver.bat'])
                if 'print' in request.query_params:

                    get_details_bills_all = DetailsBills.objects.filter(hospital = user.hospital,cash__user_id=request.data['cashier'], bills_id=request.data['bills'], deleted=False).all()
                    get_detail_account = get_prepaid_account_detail(request, reference_id= request.data['bills'],user=user, source='BILL')
                    
                    html_render = get_template('print.html')
                    if get_detail_account:
                        html_content = html_render.render(
                        {'products': get_details_bills_all, 'bills': bills,
                        'hospital': user.hospital,
                        'url': request.build_absolute_uri('/'),
                        'account_prepaid': get_detail_account,
                        'Cashier': get_cash.user.username})
                    else:

                        
                        # logo = settings.MEDIA_ROOT + '/logo.png'
                        html_content = html_render.render(
                            {'products': get_details_bills_all, 'bills': bills,
                            'hospital': user.hospital,
                            'url': request.build_absolute_uri('/'),
                            'Cashier': get_cash.user.username})
                    result = BytesIO()
                    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result,
                                            link_callback=link_callback)
                    if not pdf.err:                       
                        response = HttpResponse(content_type='application/pdf')
                        filename = 'Facture_' + str(bills.code) + '.pdf'
                        response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                        response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                        response.write(result.getvalue())
                        return response
                    else:
                        errors = {"pdf": ["Error to generate PDF."]}
                        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                        
                else:
                    serializer = self.get_serializer(bills, many=False)
                    return Response(data=serializer.data, status=status.HTTP_201_CREATED)

            else:
                errors = {"cash": ["No open checkout."]}
                return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        errors = {**bills_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
    @transaction.atomic
    def update(self, request, *args, **kwargs):
        user=self.request.user
        bills = self.get_object()
        # if 'just_print' in request.query_params:
        get_details_bills = DetailsBills.objects.filter(hospital = user.hospital,cash__user_id=request.data['cashier'],
                                                        bills=bills, deleted=False).all()
        # get_details_bills = DetailsBills.objects.filter(cash__user_id=request.data['cashier'],
        #                                                 bills_id=bills).all()

        get_detail_account = get_prepaid_account_detail(request,reference_id= bills.id,user=user, source='BILL')
        html_render = get_template('print.html')
        if get_detail_account:
            html_content = html_render.render(
            {'products': get_details_bills, 'bills': bills,
            'hospital': user.hospital,
            'url': request.build_absolute_uri('/'),
            'account_prepaid': get_detail_account,
            'Cashier': user.username})
        else:
            
            # logo = settings.MEDIA_ROOT + '/logo.png'
            html_content = html_render.render(
                {'products': get_details_bills, 'bills': bills,
                'hospital': user.hospital,
                'url': request.build_absolute_uri('/'),
                'Cashier': user.username})
        # logo = settings.MEDIA_ROOT + '/logo.png'
        html_content = html_render.render(
            {'products': get_details_bills, 'bills': bills,
                'hospital': user.hospital,
                'url': request.build_absolute_uri('/'),
                'Cashier': user.username})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-16")), result, link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(content_type='application/pdf')
            filename = 'Facture_' + str(bills.code) + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            response.write(result.getvalue())
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)
        # else:
        #     bills_form = BillsForm(request.data, instance=bills)
        #     if bills_form.is_valid():
        #         bills = bills_form.save()
        #         if 'is_accounted' in request.data and request.data['is_accounted'] == True:
        #             pass
        #             # bill = bills_form.save()
        #             # bill.save()
        #             # get_details_bills = DetailsBills.objects.filter(cash__user_id=request.data['cashier'],
        #             #                                                 bills=bills).all()
        #             # for bill in get_details_bills:
        #             #     get_details_stock = DetailsStock.objects.filter(product_id=bill.details_stock.product_id,
        #             #                                                     storage_depots_id=bill.details_stock.storage_depots_id).last()
        #             #     get_details_stock.qte_stock = get_details_stock.qte_stock - bill.quantity_served
        #             #     get_details_stock.save()
        #             # bulk_list = list()
        #             # if request.data['advances'] != '':
        #             #     get_settlement = PatientSettlement.objects.filter(bills_id=bills.id).last()
        #             #     if not get_settlement:
        #             #         get_patient_settlement = PatientSettlement.objects.first()
        #             #         if get_patient_settlement is not None:
        #             #             regex = re.compile(r'[\d]+')
        #             #             find = re.findall(regex, get_patient_settlement.code)
        #             #             code = str('%05d' % (int(find[0]) + 1))
        #             #             code_patient = get_patient_settlement.code[0:3] + code
        #             #             bulk_list.append(
        #             #                 PatientSettlement(code=code_patient, patient_id=request.data['patient'],
        #             #                                   bills_id=bills.id,
        #             #                                   payment_method=request.data['payment_method'],
        #             #                                   payment=request.data['advances'],
        #             #                                   cash_id=get_cash.id, wordings=''))
        #             #         else:
        #             #
        #             #             bulk_list.append(
        #             #                 PatientSettlement(patient_id=request.data['patient'], bills_id=bills.id,
        #             #                                   payment=request.data['advances'],
        #             #                                   payment_method=request.data['payment_method'],
        #             #                                   cash_id=get_cash.id,
        #             #                                   wordings=''))
        #             # PatientSettlement.objects.bulk_create(bulk_list)

        #             # serializer = self.get_serializer(bills, many=False)
        #             # return Response(data=serializer.data, status=status.HTTP_200_OK)
        #         else:
                  
        #             get_cash = Cash.objects.filter(hospital = user.hospital,id=request.data['cash']).last()
        #             if get_cash:
        #                 # bills.cash_id = get_cash.id
        #                 # bills.cash_code = get_cash.code
        #                 # bills.cashier_name = get_cash.user.username
        #                 # bills.timeAt = time.strftime("%H:%M:%S", time.localtime())
        #                 # # get_details_bills = DetailsBills.objects.filter(cash__user_id=request.data['cashier'],
        #                 # #                                             bills=bills).all()
        #                 # # for bill in get_details_bills:
        #                 # #     bill.bills_id = bills
        #                 # #     # bill.save()
        #                 # #     bill.save()
        #                 # bills.save()
                        
        #                 # if request.data['amount_paid'] != '0':
        #                 #     if PatientSettlement.objects.filter(hospital = user.hospital,patient_id=request.data['patient'], deleted=False,
        #                 #                                         bills_id=bills.id):
        #                 #         pass
        #                 #     else:
        #                 #         PatientSettlement.objects.create(hospital = self.request.user.hospital,patient_id=request.data['patient'], deleted=False,
        #                 #                                         bills_id=bills.id,
        #                 #                                         payment_method=request.data['payment_method'],
        #                 #                                         amount_paid=request.data['amount_paid'],
        #                 #                                 amount_received=request.data['amount_received'], cash_id=get_cash.id,
        #                 #                                         wordings='')
        #                     # subprocess.call([r'D:\runserver.bat'])

        #                 # if 'print' in request.query_params:

        #                 get_details_bills_all = DetailsBills.objects.filter(hospital = user.hospital,cash__user_id=request.data['cashier'],
        #                                                                     bills_id=bills, deleted=False).all()
        #                 html_render = get_template('print.html')
        #                 # logo = settings.MEDIA_ROOT + '/logo.png'
        #                 html_content = html_render.render(
        #                     {'products': get_details_bills_all, 'bills': bills,
        #                     'hospital': user.hospital,
        #                     'url': request.build_absolute_uri('/'),
        #                     'Cashier': get_cash.user.username})
        #                 result = BytesIO()
        #                 pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-16")), result,
        #                                         link_callback=link_callback)
        #                 if not pdf.err:
        #                     response = HttpResponse(content_type='application/pdf')
        #                     filename = 'Facture_' + str(bills.code) + '.pdf'
        #                     response['Content-Disposition'] = 'inline; filename="' + filename + '"'
        #                     response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
        #                     response.write(result.getvalue())
        #                     return response
        #                 else:
        #                     errors = {"pdf": ["Error to generate PDF."]}
        #                     return Response(data=errors, status=status.HTTP_500)
                            
                    
                    
        #             else:
        #                 serializer = self.get_serializer(bills, many=False)
        #                 return Response(data=serializer.data, status=status.HTTP_201_CREATED)

        #     errors = {**bills_form.errors}
        #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)
    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        user = self.request.user
        get_user = User.objects.filter(hospital = user.hospital,username=request.data['username'], deleted=False).first()
        if get_user.check_password(request.data['password']):
            # if get_user.has_perm('core.delete_bills'):
            get_details_bills = DetailsBills.objects.filter(hospital = user.hospital,bills_id=kwargs['pk'], deleted=False)
            if get_details_bills:
                message= f'Effectue lors de la suppresion de la facture {get_details_bills.last().bills.code}'
                save_mvt_entry = Supplies.objects.create(hospital = self.request.user.hospital, additional_info=message)
                for details in get_details_bills:
                    get_details_ingredient = DetailsBillsIngredient.objects.filter(hospital = user.hospital, details_bills_id = details.id).all()
                    for ingredient in get_details_ingredient:
                        get_ingredient = Ingredient.objects.filter(hospital = user.hospital, id = ingredient.ingredient.id).last()
                        get_ingredient.stock_quantity += ingredient.quantity
                        get_ingredient.save()
                        DetailsSupplies.objects.create(hospital = self.request.user.hospital, supplies_id=save_mvt_entry.id, ingredient_id=ingredient.id, quantity=ingredient.quantity)
                    details.deleted = True
                    details.deletedAt = timezone.now()
                    details.save()
            else:
                pass
            get_patient_settlement = PatientSettlement.objects.filter(hospital = user.hospital,bills_id=kwargs['pk'], deleted=False).last()
            if get_patient_settlement:
                get_patient_settlement.deleted = True
                get_patient_settlement.deletedAt = timezone.now()
                get_patient_settlement.save()
            else:
                pass
            get_bills = Bills.objects.filter(id=kwargs['pk'], deleted=False).last()
            get_bills.deleted = True
            get_bills.deleted_by = user.username
            get_bills.deletedAt = timezone.now()
            get_bills.save()
            return Response(status=status.HTTP_204_NO_CONTENT)
            # else:
            #     errors = {"permission": ["permission not allowed."]}
            #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            errors = {"credential": ["Username or password incorrect."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='all')
    
    def get_all_bill(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    @action(detail=True, methods=['DELETE'], url_path='delete-empty')
    def delete_empty(self, request, pk=None):
        bills = self.get_object()

        has_lines = DetailsBills.objects.filter(bills=bills, hospital=request.user.hospital).exists()

        if has_lines:
            return Response(
                {"detail": "Cannot delete, lines exist."},
                status=status.HTTP_400_BAD_REQUEST
            )

        bills.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='create-empty')
    def create_empty(self, request):

        lines = Bills.objects.annotate(
            line_count=Count('bills'),
        ).filter(line_count=0, hospital=request.user.hospital)
        lines.delete()
        bills = Bills.objects.create(
            hospital=self.request.user.hospital,
            patient=None,
            total_amount=0
        )
        return Response(data={"id": bills.id}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST'], url_path='verify_permission')
    def verify_permission(self, request, *args, **kwargs):

        get_user = User.objects.filter(username=request.data['username']).last()
        if get_user.check_password(request.data['password']):
            return Response(status=status.HTTP_200_OK)

        else:
            errors = {"user": ["No permission allowed."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['get'], url_path='details_hospitalisation')
    def get_details_hospitalisation(self, request, *args, **kwargs):
        user = self.request.user
                # 1. Optimiser la requête ORM : select_related et prefetch_related
        queryset = self.filter_queryset(self.get_queryset()).exclude(deleted=True)

        # 2. Calculer les sommes en une seule requête
        aggregates = queryset.aggregate(
            sum_ca=Sum('net_payable'),
            sum_unpaid=Sum('balance'),
            sum_paid=Sum('amount_paid')
        )

        # 3. Sérialisation avec fields limités (si vous avez un serializer dynamique)
        serializer = BillsSerializerAnalysis(
            queryset,
            many=True,
            fields=(
                'id', 'code', 'bill_type', 'createdAt', 'timeAt', 'doctor', 'patient',
                'bills_amount', 'patient_name', 'amount_received',
                'cash', 'cash_code', 'cashier_name', 'balance', 'details',
                'amount_gross', 'amount_paid'
            )
        ).data

        # 4. Préparer la réponse
        content = {
            'content': serializer,
            'sum_ca': aggregates['sum_ca'],
            'sum_unpaid': aggregates['sum_unpaid'],
            'sum_paid': aggregates['sum_paid']
        }
        if 'print' in request.query_params:
            html_render = get_template('print.html')
            # logo = settings.MEDIA_ROOT + '/logo.png'
            html_content = html_render.render(
                {'products': get_details_bills_all, 'bills': bills,
                    'hospital': user.hospital,
                    'url': request.build_absolute_uri('/'),
                    'Cashier': user.username})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-16")), result,
                                    link_callback=link_callback)
            if not pdf.err:                       
                response = HttpResponse(content_type='application/pdf')
                filename = 'Facture' + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                response.write(result.getvalue())
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='details_hospitalisation/print')
    def get_details_hospitalisation_print(self, request, *args, **kwargs):
        user=self.request.user
        queryset = self.filter_queryset(self.get_queryset()).exclude(deleted=True)
        serializer = BillsSerializer(queryset, many=True).data

        sum_unpaid = queryset.aggregate(Sum('balance'))
        sum_paid = queryset.aggregate(Sum('amount_paid'))
        sum_total = queryset.aggregate(Sum('net_payable'))
        if 'balance' in self.request.query_params:
            html_render = get_template('export_bills_analysis_unpaid.html')
        else:
            html_render = get_template('export_bills_analysis.html')
        html_content = html_render.render(
            {'bills': serializer, 'hospital': user.hospital, 'sum_unpaid': sum_unpaid['balance__sum'],
             'sum_paid': sum_paid['amount_paid__sum'], 'sum_total': sum_total['net_payable__sum'],
             'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("utf-16")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)


    @action(detail=False, methods=['post'], url_path='name/exists')
    def check_category(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            product = Bills.objects.filter(hospital = self.request.user.hospital,name__icontains=data['name'])
            if product:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get', 'post', 'DELETE'], url_path='cash_movements')
    def cash_movements(self, request, *args, **kwargs):
        user = self.request.user
        # if 'id' in self.request.query_params:
        #     get_cash = Cash.objects.filter(id=self.request.query_params.get('id')).last()
        # else:
        #     get_cash = Cash.objects.filter(user_id=user.id, is_active=True).last()

        # if get_cash:
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        
        if self.request.query_params.get("type")=='settlement':
            if user_hospital:
                if 'id' in self.request.query_params:
                    get_settlement = PatientSettlement.objects.select_related('cash').filter(hospital = user.hospital,cash_id=self.request.query_params.get('id'), deleted=False).exclude(amount_paid=0, deleted=True)
                
                    bills_other = PatientSettlement.objects.filter(hospital=user_hospital, deleted=False, cash_id=self.request.query_params.get('id')).aggregate(
                        total_cash=Sum('amount_cash'),
                        total_bank_card=Sum('amount_bank_card'),
                        total_momo=Sum('amount_momo'),
                        total_om=Sum('amount_om'),
                        total_prepaid=Sum('amount_prepaid'),
                    )
                    
                
                else:

                    get_settlement = PatientSettlement.objects.select_related('cash').filter(hospital = user.hospital,cash__user_id=self.request.user.id, cash__is_active=True, deleted=False).exclude(amount_paid=0, deleted=True)
                    bills_other = PatientSettlement.objects.filter(hospital=user_hospital, deleted=False).aggregate(
                        total_cash=Sum('amount_cash'),
                        total_bank_card=Sum('amount_bank_card'),
                        total_momo=Sum('amount_momo'),
                        total_om=Sum('amount_om'),
                        total_prepaid=Sum('amount_prepaid'),
                    )

            else:
                if 'id' in self.request.query_params:
                    get_settlement = PatientSettlement.objects.select_related('cash').filter(cash_id=self.request.query_params.get('id'), deleted=False).exclude(amount_paid=0, deleted=True)
                    bills_other = PatientSettlement.objects.filter(deleted=False, cash_id=self.request.query_params.get('id')).aggregate(
                        total_cash=Sum('amount_cash'),
                        total_bank_card=Sum('amount_bank_card'),
                        total_momo=Sum('amount_momo'),
                        total_om=Sum('amount_om'),
                        total_prepaid=Sum('amount_prepaid'),
                    )
                else:

                    get_settlement = PatientSettlement.objects.select_related('cash').filter(cash__user_id=self.request.user.id, cash__is_active=True, deleted=False).exclude(amount_paid=0, deleted=True)
                    bills_other = PatientSettlement.objects.filter(deleted=False).aggregate(
                        total_cash=Sum('amount_cash'),
                        total_bank_card=Sum('amount_bank_card'),
                        total_momo=Sum('amount_momo'),
                        total_om=Sum('amount_om'),
                        total_prepaid=Sum('amount_prepaid'),
                    )

            response = {
                        "CASH": bills_other["total_cash"] or 0,
                        "BANK_CARD": bills_other["total_bank_card"] or 0,
                        "MOMO": bills_other["total_momo"] or 0,
                        "OM": bills_other["total_om"] or 0,
                        "PREPAID": bills_other["total_prepaid"] or 0,
                    }
            
            
            serializer_settle = PatientSettlementSerializer(get_settlement, many=True)
            content = {'content': serializer_settle.data, 'statistics': response}
            return Response(data=content, status=status.HTTP_200_OK)
        else:
            if user_hospital:
                if 'id' in self.request.query_params:
                    get_cash_movment = Cash_movement.objects.select_related('cash').filter(hospital = user.hospital,cash__id=self.request.query_params.get('id')).all()
                else:

                    get_cash_movment = Cash_movement.objects.select_related('cash').filter(hospital = user.hospital, cash__user_id=user.id, cash__is_active=True).all()
            else:
                if 'id' in self.request.query_params:
                    get_cash_movment = Cash_movement.objects.select_related('cash').filter(cash__id=self.request.query_params.get('id')).all()
                else:

                    get_cash_movment = Cash_movement.objects.select_related('cash').filter(cash__user_id=user.id, cash__is_active=True).all()
            serializer = Cash_movementSerializer(get_cash_movment, many=True)
            # serializer_cash = CashSerializer(get_cash, many=False)
            content = {'content': serializer.data}
            return Response(data=content, status=status.HTTP_200_OK)
        # else:
        #     content = {'content': []}
        #     return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get', 'post', 'DELETE'], url_path='all_cash_movements')
    def all_cash_movements(self, request, *args, **kwargs):
        user=self.request.user
        get_cash = Cash.objects.filter(hospital = user.hospital, is_active=True, type_cash='CASH_COUNTERS').last()

        # serializer_cash = CashSerializer(Cash.objects.filter(id=self.request.query_params.get('id'), deleted=False), many=False,  fields=('id', 'code', 'cash_fund', 'balance','is_active', 'user', 'open_date', 'balance' ))
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:

            if 'id' in self.request.query_params:
                get_cash_movment = Cash_movement.objects.select_related('cash').filter(hospital = user_hospital,cash_id=self.request.query_params.get('id')).all()
                get_settlement = PatientSettlement.objects.select_related('cash').filter(hospital = user_hospital,cash_id=self.request.query_params.get('id'), deleted=False).exclude(amount_paid=0, deleted=True)
            else:

                get_cash_movment = Cash_movement.objects.select_related('cash').filter(hospital = user_hospital,cash__user_id=user.id, cash__is_active=True).all()
                    
                get_settlement = PatientSettlement.objects.select_related('cash').filter(hospital = user_hospital,cash__user_id=self.request.user.id, cash__is_active=True, deleted=False).exclude(amount_paid=0, deleted=True)
                    
        else:
            if 'id' in self.request.query_params:
                get_cash_movment = Cash_movement.objects.select_related('cash').filter(cash_id=self.request.query_params.get('id')).all()
                get_settlement = PatientSettlement.objects.select_related('cash').filter(cash_id=self.request.query_params.get('id'), deleted=False).exclude(amount_paid=0, deleted=True)
            else:

                get_cash_movment = Cash_movement.objects.select_related('cash').filter(cash__user_id=user.id, cash__is_active=True).all()
                
                get_settlement = PatientSettlement.objects.select_related('cash').filter(cash__user_id=self.request.user.id, cash__is_active=True, deleted=False).exclude(amount_paid=0, deleted=True)
                
        
        serializer_settle = PatientSettlementSerializer(get_settlement, many=True, fields=('id', 'amount_paid'))
        serializer = Cash_movementSerializer(get_cash_movment, many=True, fields=('id', 'amount_movement', 'type', 'cash'))

        if serializer.data:
            cashs=serializer.data
        else:
            cashs=[]
        if serializer_settle:
            settlement=serializer_settle.data
        else:
            settlement=[]
        content = {'content': {'cash':CashSerializer(get_cash, many=False).data, 'cashs':cashs, 'settlement':settlement}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post', 'get'], url_path='statistic_payment_method')
    def get_statistic_method(self, request):
        start_date_month = request.data['start_date_month_payment_method']
        year_month = {}
        startdate = ''
        enddate = ''
        if start_date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_month_payment_method']
            enddate = request.data['end_date_month_payment_method']
        stat_globale_other = {}
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            bills_other = Bills.objects.filter(hospital=user_hospital,createdAt__range=[startdate, enddate], deleted=False).aggregate(
                total_cash=Sum('amount_cash'),
                total_bank_card=Sum('amount_bank_card'),
                total_momo=Sum('amount_momo'),
                total_om=Sum('amount_om'),
                total_prepaid=Sum('amount_prepaid'),
            )
        else:
            bills_other = Bills.objects.filter(createdAt__range=[startdate, enddate], deleted=False).aggregate(
            total_cash=Sum('amount_cash'),
            total_bank_card=Sum('amount_bank_card'),
            total_momo=Sum('amount_momo'),
            total_om=Sum('amount_om'),
            total_prepaid=Sum('amount_prepaid'),
        )
        response = {
            "CASH": bills_other["total_cash"] or 0,
            "BANK_CARD": bills_other["total_bank_card"] or 0,
            "MOMO": bills_other["total_momo"] or 0,
            "OM": bills_other["total_om"] or 0,
            "PREPAID": bills_other["total_prepaid"] or 0,
        }
        finalStat = []
        for k, v in response.items():
            jsonElt = {}
            jsonElt['category'] = k
            jsonElt['turnover'] = v
            finalStat.append(jsonElt)
        if self.request.query_params.get('type') == 'pdf':
            sum_total = sum(map(lambda x: int(x['turnover']), finalStat))
            html_render = get_template('export_stat_payment_method.html')
            html_content = html_render.render(
                {'bills': finalStat, 'hospital': user_hospital, 'sum_total': sum_total, 'start_date': startdate, 'end_date': enddate,
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            content = {'content': finalStat}
            return Response(data=content, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=['post', 'get'], url_path='statistic_payment_channel')
    def get_statistic_payment_channel(self, request):
        start_date_channel = request.data['start_date_channel']
        year_month = {}
        startdate = ''
        enddate = ''
        if start_date_channel is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_channel']
            enddate = request.data['end_date_channel']

        stat_globale_other = {}
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            bills_other = Bills.objects.filter(createdAt__range=[startdate, enddate], hospital=user_hospital, deleted=False).values(category=F('bill_type')).annotate(turnover=Sum('amount_paid'))
        else:

            bills_other = Bills.objects.filter(createdAt__range=[startdate, enddate], deleted=False).values(category=F('bill_type')).annotate(turnover=Sum('amount_paid'))
        for bill in bills_other:
            stat_globale_other[bill['category']] = bill['turnover']

        final_other = stat_globale_other
        finalStat = []
        for k, v in final_other.items():
            jsonElt = {}
            jsonElt['category'] = k
            jsonElt['turnover'] = v
            finalStat.append(jsonElt)
        if self.request.query_params.get('type') == 'pdf':
            sum_total = sum(map(lambda x: int(x['turnover']), finalStat))
            html_render = get_template('export_stat_payment_method.html')
            html_content = html_render.render(
                {'bills': finalStat, 'hospital': user_hospital, 'sum_total': sum_total,
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            content = {'content': finalStat}
            return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post', 'get'], url_path='statistic')
    def get_statistic(self, request):
        start_date_month = request.data['start_date_month']
        year_month = {}
        startdate = ''
        enddate = ''
        if start_date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            
            startdate = request.data['start_date_month']
            enddate = request.data['end_date_month']

        # Filtrage de base
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:

            data = DetailsBills.objects.filter(
            deleted=False,
            dish__isnull=False,
            createdAt__range=[startdate, enddate],
            hospital=user_hospital,
                ).values(
                    'dish__category__id',
                    'dish__category__name_language'
                ).annotate(
                    total_quantity=Sum('quantity_served'),
                    total_amount=Sum('amount_net'),
                    total_lines=Count('id')
                ).order_by('-total_amount')
        else:
            data = DetailsBills.objects.filter(
            deleted=False,
            dish__isnull=False,
            createdAt__range=[startdate, enddate],
            ).values(
                'dish__category__id',
                'dish__category__name_language'
            ).annotate(
                total_quantity=Sum('quantity_served'),
                total_amount=Sum('amount_net'),
                total_lines=Count('id')
            ).order_by('-total_amount')
        if self.request.query_params.get('type') == 'pdf':
            sum_total = sum(map(lambda x: int(x['turnover']), data))
            html_render = get_template('export_stat_domaine.html')
            html_content = html_render.render(
                {'bills': data, 'hospital': user_hospital, 'sum_total': sum_total, 'start_date': startdate, 'end_date': enddate,
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            content = {'content': data}
            return Response(data=content, status=status.HTTP_200_OK)
 
    @action(detail=False, methods=['post', 'get'], url_path='statistic_product')
    def get_statistic_product(self, request):
        start_date_month = request.data['start_date_month']
        year_month = {}
        startdate = ''
        enddate = ''
        if start_date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_month']
            enddate = request.data['end_date_month']

        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            bills = DetailsBills.objects.filter(createdAt__range=[startdate, enddate], deleted=False).exclude(ingredient_id=None).values(category=F('details_stock__product__category__name')).annotate(turnover=Sum('amount_net'))
        else:
        
            bills = DetailsBills.objects.filter(createdAt__range=[startdate, enddate], deleted=False).exclude(ingredient_id=None).values(category=F('details_stock__product__category__name')).annotate(turnover=Sum('amount_net'))
        if self.request.query_params.get('type') == 'pdf':
            sum_total = bills.aggregate(Sum('turnover'))
            html_render = get_template('export_stat_domaine.html')
            html_content = html_render.render(
                {'bills': bills, 'hospital': user_hospital, 'sum_total': sum_total['turnover__sum'], 'start_date': startdate,
                 'end_date': enddate,
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            # stat_category = {'category': 'C.A'}
            # for bill in bills:
            #     stat_category[bill['category']] = bill['turnover']
            # content = {'content': {'stat_category': stat_category}}
            # labels = list()
            # values = list()
            # for bill in bills:
            #     labels.append(bill['category'])
            #     values.append(bill['turnover'])
            content = {'content': bills}
            return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post', 'get'], url_path='statistic_product_days')
    def get_statistic_product_days(self, request):
        date_month = request.data['date_month']
        year_month = {}
        startdate = ''
        enddate = ''
        if date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = get_first_date_of_month(year=int(date_month.split("-")[0]), month=int(date_month.split("-")[1]))
            enddate = get_last_date_of_month(year=int(date_month.split("-")[0]), month=int(date_month.split("-")[1]))
        
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            
            bills = Bills.objects.filter(hospital=user_hospital,createdAt__range=[startdate, enddate], deleted=False, bill_type=self.request.query_params.get('type')).values(days=F('createdAt')).annotate(turnover=Sum('amount_paid'))
        else:
            
            bills = Bills.objects.filter(createdAt__range=[startdate, enddate], deleted=False, bill_type=self.request.query_params.get('type')).values(days=F('createdAt')).annotate(turnover=Sum('amount_paid'))



        if self.request.query_params.get('type') == 'pdf':
            sum_total = bills.order_by('days').aggregate(Sum('turnover'))
            html_render = get_template('export_stat_days.html')
            html_content = html_render.render(
                {'bills': bills.order_by('days'), 'hospital': user_hospital, 'sum_total': sum_total['turnover__sum'],
                 'month': date_month.split("-")[1], 'year': date_month.split("-")[0],
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            # stat_category = {'category': 'C.A'}
            # for bill in bills:
            #     stat_category[bill['category']] = bill['turnover']
            # content = {'content': {'stat_category': stat_category}}
            # labels = list()
            # values = list()
            # for bill in bills:
            #     labels.append(bill['category'])
            #     values.append(bill['turnover'])
            content = {'content': bills.order_by('days')}
            return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post', 'get'], url_path='bills_archive')
    def get_bills_archive(self, request):
        archive(date_month=request.data['date_month'])
        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['post', 'get'], url_path='statistic_days')
    def get_statistic_days(self, request):
        date_month = request.data['start_date_month_first']
        year_month = {}
        startdate = ''
        enddate = ''
        if date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_month_first']
            enddate = request.data['end_date_month_first']

        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            bills = Bills.objects.filter(hospital=user_hospital,createdAt__range=[startdate, enddate], deleted=False).values(days=F('createdAt')).annotate(turnover=Sum('amount_paid'))
        else:
            bills = Bills.objects.filter(createdAt__range=[startdate, enddate], deleted=False).values(days=F('createdAt')).annotate(turnover=Sum('amount_paid'))

        if self.request.query_params.get('type') == 'pdf':
            sum_total = bills.order_by('days').aggregate(Sum('turnover'))
            html_render = get_template('export_stat_days.html')
            html_content = html_render.render(
                {'bills': bills.order_by('days'), 'hospital': user_hospital, 'sum_total': sum_total['turnover__sum'],
                 'month': date_month.split("-")[1], 'year': date_month.split("-")[0],
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:

            # stat_category = {'category': 'C.A'}
            # for bill in bills:
            #     stat_category[bill['category']] = bill['turnover']
            # content = {'content': {'stat_category': stat_category}}
            # labels = list()
            # values = list()
            # for bill in bills:
            #     labels.append(bill['category'])
            #     values.append(bill['turnover'])
            content = {'content': bills.order_by('days')}
            return Response(data=content, status=status.HTTP_200_OK)
 
    @action(detail=False, methods=['post', 'get'], url_path='statistic_days_entry')
    def get_statistic_days_entry(self, request):
        date_month = request.data['start_date_month_entry']
        year_month = {}
        startdate = ''
        enddate = ''
        if date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_month_entry']
            enddate = request.data['end_date_month_entry']

        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            bills = Cash_movement.objects.filter(hospital=user_hospital,type='ENTRY',createdAt__range=[startdate, enddate], deleted=False).values(days=F('createdAt')).annotate(turnover=Sum('amount_movement'))
        else:
            bills = Cash_movement.objects.filter(type='ENTRY',createdAt__range=[startdate, enddate], deleted=False).values(days=F('createdAt')).annotate(turnover=Sum('amount_movement'))

        if self.request.query_params.get('type') == 'pdf':
            sum_total = bills.order_by('days').aggregate(Sum('turnover'))
            html_render = get_template('export_stat_days.html')
            html_content = html_render.render(
                {'bills': bills.order_by('days'), 'hospital': user_hospital, 'sum_total': sum_total['turnover__sum'],
                 'month': date_month.split("-")[1], 'year': date_month.split("-")[0],
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:

            # stat_category = {'category': 'C.A'}
            # for bill in bills:
            #     stat_category[bill['category']] = bill['turnover']
            # content = {'content': {'stat_category': stat_category}}
            # labels = list()
            # values = list()
            # for bill in bills:
            #     labels.append(bill['category'])
            #     values.append(bill['turnover'])
            content = {'content': bills.order_by('days')}
            return Response(data=content, status=status.HTTP_200_OK)
    @action(detail=False, methods=['post', 'get'], url_path='statistic_days_exit')
    def get_statistic_days_exit(self, request):
        date_month = request.data['start_date_month_exit']
        year_month = {}
        startdate = ''
        enddate = ''
        if date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_month_exit']
            enddate = request.data['end_date_month_exit']


        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            bills = Cash_movement.objects.filter(hospital=user_hospital,type='EXIT',createdAt__range=[startdate, enddate], deleted=False).values(days=F('createdAt')).annotate(turnover=Sum('amount_movement'))
        else:
            bills = Cash_movement.objects.filter(type='EXIT',createdAt__range=[startdate, enddate], deleted=False).values(days=F('createdAt')).annotate(turnover=Sum('amount_movement'))

        if self.request.query_params.get('type') == 'pdf':
            sum_total = bills.order_by('days').aggregate(Sum('turnover'))
            html_render = get_template('export_stat_days.html')
            html_content = html_render.render(
                {'bills': bills.order_by('days'), 'hospital': user_hospital, 'sum_total': sum_total['turnover__sum'],
                 'month': date_month.split("-")[1], 'year': date_month.split("-")[0],
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:

            # stat_category = {'category': 'C.A'}
            # for bill in bills:
            #     stat_category[bill['category']] = bill['turnover']
            # content = {'content': {'stat_category': stat_category}}
            # labels = list()
            # values = list()
            # for bill in bills:
            #     labels.append(bill['category'])
            #     values.append(bill['turnover'])
            content = {'content': bills.order_by('days')}
            return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post', 'get'], url_path='stat_best_selling_products')
    def get_stat_best_selling_products(self, request):
        start_date_best_selling = request.data['start_date_best_selling']
        year_month = {}
        startdate = ''
        enddate = ''
        if start_date_best_selling is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = request.data['start_date_best_selling']
            enddate = request.data['end_date_best_selling']
        if 'hospital' in self.request.query_params:
            hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            hospital = self.request.user.hospital
        if hospital:

            bills = DetailsBills.objects.filter(
                hospital=hospital,
                createdAt__range=[startdate, enddate],
                deleted=False
            ).exclude(dish_id=None).values(
                'dish__id',
                'dish__name_language',
                'dish__category__name_language'
            ).annotate(
                quantity_served=Sum('quantity_served'),
                turnover=Sum('amount_net'),
                cost_production=Sum('cost_production'),
                margin=Sum('margin'),
            )

        else:

            bills = DetailsBills.objects.filter(
                createdAt__range=[startdate, enddate],
                deleted=False
            ).exclude(dish_id=None).values(
                'dish__id',
                'dish__name_language',
                'dish__category__name_language'
            ).annotate(
                quantity_served=Sum('quantity_served'),
                turnover=Sum('amount_net'),
                cost_production=Sum('cost_production'),
                margin=Sum('margin'),
            )
            # total_cost = cost_data['unit_cost'] * bill['quantity_served']

            # bill['cost_production'] = total_cost
            # bill['margin'] = bill['turnover'] - total_cost
            # print(cost_data,bill['cost_production'], bill['margin'])
                
        
            # get_product = Product.objects.filter(id=request.data['product_available']).last()
            # serializer = ProductSerializer(get_product, many=False)
            # prod = [serializer.data]
        # test = DetailsBills.objects.filter(createdAt__range=[startdate, enddate]).exclude(
        #     ingredient_id=None).all()

        # t_turnover=cotation.aggregate(Sum('turnover'))['turnover__sum']
        # t_gross_margin=cotation.aggregate(Sum('gross_margin'))['gross_margin__sum']
        # for product in prod:
        #     update_prod = dict(product)
        #     get_details_bill = DetailsBills.objects.filter(details_stock__product_id=dict(product)['id'],
        #                                                    createdAt__range=[startdate, enddate]).exclude(
        #         bills=None)
        #
        #     if get_details_bill.aggregate(
        #             Sum('amount_net'))['amount_net__sum'] is None or get_details_bill.aggregate(
        #         Sum('margin'))['margin__sum'] is None:
        #         pass
        #     else:
        #         update_prod['gross_margin'] = get_details_bill.aggregate(
        #             Sum('margin'))['margin__sum']
        #         update_prod['quantity_served'] = get_details_bill.aggregate(
        #             Sum('quantity_served'))['quantity_served__sum']
        #         update_prod['turnover'] = get_details_bill.aggregate(
        #             Sum('amount_net'))['amount_net__sum']
        #         prod_update.append(update_prod)
        content = {'content': {'product': bills, 'turnover': bills.aggregate(Sum('turnover'))['turnover__sum'],
                               'gross_margin': bills.aggregate(Sum('margin'))['margin__sum']}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post', 'get'], url_path='stat_act_patient')
    def stat_act_patient(self, request):
        start_date_act = request.data['start_date']
        year_month = {}
        startdate = ''
        enddate = ''
        if start_date_act is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = datetime.strptime(request.data['start_date'], "%Y-%m-%d").date()
            enddate = datetime.strptime(request.data['end_date'], "%Y-%m-%d").date()

        # print(datetime.strptime(startdate, "%Y-%m-%d").date().year,enddate)
        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            get_details_bill = Bills.objects.filter(hospital = user_hospital,patient_id=request.data['patient'], deleted=False,
                                                    createdAt__range=[startdate, enddate]).all()
        else:
        
            get_details_bill = Bills.objects.filter(patient_id=request.data['patient'], deleted=False,
                                                    createdAt__range=[startdate, enddate]).all()

        serializer = BillsSerializer(get_details_bill, many=True)
        # act_update.append(serializer.data[0])
        # get_bill = Bills.objects.filter(patient_id=request.data['patient'], bill_type='medical_act', createdAt__range=[startdate,enddate]).all()
        # print(get_bill)
        # serializer = BillsSerializer(get_bill, many=True)

        content = {'content': serializer.data}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get','post'], url_path='stat_day_month')
    def stat_day_month(self, request, *args, **kwargs):
        # For printing results

        start_date_month = request.data['start_date']
        year_month = {}
        startdate = ''
        enddate = ''
        today = date.today()
        if start_date_month is None:
            
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            
            startdate = request.data['start_date']
            enddate = request.data['end_date']

        last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
        previous_day = timezone.now() - timedelta(days=1)
        start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
        if 'hospital' in self.request.query_params:
            hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            hospital = self.request.user.hospital

        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today, datetime.max.time())

        prev_start = datetime.combine(previous_day, datetime.min.time())
        prev_end = datetime.combine(previous_day, datetime.max.time())
        if hospital:


            bills_today = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[today_start, today_end],
                deleted=False
            )
            bills_unpaid_today = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[today_start, today_end],
                deleted=False
            ).aggregate(Sum('balance')
            )['balance__sum']
            bills_prev = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[prev_start, prev_end],
                deleted=False
            )

            bills_month = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[startdate, enddate],
                deleted=False
            ).aggregate(Sum('amount_paid')
            )['amount_paid__sum']

            bills_unpaid_month = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[startdate, enddate],
                deleted=False
            ).aggregate(Sum('balance')
            )['balance__sum']

            bills_prev_month = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month],
                deleted=False
            ).aggregate(Sum('amount_paid')
            )['amount_paid__sum']

            bills_prev_unpaid_month = Bills.objects.filter(
                hospital=hospital,
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month],
                deleted=False
            ).aggregate(Sum('balance')
            )['balance__sum']


            settlements_today = PatientSettlement.objects.filter(
                hospital=hospital,
                createdAt__range=[today_start, today_end],
                deleted=False
            )
            settlements_previous = PatientSettlement.objects.filter(
                            hospital=hospital,
                            createdAt__range=[prev_start, prev_end],
                            deleted=False
                        )

            cash_today = Cash_movement.objects.filter(
                hospital=hospital,
                createdAt__range=[today_start, today_end],
                deleted=False
            )
            cash_previous_today = Cash_movement.objects.filter(
                hospital=hospital,
                createdAt__range=[prev_start, prev_end],
                deleted=False
            )
            sales_today = bills_today.aggregate(Sum('net_payable')
            )['net_payable__sum']
            sales_prev = bills_prev.aggregate(Sum('net_payable'),
            )['net_payable__sum']
            settlement_today = settlements_today.aggregate(Sum('amount_paid')
            )['amount_paid__sum']
            settlement_previous = settlements_previous.aggregate(Sum('amount_paid')
            )['amount_paid__sum']

            cash_types_today = cash_today.values(
                types=F('type')
            ).annotate(total=Sum('amount_movement'))

            cash_types_previous = cash_previous_today.values(
                types=F('type')
            ).annotate(total=Sum('amount_movement'))

            get_sum_month= Cash_movement.objects.filter(hospital=hospital,createdAt__range=[startdate, enddate], deleted=False).values(types=F('type')).annotate(
                total=Sum('amount_movement'))
            get_sum_previous_month= Cash_movement.objects.filter(hospital=hospital,createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).values(types=F('type')).annotate(
                total=Sum('amount_movement'))
            
            get_cash = Cash.objects.filter(hospital=hospital,is_active=True, open_date__year=today.year, open_date__month=today.month,
                                        open_date__day=today.day).last()
            if get_cash is not None:
                cash_balance = get_cash.balance
            else:
                cash_balance = 0

            get_cash_previous = Cash.objects.filter(hospital=hospital,is_active=True, open_date__year=previous_day.year,
                                                    open_date__month=previous_day.month,
                                                    open_date__day=previous_day.day).last()
            if get_cash_previous is not None:
                previous_cash_balance = get_cash_previous.balance
            else:
                previous_cash_balance = 0


            get_sum_bills_month_margin = DetailsBills.objects.filter(hospital=hospital,createdAt__range=[startdate, enddate],
                                                                    deleted=False).exclude(
                bills=None).aggregate(
                Sum('margin'))            
            get_sum_bills_month_cost_production = DetailsBills.objects.filter(hospital=hospital,createdAt__range=[startdate, enddate],
                                                                        deleted=False).exclude(
                    bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            get_sum_bills_cost_production = DetailsBills.objects.filter(hospital=hospital,createdAt__range=[today_start, today_end],
                                                                        deleted=False).exclude(
                    bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            month_margin = get_sum_bills_month_margin['margin__sum']        
            get_sum_bills_month_tva = Bills.objects.filter(hospital=hospital,createdAt__range=[startdate, enddate],
                                                                    deleted=False).aggregate(
                Sum('tva'))
            month_tva = get_sum_bills_month_tva['tva__sum']

            get_sum_bills_previous_month_margin = DetailsBills.objects.filter(hospital=hospital,
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).exclude(
                bills=None).aggregate(
                Sum('margin'))           
            get_sum_bills_previous_month_cost_production = DetailsBills.objects.filter(hospital=hospital,
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).exclude(
                bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            
            get_sum_bills_previous_cost_production = DetailsBills.objects.filter(hospital=hospital,
                createdAt__range=[prev_start, prev_end], deleted=False).exclude(
                bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            previous_month_margin = get_sum_bills_previous_month_margin['margin__sum'] 

            get_sum_bills_previous_month_tva = Bills.objects.filter(hospital=hospital,
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).aggregate(
                Sum('tva'))
            previous_month_tva = get_sum_bills_previous_month_tva['tva__sum']
            
        else:

            bills_today = Bills.objects.filter(
                createdAt__range=[today_start, today_end],
                deleted=False
            )
            bills_unpaid_today = Bills.objects.filter(
                createdAt__range=[today_start, today_end],
                deleted=False
            ).aggregate(Sum('balance')
            )['balance__sum']
            bills_prev = Bills.objects.filter(
                createdAt__range=[prev_start, prev_end],
                deleted=False
            )

            bills_month = Bills.objects.filter(
                createdAt__range=[startdate, enddate],
                deleted=False
            ).aggregate(Sum('amount_paid')
            )['amount_paid__sum']

            bills_unpaid_month = Bills.objects.filter(
                createdAt__range=[startdate, enddate],
                deleted=False
            ).aggregate(Sum('balance')
            )['balance__sum']

            bills_prev_month = Bills.objects.filter(
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month],
                deleted=False
            ).aggregate(Sum('amount_paid')
            )['amount_paid__sum']

            bills_prev_unpaid_month = Bills.objects.filter(
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month],
                deleted=False
            ).aggregate(Sum('balance')
            )['balance__sum']


            settlements_today = PatientSettlement.objects.filter(
                createdAt__range=[today_start, today_end],
                deleted=False
            )
            settlements_previous = PatientSettlement.objects.filter(
                            createdAt__range=[prev_start, prev_end],
                            deleted=False
                        )

            cash_today = Cash_movement.objects.filter(
                createdAt__range=[today_start, today_end],
                deleted=False
            )
            cash_previous_today = Cash_movement.objects.filter(
                createdAt__range=[prev_start, prev_end],
                deleted=False
            )
            sales_today = bills_today.aggregate(Sum('net_payable')
            )['net_payable__sum']
            sales_prev = bills_prev.aggregate(Sum('net_payable'),
            )['net_payable__sum']
            settlement_today = settlements_today.aggregate(Sum('amount_paid')
            )['amount_paid__sum']
            settlement_previous = settlements_previous.aggregate(Sum('amount_paid')
            )['amount_paid__sum']

            cash_types_today = cash_today.values(
                types=F('type')
            ).annotate(total=Sum('amount_movement'))

            cash_types_previous = cash_previous_today.values(
                types=F('type')
            ).annotate(total=Sum('amount_movement'))

            get_sum_month= Cash_movement.objects.filter(createdAt__range=[startdate, enddate], deleted=False).values(types=F('type')).annotate(
                total=Sum('amount_movement'))
            get_sum_previous_month= Cash_movement.objects.filter(createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).values(types=F('type')).annotate(
                total=Sum('amount_movement'))
            
            get_cash = Cash.objects.filter(is_active=True, open_date__year=today.year, open_date__month=today.month,
                                        open_date__day=today.day).last()
            if get_cash is not None:
                cash_balance = get_cash.balance
            else:
                cash_balance = 0

            get_cash_previous = Cash.objects.filter(is_active=True, open_date__year=previous_day.year,
                                                    open_date__month=previous_day.month,
                                                    open_date__day=previous_day.day).last()
            if get_cash_previous is not None:
                previous_cash_balance = get_cash_previous.balance
            else:
                previous_cash_balance = 0


            get_sum_bills_month_margin = DetailsBills.objects.filter(createdAt__range=[startdate, enddate],
                                                                    deleted=False).exclude(
                bills=None).aggregate(
                Sum('margin'))            
            get_sum_bills_month_cost_production = DetailsBills.objects.filter(createdAt__range=[startdate, enddate],
                                                                        deleted=False).exclude(
                    bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            get_sum_bills_cost_production = DetailsBills.objects.filter(createdAt__range=[today_start, today_end],
                                                                        deleted=False).exclude(
                    bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            month_margin = get_sum_bills_month_margin['margin__sum']        
            get_sum_bills_month_tva = Bills.objects.filter(createdAt__range=[startdate, enddate],
                                                                    deleted=False).aggregate(
                Sum('tva'))
            month_tva = get_sum_bills_month_tva['tva__sum']

            get_sum_bills_previous_month_margin = DetailsBills.objects.filter(
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).exclude(
                bills=None).aggregate(
                Sum('margin'))           
            get_sum_bills_previous_month_cost_production = DetailsBills.objects.filter(
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).exclude(
                bills=None).aggregate(
                Sum('cost_production'))['cost_production__sum']
            
            get_sum_bills_previous_cost_production = DetailsBills.objects.filter(
                createdAt__range=[prev_start, prev_end], deleted=False).exclude(
                bills=None).aggregate(Sum('cost_production'))['cost_production__sum']
            previous_month_margin = get_sum_bills_previous_month_margin['margin__sum'] 

            get_sum_bills_previous_month_tva = Bills.objects.filter(
                createdAt__range=[start_day_of_prev_month, last_day_of_prev_month], deleted=False).aggregate(
                Sum('tva'))
            previous_month_tva = get_sum_bills_previous_month_tva['tva__sum']
            

        if self.request.query_params:

            pass
        else:
            sale_of_the_day_exit, sale_of_the_day_entry = split_entry_exit(cash_types_today)

            sale_of_the_previous_day_exit, sale_of_the_previous_day_entry = split_entry_exit(cash_types_previous)

            sale_of_the_month_exit, sale_of_the_month_entry = split_entry_exit(get_sum_month)

            sale_of_the_previous_month_exit, sale_of_the_previous_month_entry = split_entry_exit(get_sum_previous_month)
        content = {'content': {'sale_of_the_day_entry': sale_of_the_day_entry,
            'sale_of_the_day_exit': sale_of_the_day_exit,
            'sale_of_the_month_entry': sale_of_the_month_entry,
            'sale_of_the_month_exit': sale_of_the_month_exit,
            'sale_of_the_previous_day_entry': sale_of_the_previous_day_entry,
            'sale_of_the_previous_day_exit': sale_of_the_previous_day_exit,
            'sale_of_the_previous_month_entry': sale_of_the_previous_month_entry,
            'sale_of_the_previous_month_exit': sale_of_the_previous_month_exit,
            'sales_today': sales_today,
            'cost_production': get_sum_bills_cost_production,
            'back_sale_of_the_day_entry': get_back(current=sale_of_the_day_entry,
                                                        previous=sale_of_the_previous_day_entry),
            'back_sale_of_the_month_exit': get_back(current=sale_of_the_month_exit,
                                                        previous=sale_of_the_previous_month_exit),
            'back_sale_of_the_day_exit': get_back(current=sale_of_the_day_exit,
                                                        previous=sale_of_the_previous_day_exit),                
            'back_sale_of_the_day': get_back(current=sales_today,
                                                        previous=sales_prev),                
            'back_cost_production': get_back(current=get_sum_bills_cost_production,
                                                        previous=get_sum_bills_previous_cost_production),
            'back_sale_of_the_month_entry': get_back(current=sale_of_the_month_entry,
                                                        previous=sale_of_the_previous_month_entry),

                                'day_settlement': settlement_today,
                                'cash_balance': cash_balance,
                                'back_cash_balance': get_back(current=cash_balance, previous=previous_cash_balance),
                                'back_day_settlement': get_back(current=settlement_today,
                                                                previous=settlement_previous),
                                'sale_of_the_month': bills_month,
                                'daily_expenses': sale_of_the_day_exit,
                                'back_daily_expenses': get_back(current=sale_of_the_day_exit,
                                                                previous=sale_of_the_previous_day_exit),
                                'back_sale_of_the_month': get_back(current=bills_month,
                                                                    previous=bills_prev_month),
                                'cumulative_month_unpaid': bills_unpaid_month,
                                'cumulative_unpaid': bills_unpaid_today,
                                'month_margin': month_margin,
                                'sale_cost_production_month': get_sum_bills_month_cost_production,
                                'month_tva': month_tva,
                                'back_month_margin': get_back(current=month_margin, previous=previous_month_margin),
                                'back_cost_production_month': get_back(current=get_sum_bills_month_cost_production, previous=get_sum_bills_previous_month_cost_production),
                                'back_month_tva': get_back(current=month_tva, previous=previous_month_tva),

                                }}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get','post'], url_path='stat_secretariat')
    def stat_secretariats(self, request, *args, **kwargs):
        # For printing results
        date_month = request.data['date_month']
        year_month = {}
        startdate = ''
        enddate = ''
        if date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = get_first_date_of_month(year=int(date_month.split("-")[0]), month=int(date_month.split("-")[1]))
            enddate = get_last_date_of_month(year=int(date_month.split("-")[0]), month=int(date_month.split("-")[1]))

        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital
        if user_hospital:
            get_doctor = Doctor.objects.filter(hospital=user_hospital).values(types=F('type')).annotate(
                total=Count('id'))
            get_appointment = Appointment.objects.filter(hospital=user_hospital,createdAt__range=[startdate, enddate]).values(status=F('appointment_status')).annotate(total=Count('id'))
            get_appointment_shape = Appointment.objects.filter(hospital=user_hospital,createdAt__range=[startdate, enddate]).values(shape=F('patient_shape')).annotate(total=Count('id'))
            statGender = Patient.objects.filter(hospital=user_hospital).values(type=F('gender')).annotate(total=Count('id'))
        else:
            get_doctor = Doctor.objects.all().values(types=F('type')).annotate(
                        total=Count('id'))
            get_appointment = Appointment.objects.filter(createdAt__range=[startdate, enddate]).values(status=F('appointment_status')).annotate(
                total=Count('id'))
            get_appointment_shape = Appointment.objects.filter(createdAt__range=[startdate, enddate]).values(shape=F('patient_shape')).annotate(
                total=Count('id'))
            statGender = Patient.objects.all().values(type=F('gender')).annotate(
                total=Count('id'))

        content = {'content': {'patients': statGender,
                               'doctors': get_doctor, 'appointments_status': get_appointment, 'appointments_shape': get_appointment_shape}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='stat_month_doctor')
    def stat_month_doctor(self, request, *args, **kwargs):
        # For printing results
        today = timezone.now()
        year = today.year
        month = today.month
        startdate = get_first_date_of_month(year=year, month=month)
        enddate = get_last_date_of_month(year=year, month=month)

        last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
        previous_day = timezone.now() - timedelta(days=1)
        start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)

        if 'hospital' in self.request.query_params:
            user_hospital = Hospital.objects.filter(id=self.request.query_params.get("hospital")).last()
        else:
            user_hospital = self.request.user.hospital

        if user_hospital:
            get_count_patient_month = Bills.objects.filter(hospital=user_hospital,doctor_id=request.query_params.get("id"), deleted=False,createdAt__range=[startdate, enddate]).count()
            get_count_patient_previous_month = Bills.objects.filter(hospital=user_hospital,doctor_id=request.query_params.get("id"), deleted=False,createdAt__range=[start_day_of_prev_month,
                                                                                    last_day_of_prev_month]).count()

            get_ca_bills_month = Bills.objects.filter(hospital=user_hospital,doctor_id=request.query_params.get("id"), deleted=False,createdAt__range=[startdate, enddate]).aggregate(Sum('amount_paid'))
            get_share_bills_month = DetailsBills.objects.filter(hospital=user_hospital,doctor_id=request.query_params.get("id"), deleted=False,createdAt__range=[startdate, enddate]).aggregate(
                Sum('share_doctor'))['share_doctor__sum']
            get_share_bills_previous_month = DetailsBills.objects.filter(hospital=user_hospital,doctor_id=request.query_params.get("id"), deleted=False,createdAt__range=[start_day_of_prev_month,last_day_of_prev_month]).aggregate(
                Sum('share_doctor'))['share_doctor__sum']
            get_ca_bills_previous_month = Bills.objects.filter(hospital=user_hospital,doctor_id=request.query_params.get("id"), deleted=False,createdAt__range=[start_day_of_prev_month,last_day_of_prev_month]).aggregate(Sum('amount_paid'))
            ca_bills_previous_month = get_ca_bills_previous_month['amount_paid__sum']
            ca_bills_month = get_ca_bills_month['amount_paid__sum']
        
        else:
            get_count_patient_month = Bills.objects.filter(doctor_id=request.query_params.get("id"), deleted=False,
                                                        createdAt__range=[startdate, enddate]).count()
            get_count_patient_previous_month = Bills.objects.filter(doctor_id=request.query_params.get("id"), deleted=False,
                                                                    createdAt__range=[start_day_of_prev_month,
                                                                                    last_day_of_prev_month]).count()

            get_ca_bills_month = Bills.objects.filter(doctor_id=request.query_params.get("id"), deleted=False,
                                                    createdAt__range=[startdate, enddate]).aggregate(
                Sum('amount_paid'))
            get_share_bills_month = DetailsBills.objects.filter(doctor_id=request.query_params.get("id"), deleted=False,
                                                    createdAt__range=[startdate, enddate]).aggregate(
                Sum('share_doctor'))['share_doctor__sum']
            get_share_bills_previous_month = DetailsBills.objects.filter(doctor_id=request.query_params.get("id"), deleted=False,
                                                            createdAt__range=[start_day_of_prev_month,
                                                                                last_day_of_prev_month]).aggregate(
                Sum('share_doctor'))['share_doctor__sum']
            get_ca_bills_previous_month = Bills.objects.filter(doctor_id=request.query_params.get("id"), deleted=False,
                                                            createdAt__range=[start_day_of_prev_month,
                                                                                last_day_of_prev_month]).aggregate(
                Sum('amount_paid'))
            ca_bills_previous_month = get_ca_bills_previous_month['amount_paid__sum']
            ca_bills_month = get_ca_bills_month['amount_paid__sum']
            

        content = {'content': {'patient_month': get_count_patient_month,
                               'back_patient': get_back(current=get_count_patient_month,
                                                        previous=get_count_patient_previous_month),
                               'CA_month': ca_bills_month,
                               'share': get_share_bills_month,
                               'back_share': get_back(current=get_share_bills_month, previous=get_share_bills_previous_month),
                               'back_ca': get_back(current=ca_bills_month, previous=ca_bills_previous_month),
                               }}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='bills_analysis')
    def analysis_bills(self, request, *args, **kwargs):
        try:
            start_date = self.request.query_params.get("start_date")
            today = timezone.now()
            # if start_date:
            # year = start_date.split("-")[0]
            # month = start_date.split("-")[1]
            # else:
                
                # year = today.year
                # month = today.month
            # year_today = today.year
            # month_today = today.month
            # if int(year) == int(year_today) and int(month) == int(month_today):
            queryset = self.filter_queryset(self.get_queryset()).exclude( cash__isnull=True)
            sum_ca = queryset.aggregate(Sum('bills_amount'))['bills_amount__sum']
            sum_paid = queryset.aggregate(Sum('amount_paid'))['amount_paid__sum']
            sum_unpaid = queryset.aggregate(Sum('balance'))['balance__sum']
            serializer = BillsSerializerAnalysis(queryset, many=True, fields=(
                'id', 'code', 'bill_type', 'createdAt', 'timeAt', 'patient', 'bills_amount','amount_received', 'cash',
                'balance', 'refund','amount_gross','amount_paid')).data
            content = {'content': serializer, 'sum_ca': sum_ca,'sum_ca': sum_ca, 'sum_paid': sum_paid, 'sum_unpaid': sum_unpaid}
            # else:
                # content = get_archive(year, month, self.request.query_params)
            return Response(data=content, status=status.HTTP_200_OK)
        except Exception as e:
            print("Error in creating", e)
            return Response(data=e, status=status.HTTP_502_BAD_GATEWAY)

    @action(detail=False, methods=['get'], url_path='bills_analysis/print')
    def print(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset()).exclude(deleted=True)
        serializer = BillsSerializerAnalysis(queryset, many=True, fields=(
                'id', 'code', 'bill_type', 'createdAt', 'timeAt', 'patient', 'type_patient', 'bills_amount','amount_received', 'cash',
                'balance', 'refund','amount_gross','amount_paid','net_payable')).data
        sum_unpaid = queryset.aggregate(Sum('balance'))
        sum_paid = queryset.aggregate(Sum('amount_paid'))
        sum_total = queryset.aggregate(Sum('net_payable'))
        if 'balance' in self.request.query_params:
            html_render = get_template('export_bills_analysis_unpaid.html')
        else:
            html_render = get_template('export_bills_analysis.html')
        html_content = html_render.render(
            {'bills': serializer, 'hospital': self.request.user.hospital, 'sum_unpaid': sum_unpaid['balance__sum'],
             'sum_paid': sum_paid['amount_paid__sum'], 'sum_total': sum_total['net_payable__sum'],
             'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("utf-16")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)

    @action(detail=False, methods=['get'], url_path='bills_analysis/export')
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset()).exclude(deleted=True)
        serializer = BillsSerializerAnalysis(queryset, many=True).data
        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Factures', index=1)
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.freeze_panes = 'I1'

        columns = ['Code', 'Type', 'Client', 'Type patient', 'Caissier', 'Session de Caisse', 'Montant', 'Avance',
                   'Solde','Rendu',
                   'Date']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        # Iterate through all coins

        for _, act in enumerate(serializer, 1):
            row_num += 1

            row = [
                act['code'],
                act['bill_type'],
                act['patient']['name'],
                act['patient']['type_patient']['title'],
                act['cash']['user']['username'],
                act['cash']['code'],
                act['net_payable'],
                act['amount_paid'],
                act['balance'],
                act['refund'],
                act['createdAt']
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(excelfile)
        response = HttpResponse(content_type='application/vnd.ms-excel')

        # tell the browser what the file is named
        response['Content-Disposition'] = 'attachment;filename="some_file_name.xlsx"'

        # put the spreadsheet data into the response
        response.write(excelfile.getvalue())

        # return the response
        return response

    @action(detail=False, methods=['get'], url_path='current')
    def current_bills(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset()).filter(cash__user_id=self.request.user.id, cash__is_active=True, deleted=False)

        serializer = BillsSerializer(queryset, many=True, fields=('id', 'bills','cash', 'code', 'patient', 'patient_account', 'amount_received', 'amount_paid', 'phone_number', 'bills_date', 'bill_type', 'additional_info', 'refund', 'tva', 'bills_amount', 'balance', 'refund', 'createdAt', 'timeAt', 'bank_card_number', 'phone_number_momo', 'phone_number_om', 'amount_om', 'amount_momo', 'amount_prepaid','amount_bank_card', 'amount_cash')).data
        content = {'content': serializer}
        return Response(data=content, status=status.HTTP_200_OK)  
      
    @action(detail=False, methods=['get'], url_path='details_client')
    def details_client(self, request, *args, **kwargs):
       
        
        if self.request.query_params.get('type')=='pdf':
            qs = (
                    DetailsBills.objects
                    .filter(deleted=False, patient_id=self.request.query_params.get("patient"))
                    .values(   
                        'patient__name',
                        'dish_id',  # optional
                        'dish__name',         # optional
                    )
                    .annotate(
                        total_quantity=Sum('quantity_served'),
                        total_amount_net=Sum('amount_net'),
                        total_cost_production=Sum('cost_production'),
                        total_margin=Sum('margin'),
                    )
                    .order_by( 'dish_id')
                )
            queryset = self.filter_queryset(self.get_queryset()).filter(hospital_id=self.request.user.hospital.id)
            serializer = self.get_serializer(queryset, many=True, ).data
            html_render = get_template('details_client_dish.html')
            # logo = settings.MEDIA_ROOT + '/logo.png'
            html_content = html_render.render(
                {'lignes': qs,
                'hospital': self.request.user.hospital,
                'total_commandes': qs.aggregate(Sum('total_quantity'))['total_quantity__sum'],
                'total_ca': qs.aggregate(Sum('total_amount_net'))['total_amount_net__sum'],
                'date': datetime.today().strftime("%Y-%m-%d"),
                'patient_name': qs[0],
                'url': request.build_absolute_uri('/')})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(content_type='application/pdf')
                filename = 'Extrait' + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                response.write(result.getvalue())
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            queryset = self.filter_queryset(self.get_queryset())

            get_account_prepaid = PatientAccount.objects.filter(patient_id=self.request.query_params.get("patient"), type_account='PREPAID').last()
            get_account_private = PatientAccount.objects.filter(patient_id=self.request.query_params.get("patient"), type_account='PRIVATE').last()
            content = {'content': {'amount_om':queryset.aggregate(Sum('amount_om'))['amount_om__sum'],
            'amount_prepaid':queryset.aggregate(Sum('amount_prepaid'))['amount_prepaid__sum'],
            'amount_momo':queryset.aggregate(Sum('amount_momo'))['amount_momo__sum'],
            'amount_bank_card':queryset.aggregate(Sum('amount_bank_card'))['amount_bank_card__sum'],
            'amount_cash':queryset.aggregate(Sum('amount_cash'))['amount_cash__sum'],
            'balance_private':get_account_private.balance,
            'balance_prepaid':get_account_prepaid.balance
            }}
            return Response(data=content, status=status.HTTP_200_OK)
    # @action(detail=False, methods=['post', 'get'], url_path='statistic_dom')
    # def get_statistic_dom(self, request):
    #     stat_date_month = request.data['start_date_dom']
    #     year_month = {}
    #     startdate = ''
    #     enddate = ''
    #     if stat_date_month is None:
    #         today = date.today()
    #         year_month['year'] = today.year
    #         year_month['month'] = today.month
    #         startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
    #         enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
    #     else:
    #         startdate = request.data['start_date_dom']
    #         enddate = request.data['end_date_dom']
    #     get_product = Product.objects.all()
    #
    #
    #     stat_category = {'category': 'qte'}
    #     for product in get_product:
    #         get_details_bill = DetailsBills.objects.filter(
    #             details_stock__product__category_id=product.category_id,
    #             createdAt__range=[startdate,enddate]).aggregate(
    #             Sum('quantity_served'))
    #         if get_details_bill['quantity_served__sum'] is None:
    #             stat_category[product.category.name] = 0
    #         else:
    #             stat_category[product.category.name] = get_details_bill['quantity_served__sum']
    #     get_medical_act = Medical_act.objects.all()
    #     stat_domaine = {'domaine': 'qte'}
    #     for act in get_medical_act:
    #         get_details_bill = DetailsBills.objects.filter(
    #             medical_act__medical_areas_id=act.medical_areas_id,
    #             createdAt__range=[startdate,enddate]).aggregate(
    #             Sum('quantity_served'))
    #         if get_details_bill['quantity_served__sum'] is None:
    #             stat_domaine[act.medical_areas.name] = 0
    #         else:
    #             stat_domaine[act.medical_areas.name] = get_details_bill['quantity_served__sum']
    #     content = {'content': {'stat_category':stat_category,'stat_domaine':stat_domaine}}
    #     return Response(data=content, status=status.HTTP_200_OK)


class PatientSettlementViewSet(viewsets.ModelViewSet):
    queryset = PatientSettlement.objects.filter(deleted=False)
    serializer_class = PatientSettlementSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = PatientSettlementFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return PatientSettlement.objects.filter( deleted=False, hospital=user_hospital)
        return PatientSettlement.objects.filter(deleted=False)

    # def get_queryset(self):
    #     return PatientSettlement.objects.filter(cash__user_id=self.request.user.id, cash__is_active=True,
    #                                             deleted=False).all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        patient_settlement_form = PatientSettlementForm(request.data)
        if patient_settlement_form.is_valid():
            user = self.request.user
            get_cash = Cash.objects.filter(user_id=user.id, hospital = user.hospital, is_active=True).last()
            if get_cash:
                get_bills = Bills.objects.filter(patient_id=request.data['patient'], hospital = user.hospital, deleted=False).all()
                dette = get_bills.aggregate(Sum('balance'))['balance__sum']
                patient_settlement = patient_settlement_form.save()
                patient_settlement.hospital = user.hospital
                patient_settlement.cash_id = get_cash.id
                patient_settlement.current_balance = dette

                remaining = request.data['amount_paid']

                for bill in get_bills:
                    if remaining <= 0:
                        break

                    if remaining >= bill.balance:
                        remaining -= bill.balance
                        bill.balance = 0
                        bill.status = 'PAID'
                    else:
                        bill.balance -= remaining
                        remaining = 0
                        bill.status = 'PARTIAL'

                    bill.save()
                if 'overpayment_action' in request.data:
                    patient_settlement.overpayment_action=request.data['overpayment_action']
                    if request.data['overpayment_action'] == 'PREPAID':
                        get_account_prepaid=PatientAccount.objects.filter(hospital = user.hospital,patient_id=request.data['patient'], type_account='PREPAID').last()
                        get_account_prepaid.balance += float(request.data['refund'])
                        get_account_prepaid.save()
                        DetailsPatientAccount.objects.create(type_operation = 'CREDIT', hospital = user.hospital,patient_account_id=get_account_prepaid.id, source='SETTEL',reference_id=patient_settlement.id, balance = int(request.data['refund']), balance_before = get_account_prepaid.balance - int(request.data['refund']), balance_after = get_account_prepaid.balance, user_id = user.id)

                get_bills1 = Bills.objects.filter(patient_id=request.data['patient'], hospital = user.hospital,deleted=False).all()
                new_balance = get_bills1.aggregate(Sum('balance'))['balance__sum']
                patient_settlement.new_balance = new_balance
                # patient_settlement.bills_id = get_cash.id
                patient_settlement.timeAt = time.strftime("%H:%M:%S", time.localtime())
                patient_settlement.save()
                if 'print' in request.query_params:
                    # get_details_bills = DetailsBills.objects.filter(cash__user_id=request.data['cashier'],
                    #                                                 bills_id=bills).all()
                    html_render = get_template('patient_settlement.html')
                    get_detail_account = get_prepaid_account_detail(request, reference_id=patient_settlement.id,user=user, source='SETTLE')
                    # logo = settings.MEDIA_ROOT + '/logo.png'
                    if get_detail_account:

                        html_content = html_render.render(
                            {'patient_settlement': patient_settlement, 'dette': dette,
                            'reste': int(get_bills.aggregate(Sum('balance'))['balance__sum']),
                            'hospital': self.request.user.hospital,
                            'account_prepaid': get_detail_account,
                            'url': request.build_absolute_uri('/'),
                            'Cashier': user.username})
                    else:
                        html_content = html_render.render(
                        {'patient_settlement': patient_settlement, 'dette': dette,
                         'reste': int(get_bills.aggregate(Sum('balance'))['balance__sum']),
                         'hospital': self.request.user.hospital,
                         'url': request.build_absolute_uri('/'),
                         'Cashier': user.username})
                    result = BytesIO()
                    pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
                    if not pdf.err:
                        response = HttpResponse(content_type='application/pdf')
                        filename = 'Facture_' + str(patient_settlement.code) + '.pdf'
                        response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                        response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                        response.write(result.getvalue())
                        return response
                    else:
                        errors = {"pdf": ["Error to generate PDF."]}
                        return Response(data=errors, status=status.HTTP_500)
                else:
                    serializer = self.get_serializer(patient_settlement, many=False)
                    return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            else:
                errors = {"cash": ["No Cash open."]}
                return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        errors = {**patient_settlement_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        user = self.request.user
        patient_settlement = self.get_object()
        patient_settlement_form = PatientSettlementForm(request.data, instance=patient_settlement)
        if patient_settlement_form.is_valid():
            patient_settlement = patient_settlement_form.save()
            patient_settlement.timeAt = time.strftime("%H:%M:%S", time.localtime())
            patient_settlement.save()
            if 'print' in request.query_params:
                # get_details_bills = DetailsBills.objects.filter(cash__user_id=request.data['cashier'],
                #                                                 bills_id=bills).all()
                html_render = get_template('patient_settlement.html')
                # logo = settings.MEDIA_ROOT + '/logo.png'
                get_detail_account = get_prepaid_account_detail(request, reference_id= patient_settlement.id,user=user, source='SETTLE')
                    # logo = settings.MEDIA_ROOT + '/logo.png'
                if get_detail_account:

                    html_content = html_render.render(
                        {'patient_settlement': patient_settlement, 'dette': patient_settlement.current_balance,
                        'reste': patient_settlement.new_balance,
                        'hospital': self.request.user.hospital,
                        'account_prepaid': get_detail_account,
                        'url': request.build_absolute_uri('/'),
                        'Cashier': patient_settlement.cash.user.username})
                else:
                    html_content = html_render.render(
                    {'patient_settlement': patient_settlement, 'dette': patient_settlement.current_balance,
                        'reste': patient_settlement.new_balance,
                        'hospital': self.request.user.hospital,
                        'url': request.build_absolute_uri('/'),
                        'Cashier': patient_settlement.cash.user.username})
                
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                    link_callback=link_callback)
                if not pdf.err:
                    response = HttpResponse(content_type='application/pdf')
                    filename = 'Facture_' + str(patient_settlement.code) + '.pdf'
                    response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                    response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                    response.write(result.getvalue())
                    return response
                else:
                    errors = {"pdf": ["Error to generate PDF."]}
                    return Response(data=errors, status=status.HTTP_500)
            else:
                serializer = self.get_serializer(patient_settlement, many=False)
                return Response(data=serializer.data, status=status.HTTP_200_OK)
            
        errors = {**patient_settlement_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        patient_settlement = self.get_object()
        patient_settlement.deleted = True
        patient_settlement.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_category(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            product = Product.objects.filter(name__icontains=data['name'])
            if product:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_cash(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        bills_other = queryset.aggregate(
                        total_cash=Sum('amount_cash'),
                        total_bank_card=Sum('amount_bank_card'),
                        total_momo=Sum('amount_momo'),
                        total_om=Sum('amount_om'),
                        total_prepaid=Sum('amount_prepaid'),
                    )

        response = {
                    "CASH": bills_other["total_cash"] or 0,
                    "BANK_CARD": bills_other["total_bank_card"] or 0,
                    "MOMO": bills_other["total_momo"] or 0,
                    "OM": bills_other["total_om"] or 0,
                    "PREPAID": bills_other["total_prepaid"] or 0,
                }
            
            
        content = {'content': serializer, 'statistics': response}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get', 'post'], url_path='details_patients_settlements',
            permission_classes=[AllowAny])
    def details(self, request, *args, **kwargs):
        if request.method == 'GET':
            if request.query_params.get("patient") != 'null':
                get_bills = Bills.objects.filter(
                    patient_id=request.query_params.get("patient"),
                    balance__gt=0
                ).order_by('createdAt')
                serializer = BillsSerializer(get_bills, many=True)
                return Response(data=serializer.data, status=status.HTTP_200_OK)
        else:
            get_details = DetailsSupplies.objects.create(hospital = self.request.user.hospital,supplies_id=request.data['id_supplies'],
                                                         product_id=request.data['product']['id'],
                                                         quantity=request.data['quantity'],
                                                         total_amount=request.data['total_amount'],
                                                         arrival_price=request.data['arrival_price'],
                                                         product_code=request.data['product_code'],
                                                         product_name=request.data['product_name'])
            details = DetailsSuppliesSerializer(get_details, many=False)
            return Response(data=details.data, status=status.HTTP_200_OK)

class SuppliersViewSet(viewsets.ModelViewSet):
    queryset = Suppliers.objects.filter(deleted=False)
    serializer_class = SuppliersSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = SuppliersFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Suppliers.objects.filter(deleted=False, hospital=user_hospital)
        return Suppliers.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        suppliers_form = SuppliersForm(request.data)
        if suppliers_form.is_valid():
            suppliers = suppliers_form.save()
            suppliers.hospital = self.request.user.hospital
            suppliers.save()
            serializer = self.get_serializer(suppliers, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**suppliers_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        suppliers = self.get_object()
        suppliers_form = SuppliersForm(request.data, instance=suppliers)
        if suppliers_form.is_valid():
            suppliers = suppliers_form.save()
            suppliers.save()
            serializer = self.get_serializer(suppliers, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**suppliers_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        suppliers = self.get_object()
        suppliers.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                data = {}
                for dbframe in dbframe.itertuples():
                    get_obj = Suppliers.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.Nom).last()
                    if get_obj:
                        get_obj.name = checkContent(content=dbframe.Nom)
                        get_obj.phone = checkContentPhone(content=dbframe.Phone)
                        get_obj.mailbox = checkContent(content=dbframe.Boite_postale)    
                        get_obj.city = checkContent(content=dbframe.Ville)
                        get_obj.country = checkContent(content=dbframe.Pays)
                        get_obj.fax = checkContent(content=dbframe.Fax)
                        get_obj.taxpayer_number = checkContent(content=dbframe.Numero_contribuable)        
                        get_obj.email = checkContent(content=dbframe.Email)
                        get_obj.name_representative = checkContent(content=dbframe.Nom_representant)
                        get_obj.phone_representative =checkContentPhone(content=dbframe.Phone_representant)
                        get_obj.save()
                    else:
                        Suppliers.objects.create(hospital = self.request.user.hospital,name = checkContent(content=dbframe.Nom),
                        phone = checkContentPhone(content=dbframe.Phone),
                        mailbox = checkContent(content=dbframe.Boite_postale),       
                        city = checkContent(content=dbframe.Ville),
                        country = checkContent(content=dbframe.Pays),
                        fax = checkContent(content=dbframe.Fax),
                        taxpayer_number = checkContent(content=dbframe.Numero_contribuable),      
                        email = checkContent(content=dbframe.Email),
                        name_representative = checkContent(content=dbframe.Nom_representant),
                        phone_representative = checkContentPhone(content=dbframe.Phone_representant))

        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_exp(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_suppliers(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            suppliers = Suppliers.objects.filter(hospital=self.request.user.hospital,name__icontains=data['name'])
            if suppliers:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

class Type_patientViewSet(viewsets.ModelViewSet):
    queryset = Type_patient.objects.filter(deleted=False)
    serializer_class = Type_patientSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = Type_patientFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user = self.request.user

        qs = (
            Type_patient.objects
            .select_related('hospital')   # très important
        )

        if user.hospital_id:
            qs = qs.filter(hospital_id=user.hospital_id)

        return qs
   
    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = Type_patientForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = Type_patientForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'], url_path='all')
    def get_all_exp(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    @action(detail=False, methods=['post'], url_path='title/exists', permission_classes=[AllowAny])
    def check_obj(self, request, *args, **kwargs):
        data = request.data
        errors = {"title": ["This field already exists."]}
        if 'title' in data:
            obj = Type_patient.objects.filter(title__icontains=data['title'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

from rest_framework.filters import SearchFilter
class PatientViewSet(viewsets.ModelViewSet):
    queryset = Patient.objects.filter(deleted=False)
    serializer_class = PatientSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = PatientFilter
    search_fields = ['^name', '^phone']
    filter_backends = (filters.DjangoFilterBackend, SearchFilter)


    def get_queryset(self):
        user = self.request.user

        qs = (
            Patient.objects
            .select_related('hospital')   # très important
            .filter(deleted=False)
        )

        if user.hospital_id:
            qs = qs.filter(hospital_id=user.hospital_id)

        return qs

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        patient_form = PatientForm(request.data)
        if patient_form.is_valid():
            patient = patient_form.save()
            patient.hospital = self.request.user.hospital
            patient.save()
            get_account=PatientAccount.objects.filter(patient=patient, type_account='PRIVATE').last()
            if get_account:
                pass
            else:
                PatientAccount.objects.create(hospital = self.request.user.hospital,patient=patient, type_account='PRIVATE')
        
            get_account=PatientAccount.objects.filter(patient=patient, type_account='PREPAID').last()
            if get_account:
                pass
            else:
                PatientAccount.objects.create(hospital = self.request.user.hospital,patient=patient, type_account='PREPAID')
        
            serializer = self.get_serializer(patient, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**patient_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        patient = self.get_object()
        patient_form = PatientForm(request.data, instance=patient)
        if patient_form.is_valid():
            patient = patient_form.save()
            patient.hospital = self.request.user.hospital
            patient.save()
            serializer = self.get_serializer(patient, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**patient_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        patient = self.get_object()
        get_appointment = Appointment.objects.filter(patient_id=patient.id).last()
        get_consultation = Consultation.objects.filter(patient_id=patient.id).last()
        get_Bills = Bills.objects.filter(patient_id=patient.id, deleted=False).last()
        get_PatientSettlement = PatientSettlement.objects.filter(patient_id=patient.id, deleted=False).last()
        if get_appointment or get_consultation or get_Bills or get_PatientSettlement:
            errors = {"patient": ["The patient is used."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            patient.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['post', 'get'], url_path='days_patients')
    def get_days_patients(self, request):
        date_month = request.data['date_month']
        year_month = {}
        startdate = ''
        enddate = ''
        if date_month is None:
            today = date.today()
            year_month['year'] = today.year
            year_month['month'] = today.month
            startdate = get_first_date_of_month(year=int(year_month['year']), month=int(year_month['month']))
            enddate = get_last_date_of_month(year=int(year_month['year']), month=int(year_month['month']))

        else:
            startdate = get_first_date_of_month(year=int(date_month.split("-")[0]), month=int(date_month.split("-")[1]))
            enddate = get_last_date_of_month(year=int(date_month.split("-")[0]), month=int(date_month.split("-")[1]))
        patient = Patient.objects.filter(createdAt__range=[startdate, enddate], deleted=False).values(
                days=F('createdAt')
            ).annotate(turnover=Count('id'))


        if self.request.query_params.get('type') == 'pdf':
            html_render = get_template('export_stat_days.html')
            html_content = html_render.render(
                {'bills': patient.order_by('days'), 'hospital': self.request.user.hospital,
                 'month': date_month.split("-")[1], 'year': date_month.split("-")[0],
                 'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
                response['Content-Disposition'] = 'inline; filename="' + filename + '"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
                return response
            else:
                errors = {"pdf": ["Error to generate PDF."]}
                return Response(data=errors, status=status.HTTP_500)
        else:
            # stat_category = {'category': 'C.A'}
            # for bill in bills:
            #     stat_category[bill['category']] = bill['turnover']
            # content = {'content': {'stat_category': stat_category}}
            # labels = list()
            # values = list()
            # for bill in bills:
            #     labels.append(bill['category'])
            #     values.append(bill['turnover'])
            content = {'content': patient.order_by('days')}
            return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='print')
    def print(self, request, *args, **kwargs):
        if self.request.query_params == {}:
            queryset = Patient.objects.all()
        else:
            queryset = self.filter_queryset(self.get_queryset())
        patient_update = []
        for patient_id in PatientSerializer(queryset, many=True).data:
            get_bills = Bills.objects.filter(patient_id=dict(patient_id)['id'], deleted=False).all()
            update_patient = dict(patient_id)
            update_patient['solde'] = get_bills.aggregate(Sum('balance'))['balance__sum']
            update_patient['total_bills'] = get_bills.aggregate(Sum('net_payable'))['net_payable__sum']
            update_patient['total_payment'] = get_bills.aggregate(Sum('amount_paid'))['amount_paid__sum']
            patient_update.append(update_patient)

        html_render = get_template('export_patient.html')
        html_content = html_render.render(
            {'patients': patient_update, 'hospital': self.request.user.hospital,
             'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)

    @action(detail=False, methods=['get'], url_path='export')
    def exports(self, request, *args, **kwargs):
        if self.request.query_params == {}:
            queryset = Patient.objects.all()
        else:
            queryset = self.filter_queryset(self.get_queryset())
        patient_update = []
        for patient_id in PatientSerializer(queryset, many=True).data:
            get_bills = Bills.objects.filter(patient_id=dict(patient_id)['id'], deleted=False).all()
            update_patient = dict(patient_id)
            update_patient['solde'] = get_bills.aggregate(Sum('balance'))['balance__sum']
            update_patient['total_bills'] = get_bills.aggregate(Sum('net_payable'))['net_payable__sum']
            update_patient['total_payment'] = get_bills.aggregate(Sum('amount_paid'))['amount_paid__sum']
            patient_update.append(update_patient)
        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Liste des patients', index=1)
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.freeze_panes = 'I1'

        columns = ['Code', 'Nom', 'Age', 'Sexe', 'Date_naissance', 'Nom_mere', 'Tel', 'Enfant',
                   'Date_ID', 'Num_ID', 'Nom_urgence', 'Contact_urgence', 'Ville', 'Quartier', 'Religion', 'Forme',
                   'Groupe_sanguin', 'Electrophoreze', 'Solde', 'Total facture', 'Total paiement']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        # Iterate through all coins

        for _, act in enumerate(patient_update, 1):
            row_num += 1

            row = [
                act['code'],
                act['name'],
                act['age'],
                act['gender'],
                act['dateNaiss'],
                act['mother_name'],
                act['phone'],
                act['child'],
                act['date_id'],
                act['number_id'],
                act['emergency_name'],
                act['emergency_contact'],
                act['city']['name'],
                act['district'],
                act['religion'],
                act['shape'],
                act['blood_group'],
                act['electrophoresis'],
                act['solde'],
                act['total_bills'],
                act['total_payment']
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(excelfile)
        response = HttpResponse(content_type='application/vnd.ms-excel')

        # tell the browser what the file is named
        response['Content-Disposition'] = 'attachment;filename="some_file_name.xlsx"'

        # put the spreadsheet data into the response
        response.write(excelfile.getvalue())

        # return the response
        return response



    @action(detail=False, methods=['get'], url_path='allPatients')
    def allPatients(self, request, *args, **kwargs):
        # Préfiltrage (tu peux personnaliser selon les champs utiles)
        queryset = self.filter_queryset(self.get_queryset())

        # Optimisation : ne charger que certains champs (si utile)
        # queryset = queryset.only('id', 'first_name', 'last_name')  # exemple

        # Pagination DRF
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        # Fallback si pagination désactivée
        serializer = self.get_serializer(queryset, many=True, fields = ('id', 'code','type_patient', 'is_assured', 'insurance', 'name'))
        return Response({'content': serializer.data}, status=status.HTTP_200_OK)
    # @action(detail=False, methods=['get'], url_path='allPatients')
    # def allPatients(self, request, *args, **kwargs):
    #     queryset = self.filter_queryset(self.get_queryset())
    #     serializer = self.get_serializer(queryset, many=True,fields = ('id', 'code','type_patient', 'is_assured', 'insurance_agency', 'name', 'account', 'prepaid'))
    #     content = {'content': serializer.data}
    #     return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='name/exists')
    def check_name(self, request, *args, **kwargs):
        errors = {"name": ["This field already exists."]}
        if 'name' in request.data:
            users = Patient.objects.filter(name__icontains=request.data['name'])
            if users:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_patients(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        patient_update = []
        for patient_id in PatientSerializer(queryset, many=True).data:
            get_bills = Bills.objects.filter(patient_id=dict(patient_id)['id'], deleted=False).all()
            update_patient = dict(patient_id)
            update_patient['solde'] = get_bills.aggregate(Sum('balance'))['balance__sum']
            update_patient['total_bills'] = get_bills.aggregate(Sum('net_payable'))['net_payable__sum']
            update_patient['total_payment'] = get_bills.aggregate(Sum('amount_paid'))['amount_paid__sum']
            patient_update.append(update_patient)

        content = {'content': patient_update}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                for dbframe in dbframe.itertuples():
                    get_patient = Patient.objects.filter(name__icontains=dbframe.Nom, dateNaiss=dbframe.Date_naissance).last()
                    if get_patient:
                        get_patient.name = dbframe.Nom
                        get_patient.phone = dbframe.Contact
                        get_patient.dateNaiss = dbframe.Date_naissance
                        get_patient.city_it = get_city.id
                        get_patient.scale_price = dbframe.Valeur_cle
                        get_patient.district = dbframe.Quartier
                        get_patient.religion = dbframe.Religion
                        get_patient.emergency_name = dbframe.Nom_urgence
                        get_patient.emergency_contact = dbframe.Contact_urgence
                        get_patient.mother_name = dbframe.Nom_mere
                        get_patient.save()
                    else:
                        if str(dbframe.ville) == 'nan':
                            get_city = City.objects.filter(name__icontains='default').last()
                            if get_city:
                                Patient.objects.create(hospital = self.request.user.hospital,name=checkContent(content=dbframe.Nom),
                                                    phone=checkContent(content=dbframe.Contact),
                                                    dateNaiss=checkContent(content=dbframe.Date_naissance),
                                                    city_it=get_city.id,
                                                    scale_price=checkContent(content=dbframe.Valeur_cle),
                                                    district=checkContent(content=dbframe.Quartier),
                                                    religion=checkContent(content=dbframe.Religion),
                                                    emergency_name=checkContent(content=dbframe.Nom_urgence),
                                                    emergency_contact=checkContent(content=dbframe.Contact_urgence),
                                                    mother_name=checkContent(content=dbframe.Nom_mere))
                            else:
                                create_city = City.objects.create(hospital = self.request.user.hospital,name='default')
                                Patient.objects.create(hospital = self.request.user.hospital,name=checkContent(content=dbframe.Nom),
                                                    phone=checkContent(content=dbframe.Contact),
                                                    dateNaiss=checkContent(content=dbframe.Date_naissance),
                                                    city_it=create_city.id,
                                                    scale_price=checkContent(content=dbframe.Valeur_cle),
                                                    district=checkContent(content=dbframe.Quartier),
                                                    religion=checkContent(content=dbframe.Religion),
                                                    emergency_name=checkContent(content=dbframe.Nom_urgence),
                                                    emergency_contact=checkContent(content=dbframe.Contact_urgence),
                                                    mother_name=checkContent(content=dbframe.Nom_mere))
                        else:
                            get_city = City.objects.filter(name__icontains=dbframe.ville).last()
                            if get_city:
                                Patient.objects.create(hospital = self.request.user.hospital,name=checkContent(content=dbframe.Nom),
                                               phone=checkContent(content=dbframe.Contact),
                                               dateNaiss=checkContent(content=dbframe.Date_naissance),
                                               city_id=get_city.id,
                                               scale_price=checkContent(content=dbframe.Valeur_cle),
                                               district=checkContent(content=dbframe.Quartier),
                                               religion=checkContent(content=dbframe.Religion),
                                               emergency_name=checkContent(content=dbframe.Nom_urgence),
                                               emergency_contact=checkContent(content=dbframe.Contact_urgence),
                                               mother_name=checkContent(content=dbframe.Nom_mere))
                            else:
                                get_region = Region.objects.filter(name__icontains='default').last()
                                if get_region:
                                    create_city = City.objects.create(hospital = self.request.user.hospital,region_id=get_region.id,name=dbframe.Nom)
                                    Patient.objects.create(hospital = self.request.user.hospital,name=checkContent(content=dbframe.Nom),
                                               phone=checkContent(content=dbframe.Contact),
                                               dateNaiss=checkContent(content=dbframe.Date_naissance),
                                               city_it=create_city.id,
                                               scale_price=checkContent(content=dbframe.Valeur_cle),
                                               district=checkContent(content=dbframe.Quartier),
                                               religion=checkContent(content=dbframe.Religion),
                                               emergency_name=checkContent(content=dbframe.Nom_urgence),
                                               emergency_contact=checkContent(content=dbframe.Contact_urgence),
                                               mother_name=checkContent(content=dbframe.Nom_mere))
                                else:
                                    create_region = Region.objects.create(hospital = self.request.user.hospital,name='default')
                                    create_city = City.objects.create(hospital = self.request.user.hospital,region_id=create_region.id,name=dbframe.Nom)
                                    Patient.objects.create(hospital = self.request.user.hospital,name=checkContent(content=dbframe.Nom),
                                               phone=checkContent(content=dbframe.Contact),
                                               dateNaiss=checkContent(content=dbframe.Date_naissance),
                                               city_it=create_city.id,
                                               scale_price=checkContent(content=dbframe.Valeur_cle),
                                               district=checkContent(content=dbframe.Quartier),
                                               religion=checkContent(content=dbframe.Religion),
                                               emergency_name=checkContent(content=dbframe.Nom_urgence),
                                               emergency_contact=checkContent(content=dbframe.Contact_urgence),
                                               mother_name=checkContent(content=dbframe.Nom_mere))
                                
                            
        return Response(status=status.HTTP_200_OK)

class PatientAccountViewSet(viewsets.ModelViewSet):
    queryset = PatientAccount.objects.all()
    serializer_class = PatientAccountSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = PatientAccountFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return PatientAccount.objects.filter(hospital=user_hospital)
        return PatientAccount.objects.all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = PatientAccountForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = PatientAccountForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='all')
    def get_all_depart(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='title/exists', permission_classes=[AllowAny])
    def check_account(self, request, *args, **kwargs):
        data = request.data
        errors = {"title": ["This field already exists."]}
        if 'title' in data:
            obj = PatientAccount.objects.filter(hospital=self.request.user.hospital,title__icontains=data['title'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


class DetailsSuppliesViewSet(viewsets.ModelViewSet):
    queryset = DetailsSupplies.objects.all()
    serializer_class = DetailsSuppliesSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DetailsSuppliesFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        
        user = self.request.user
        if user.hospital:
            if self.request.GET.get("supplies"):
                if (
                    self.request.GET.get("supplies") == "undefined"
                    or self.request.GET.get("supplies") == "null"
                ):
                    return DetailsSupplies.objects.filter(
                        user_id=user.id, hospital=user.hospital, supplies__is_accounted = False
                    ).all()
                else:
                    if self.request.GET.get("supplies") == "reset":
                        supplies = DetailsSupplies.objects.filter(
                            user_id=user.id, supplies_id=None, hospital=user.hospital
                        ).all()
                        supplies.delete()
                        return DetailsSupplies.objects.none()
                    else:
                        return DetailsSupplies.objects.filter(
                            supplies_id=self.request.GET.get("supplies"), hospital=user.hospital
                        ).all()
            else:
                return DetailsSupplies.objects.filter(
                    user_id=user.id, supplies_id=None, hospital=user.hospital
                ).all()
        else:
            if self.request.GET.get("supplies"):
                if (
                    self.request.GET.get("supplies") == "undefined"
                    or self.request.GET.get("supplies") == "null"
                ):
                    return DetailsSupplies.objects.filter(user_id=user.id, supplies_id=None).all()
                else:
                    if self.request.GET.get("supplies") == "reset":
                        supplies = DetailsSupplies.objects.filter(user_id=user.id, supplies_id=None).all()
                        supplies.delete()
                        return DetailsSupplies.objects.none()
                    else:
                        return DetailsSupplies.objects.filter(supplies_id=self.request.GET.get("supplies")).all()
            else:
                return DetailsSupplies.objects.filter(user_id=user.id, supplies_id=None).all()


    @transaction.atomic
    def create(self, request, *args, **kwargs):
        user = self.request.user
        # if getHospital.is_inventory == False:
        detailsSupplies_form = DetailsSuppliesForm(data=request.data)
        get_supplies = Supplies.objects.filter(id = request.data['supplies'], is_accounted = False, user_id = user.id).last()
        if detailsSupplies_form.is_valid():
            
            # get_supplies = Supplies.objects.filter(id=request.data['supplies']).last()
            ingredient_id = request.data.get("ingredient")

            # Start with common filters
            filters = {
                "hospital": user.hospital,
                "user_id": user.id,
                "supplies": get_supplies,
            }
            # Add the specific item type

            if ingredient_id:
                filters["ingredient_id"] = ingredient_id
            else:
                get_details = None
                errors = {"supplies": ["No type of product specify."]}
                return Response(errors, status=status.HTTP_400_BAD_REQUEST)

            # Get the last matching entry
            get_details = DetailsSupplies.objects.filter(**filters).last()
            if get_details:
                get_details.quantity_two = request.data["quantity_two"]
                get_details.quantity = request.data["quantity"]
                get_details.save()
                return Response(status=status.HTTP_201_CREATED)

            else:
                
                detailsSupplies = detailsSupplies_form.save()
                detailsSupplies.hospital = user.hospital
                detailsSupplies.storage_depots = request.data["storage_depots"]
                detailsSupplies.supplies = get_supplies
                unit_cost = detailsSupplies.total_amount / detailsSupplies.quantity
                print(unit_cost)
                detailsSupplies.unit_price = unit_cost
                # if detailsSupplies.ingredient.stock_quantity > 0:
                #     old_value = Decimal(detailsSupplies.ingredient.stock_quantity) * Decimal(detailsSupplies.ingredient.price_per_unit)
                #     new_value = Decimal(detailsSupplies.quantity) * Decimal(unit_cost)
                #     total_qty = Decimal(detailsSupplies.ingredient.stock_quantity) + Decimal(detailsSupplies.quantity)
                #     price = (old_value + new_value) / total_qty
                #     detailsSupplies.cmup = price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                # else:
                #     detailsSupplies.cmup = Decimal(unit_cost).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                detailsSupplies.user_id = user.id
                detailsSupplies.timeAt = time.strftime(
                    "%H:%M:%S", time.localtime()
                )
                detailsSupplies.save()
                get_supplies.supply_amount += detailsSupplies.total_amount
                get_supplies.suppliers_id = request.data['suppliers']
                if 'additional_info' in request.data:
                    get_supplies.additional_info = request.data['additional_info']
                get_supplies.save()
                serializer = self.get_serializer(detailsSupplies, many=False)
                return Response(
                    data=serializer.data, status=status.HTTP_201_CREATED
                )
        errors = {**detailsSupplies_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        # else:
        #     errors = {"is_inventory": ["Current inventory."]}
        #     return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        path = request.path
        end_path = path.rsplit("/", 1)[-1]
        get_bills = Supplies.objects.filter(id=end_path).last()
        
        detailsSupplies_form = DetailsSuppliesForm(data=request.data)
        if detailsSupplies_form.is_valid():
            user = self.request.user
        
            detailsSupplies = detailsSupplies_form.save()
            detailsSupplies.hospital = user.hospital
            detailsSupplies.storage_depots_id = request.data[
                "storage_depots"
            ]
            detailsSupplies.supplies_id = end_path
            detailsSupplies.user_id = user.id
            detailsSupplies.timeAt = time.strftime(
                "%H:%M:%S", time.localtime()
            )
            detailsSupplies.save()
            serializer = self.get_serializer(detailsSupplies, many=False)
            return Response(
                data=serializer.data, status=status.HTTP_201_CREATED
            )

        errors = {**detailsSupplies_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        path = request.path
        end_path = path.rsplit('/', 1)[-1]
        get_details = DetailsSupplies.objects.filter(hospital=self.request.user.hospital,id=end_path).last()
        if get_details:
            get_details.delete()
            return Response(status=status.HTTP_200_OK)
        else:
            errors = {"errors": ["Cannot delete item"]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='noneSupplies')
    def noneSupplies(self, request, *args, **kwargs):
        user = self.request.user
        DetailsSupplies.objects.filter(hospital=self.request.user.hospital,user_id=user.id, supplies_id=None).all().delete()
        return Response(status=status.HTTP_200_OK)


class DetailsPatientAccountViewSet(viewsets.ModelViewSet):
    queryset = DetailsPatientAccount.objects.all()
    serializer_class = DetailsPatientAccountSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DetailsPatientAccountFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return DetailsPatientAccount.objects.filter(hospital=user_hospital)
        return DetailsPatientAccount.objects.all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = DetailsPatientAccountForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            get_last_account = PatientAccount.objects.filter(hospital = self.request.user.hospital,id=request.data['patient_account'], type_account='PREPAID').last()
            get_last = DetailsPatientAccount.objects.filter(hospital = self.request.user.hospital,patient_account_id=get_last_account.id).last()
            obj.user_id = self.request.user.id
            obj.balance_before = get_last_account.balance
            obj.balance = int(request.data['balance'])
            if request.data['type_operation'] == 'CREDIT':
                if get_last:
                    amount_after = int(get_last_account.balance) + int(request.data['balance'])
                else:
                    amount_after = int(request.data['balance'])
            elif request.data['type_operation'] == 'DEBIT':
                if get_last:
                    amount_after = int(get_last_account.balance) - int(request.data['balance'])
                else:
                    amount_after = int(request.data['balance'])
            obj.balance_after = amount_after
            get_last_account.balance = amount_after 
            get_last_account.save()
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = DetailsPatientAccountForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            get_last_account = PatientAccount.objects.filter(hospital = self.request.user.hospital,id=request.data['patient_account'], type_account='PREPAID').last()
            get_last = DetailsPatientAccount.objects.filter(hospital = self.request.user.hospital,patient_account_id=get_last_account.id).last()
            obj.balance_before = get_last_account.balance
            obj.balance = int(request.data['balance'])
            if request.data['type_operation'] == 'CREDIT':
                amount_after = int(get_last.amount_after) + int(request.data['balance'])
            elif request.data['type_operation'] == 'DEBIT':
                amount_after = int(get_last.amount_after) - int(request.data['balance'])
            obj.balance_after = amount_after
            get_last_account.balance = amount_after 
            get_last_account.save()
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    
    @action(detail=False, methods=['get'], url_path='all')
    def get_all(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)    
    
    @action(detail=False, methods=['get'], url_path='export')
    def get_export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset()).filter(hospital_id=self.request.user.hospital.id)
        serializer = self.get_serializer(queryset, many=True, ).data
        html_render = get_template('extrait_compte.html')
        # logo = settings.MEDIA_ROOT + '/logo.png'
        html_content = html_render.render(
            {'lignes': serializer,
            'hospital': self.request.user.hospital,
            'debit':queryset.filter(type_operation='DEBIT').aggregate(Sum('balance'))['balance__sum'],
            'credit':queryset.filter(type_operation='CREDIT').aggregate(Sum('balance'))['balance__sum'],
            'date': datetime.today().strftime("%Y-%m-%d"),
            'balance_available': queryset[0].patient_account,
            'url': request.build_absolute_uri('/')})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                            link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(content_type='application/pdf')
            filename = 'Extrait' + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            response.write(result.getvalue())
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        # return Response(data=content, status=status.HTTP_200_OK)


class InsuranceViewSet(viewsets.ModelViewSet):
    queryset = Insurance.objects.filter(deleted=False)
    serializer_class = InsuranceSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = InsuranceFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Insurance.objects.filter(deleted=False, hospital=user_hospital)
        return Insurance.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        insurance_form = InsuranceForm(request.data)
        if insurance_form.is_valid():
            insurance = insurance_form.save()
            insurance.hospital = self.request.user.hospital
            insurance.save()
            serializer = self.get_serializer(insurance, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**insurance_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        insurance = self.get_object()
        insurance_form = InsuranceForm(request.data, instance=insurance)
        if insurance_form.is_valid():
            insurance = insurance_form.save()
            insurance.save()
            serializer = self.get_serializer(insurance, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**insurance_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        insurance = self.get_object()
        insurance.deleted = True
        insurance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='all')
    def get_all_insurance(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get', 'post'], url_path='details_insurances',
            permission_classes=[AllowAny])
    def details_insurances(self, request, *args, **kwargs):
        if request.method == 'GET':
            if request.query_params.get("insurance") != 'null':
                get_bills = Bills.objects.filter(hospital=self.request.user.hospital,patient__insurance_id=request.query_params.get("insurance"), deleted=False).all()
                serializer = BillsSerializer(get_bills, many=True)
                content={'content':serializer.data}
                return Response(data=content, status=status.HTTP_200_OK)
        else:
            get_details = DetailsSupplies.objects.create(hospital = self.request.user.hospital,supplies_id=request.data['id_supplies'],
                                                         product_id=request.data['product']['id'],
                                                         quantity=request.data['quantity'],
                                                         total_amount=request.data['total_amount'],
                                                         arrival_price=request.data['arrival_price'],
                                                         product_code=request.data['product_code'],
                                                         product_name=request.data['product_name'])
            details = DetailsSuppliesSerializer(get_details, many=False)
            return Response(data=details.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='upload', permission_classes=[AllowAny])
    def upload_file(self, request, *args, **kwargs):
        if request.FILES:
            if "file" in request.FILES:
                file = request.FILES.get("file")
                # fs = FileSystemStorage()
                # filename = fs.save(file.name, file)
                # uploaded_file_url = fs.url(filename)
                empexceldata = pd.read_excel(file)
                dbframe = empexceldata
                data = {}
                for dbframe in dbframe.itertuples():
                    get_obj = City.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.Nom).last()
                    if get_obj:
                        get_obj.name = dbframe.Nom
                        get_obj.save()
                    else:
                        if dbframe.Ville:
                            get_region = City.objects.filter(hospital=self.request.user.hospital,name__icontains='default').last()
                            if get_region:
                                Insurance.objects.create(hospital = self.request.user.hospital,city_id=get_region.id,name=dbframe.Nom,mailbox=checkContent(content=dbframe.Boite_postale),number=checkContent(content=dbframe.Phone),responsible=checkContent(content=dbframe.Responsable),percent=checkContent(content=dbframe.Pourcentage))
                            else:
                                create_region = City.objects.create(hospital = self.request.user.hospital,name='default')
                                Insurance.objects.create(hospital = self.request.user.hospital,city_id=create_region.id,name=dbframe.Nom,mailbox=checkContent(content=dbframe.Boite_postale),number=checkContent(content=dbframe.Phone),responsible=checkContent(content=dbframe.Responsable),percent=checkContent(content=dbframe.Pourcentage))
                        else:

                            get_region = City.objects.filter(hospital=self.request.user.hospital,name__icontains=dbframe.Ville).last()
                            if get_region:
                                Insurance.objects.create(hospital = self.request.user.hospital,city_id=get_region.id,name=dbframe.Nom,mailbox=checkContent(content=dbframe.Boite_postale),number=checkContent(content=dbframe.Phone),responsible=checkContent(content=dbframe.Responsable),percent=checkContent(content=dbframe.Pourcentage))
                            else:
                                create_region = City.objects.create(hospital = self.request.user.hospital,name=dbframe.Ville)
                                Insurance.objects.create(hospital = self.request.user.hospital,city_id=create_region.id,name=dbframe.Nom,mailbox=checkContent(content=dbframe.Boite_postale),number=checkContent(content=dbframe.Phone),responsible=checkContent(content=dbframe.Responsable),percent=checkContent(content=dbframe.Pourcentage))
                            

        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_name(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            reagent = Insurance.objects.filter(hospital=self.request.user.hospital,name__icontains=data['name'])
            if reagent:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='email/exists', permission_classes=[AllowAny])
    def check_email(self, request, *args, **kwargs):
        data = request.data
        errors = {"email": ["This field already exists."]}
        if 'email' in data:
            reagent = Insurance.objects.filter(hospital=self.request.user.hospital,email__icontains=data['email'])
            if reagent:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='number/exists', permission_classes=[AllowAny])
    def check_number(self, request, *args, **kwargs):
        data = request.data
        errors = {"number": ["This field already exists."]}
        if 'number' in data:
            reagent = Insurance.objects.filter(hospital=self.request.user.hospital,number__icontains=data['number'])
            if reagent:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='phone/exists', permission_classes=[AllowAny])
    def check_phone(self, request, *args, **kwargs):
        data = request.data
        errors = {"phone": ["This field already exists."]}
        if 'phone' in data:
            reagent = Insurance.objects.filter(hospital=self.request.user.hospital,phone__icontains=data['phone'])
            if reagent:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def profile_view(request):
    user = request.user
    if not isinstance(user, User):
        return Response(data={'message': _('Profile not found')}, status=status.HTTP_404_NOT_FOUND)
    profile = user.profile
    profile_form = ProfileForm(instance=profile, data=request.data or {})
    if profile_form.is_valid():
        profile = profile_form.save()
        serializer = ProfileSerializer(profile, many=False)
        return Response(
            data=serializer.data,
            status=status.HTTP_200_OK
        )
    return Response(
        data=profile_form.errors,
        status=status.HTTP_400_BAD_REQUEST
    )


from rest_framework_simplejwt.views import TokenRefreshView, TokenObtainPairView


class RefreshTokenView(TokenRefreshView):
    serializer_class = TokenRefreshSerializer


class LoginView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer

from django.core.handlers import exception
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_logout(request):
    if request.method == 'POST':
        try:
            get_user = User.objects.filter(id=request.user.id).last()
            get_user.is_online = False
            get_user.save()
            return Response(status=status.HTTP_200_OK)
        except:
            error = {'error': 'Error'}
            return Response(data=error, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class Storage_depotsViewSet(viewsets.ModelViewSet):
    queryset = Storage_depots.objects.filter(deleted=False)
    serializer_class = Storage_depotsSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = Storage_depotsFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        if self.request.user.hospital:
            user_hospital = self.request.user.hospital
            return Storage_depots.objects.filter(deleted=False, hospital=user_hospital)
        else:
            return Storage_depots.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = Storage_depotsForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.name_language = request.data['name_language']
            obj.save()
            for translate in self.request.data['name_language']:
                get_translate = WarehouseTranslation.objects.filter(hospital = self.request.user.hospital, warehouse_id=obj.id, language=translate['language']).last()
                if get_translate:
                    get_translate.name = translate['name']
                    get_translate.save()
                else:
                    WarehouseTranslation.objects.create(hospital = self.request.user.hospital, user=self.request.user, warehouse_id=obj.id, language=translate['language'], name = translate['name'])
            
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = Storage_depotsForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            for translate in self.request.data['name_language']:
                get_translate = WarehouseTranslation.objects.filter(hospital = self.request.user.hospital, warehouse_id=obj.id, language=translate['language']).last()
                if get_translate:
                    get_translate.name = translate['name']
                    get_translate.save()
                else:
                    WarehouseTranslation.objects.create(hospital = self.request.user.hospital, user=self.request.user, warehouse_id=obj.id, language=translate['language'], name = translate['name'])
            
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.deleted = True
        obj.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='name/exists', permission_classes=[AllowAny])
    def check_depart(self, request, *args, **kwargs):
        data = request.data
        errors = {"name": ["This field already exists."]}
        if 'name' in data:
            storage_depots = Storage_depots.objects.filter(name__icontains=data['name'])
            if storage_depots:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], url_path='all')
    def get_all_cash(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, ).data
        content = {'content': serializer}
        # patient_id = self.request.query_params.get("patient_id")
        # get_bills = Bills.objects.filter(patient_id=patient_id).aggregate(Sum('balance'))['balance__sum']
        # content = {'content': {'solde_patient': get_bills}}
        return Response(data=content, status=status.HTTP_200_OK)



class DetailsStock_movementViewSet(viewsets.ModelViewSet):
    queryset = DetailsStock_movement.objects.all()
    serializer_class = DetailsStock_movementSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DetailsStock_movementFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user = self.request.user

        qs = (
            DetailsStock_movement.objects
            .select_related('hospital')   # très important
        )

        if user.hospital_id:
            qs = qs.filter(hospital_id=user.hospital_id)

        return qs
    
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        user = self.request.user
        if user.hospital.is_inventory == False:
            detailsStock_movement_form = DetailsStock_movementForm(data=request.data)
            if detailsStock_movement_form.is_valid():
                get_detailsStock_movement = DetailsStock_movement.objects.filter(hospital = user.hospital,
                    ingredient_id=request.data['ingredient'],
                    storage_depots_id=request.data[
                        'storage_depots'],
                    stock_movement=request.data['stock_movement'], user_id=user.id).last()
                if get_detailsStock_movement:
                    if request.data['type_movement'] == 'TRANSFER' and int(request.data['storage_depots']) == int(request.data['storage_depots_dest']):
                        errors = {"errors": ["Destination must be different from source."]}
                        return Response(errors, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        
                        get_detailsStock_movement.quantity = request.data['quantity']
                        get_detailsStock_movement.save()
                        return Response(status=status.HTTP_201_CREATED)
                        


                
                # else:
                #     get_details_stock.qte_stock = get_details_stock.qte_stock + int(
                #         get_detailsStock_movement.quantity) - int(
                #         request.data['quantity'])
                #     get_details_stock.save()
                #     get_detailsStock_movement.quantity = get_detailsStock_movement.quantity + int(
                #         get_detailsStock_movement.quantity) - int(request.data['quantity'])
                #     get_detailsStock_movement.total_amount = get_detailsStock_movement.total_amount + int(
                #         get_detailsStock_movement.total_amount) - int(
                #         request.data['total_amount'])
                #     get_detailsStock_movement.unit_price = request.data['unit_price']
                #     get_detailsStock_movement.save()
                #     return Response(status=status.HTTP_201_CREATED)
                else:
                    
                    detailsStock_movement = detailsStock_movement_form.save()
                    detailsStock_movement.hospital_id = user.hospital.id
                    detailsStock_movement.stock_movement_id = request.data['stock_movement']
                    detailsStock_movement.user_id = user.id
                    detailsStock_movement.timeAt = time.strftime("%H:%M:%S", time.localtime())
                    detailsStock_movement.save()
                    return Response(status=status.HTTP_201_CREATED)
                        

            errors = {**detailsStock_movement_form.errors}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            errors = {"is_inventory": ["Current inventory."]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        path = request.path
        end_path = path.rsplit('/', 1)[-1]
        get_details = DetailsSupplies.objects.filter(hospital = self.request.user.hospital,id=end_path).last()
        if get_details:
            get_details.supplies_id = request.data['supplies']
            get_details.product_id = get_details.product
            get_details.quantity = request.data['quantity']
            get_details.total_amount = request.data['total_amount']
            get_details.product_code = request.data['product_code']
            get_details.product_name = request.data['product_name']
            get_details.arrival_price = request.data['arrival_price']
            get_details.save()
            get_sum_details = DetailsSupplies.objects.filter(hospital = self.request.user.hospital,supplies_id=request.data['supplies']).aggregate(
                Sum('total_amount'))
            get_supplies = Supplies.objects.filter(hospital = self.request.user.hospital,id=get_details.supplies.id).last()
            get_supplies.supply_amount = get_sum_details['total_amount__sum']
            get_supplies.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            errors = {"errors": ["Already exist"]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        path = request.path
        end_path = path.rsplit('/', 1)[-1]
        get_details = DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,id=end_path, stock_movement=None).last()
        if get_details:
            get_details.delete()
            return Response(status=status.HTTP_200_OK)
        else:
            errors = {"errors": ["Cannot delete item"]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='details')
    def get_obj_details(self, request):
        user = self.request.user
        if self.request.GET.get("stock_movement"):
            if self.request.GET.get("stock_movement") == 'undefined' or self.request.GET.get(
                    "stock_movement") == 'null':
                serializer = self.get_serializer(DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,user_id=user.id, stock_movement_id=None,
                                                            deleted=False).all(), many=True)
                content={"content":serializer.data}
                return Response(data=content,status=status.HTTP_200_OK)
            elif self.request.GET.get("stock_movement") == 'reset':
                DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,user_id=user.id, stock_movement_id=None, deleted=False).all().delete()
                return Response(status=status.HTTP_200_OK)
            else:
                serializer = self.get_serializer(DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,
                stock_movement_id=self.request.GET.get("stock_movement"), deleted=False).all(), many=True)
                content={"content":serializer.data}
                return Response(data=content,status=status.HTTP_200_OK)
        else:
            serializer = self.get_serializer(DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,user_id=user.id, stock_movement_id=None, deleted=False).all(), many=True)
            content={"content":serializer.data}
            return Response(data=content,status=status.HTTP_200_OK)


    @action(detail=False, methods=['get'], url_path='noneMov')
    def noneMov(self, request, *args, **kwargs):
        user = self.request.user
        DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,user_id=user.id, stock_movement_id=None).all().delete()
        return Response(status=status.HTTP_200_OK)

    # @action(detail=False, methods=['get'], url_path='stock_available')
    # def stock_available(self, request):
    #     get_detail_depot = DetailsStock.objects.filter(storage_depots_id=request.query_params.get("id"))
    #     serializer = DetailsStockSerializer(get_detail_depot, many=True)
    #     content = {'content': serializer.data}
    #     return Response(data=content, status=status.HTTP_200_OK)

    # @action(detail=False, methods=['get'], url_path='get_items')
    # def get_items(self, request):
    #
    #     startdate = request.query_params.get("start_date")
    #     enddate = request.query_params.get("end_date")
    #     get_product_supplies = DetailsSupplies.objects.filter(storage_depots_id=request.query_params.get("id"),
    #                                                           product_id=request.query_params.get("product"),
    #                                                           createdAt__range=[startdate,enddate])
    #
    #     serializer_supplies = DetailsSuppliesSerializer(get_product_supplies, many=True)
    #     get_product_bills = DetailsBills.objects.filter(storage_depots_id=request.query_params.get("id"),
    #                                                     details_stock__product__id=request.query_params.get("product"),
    #                                                     createdAt__range=[startdate,enddate])
    #     serializer_bills = DetailsBillsSerializer(get_product_bills, many=True)
    #     get_product_stock = DetailsStock_movement.objects.filter(storage_depots_id=request.query_params.get("id"),
    #                                                              details_stock__product__id=request.query_params.get(
    #                                                                  "product"), createdAt__range=[startdate,enddate])
    #     serializer_stock = DetailsStock_movementSerializer(get_product_stock, many=True)
    #     content = {'content': {'bills': serializer_bills.data, 'supplies': serializer_supplies.data,
    #                            'stock': serializer_stock.data, }}
    #     return Response(data=content, status=status.HTTP_200_OK)
    #
    @action(detail=False, methods=['get'], url_path='get_items')
    def get_items(self, request):
        startdate = self.request.query_params.get("start_date")
        enddate = self.request.query_params.get("end_date")

        # get_product_bills = DetailsBills.objects.filter(medical_act=None, bills__storage_depots_id=self.request.query_params.get("storage_depots"),
        #                                                 deleted=False,
        #                                                 details_stock__product_id=self.request.query_params.get("product"),
        #                                                 createdAt__range=[startdate, enddate])
        if 'product' in self.request.query_params:
            get_product_bills = DetailsBills.objects.filter(hospital = self.request.user.hospital,medical_act=None,
                                                            bills__storage_depots_id=self.request.query_params.get(
                                                                "storage_depots"),
                                                            deleted=False,
                                                            details_stock__product_id=self.request.query_params.get(
                                                                "product"),
                                                            createdAt__range=[startdate, enddate])

            get_product_supplies = DetailsSupplies.objects.filter(hospital = self.request.user.hospital,
                storage_depots_id=self.request.query_params.get("storage_depots"),
                product_id=self.request.query_params.get("product"),
                createdAt__range=[startdate, enddate])

            get_product_stock = DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,
                storage_depots_id=self.request.query_params.get("storage_depots"),
                details_stock__product_id=self.request.query_params.get("product"),
                createdAt__range=[startdate, enddate])
            content = {'content': {'bill': DetailsBillsSerializer(get_product_bills, many=True).data,
                                   'bills': DetailsBillsSerializer(get_product_bills,
                                                                   many=True).data + DetailsSuppliesSerializer(
                                       get_product_supplies, many=True).data + DetailsStock_movementSerializer(
                                       get_product_stock,
                                       many=True).data}}

        else:
            # get_product_bills = DetailsBills.objects.filter(medical_act=None,
            #                                                 bills__storage_depots_id=self.request.query_params.get(
            #                                                     "storage_depots"),
            #                                                 deleted=False,
            #                                                 createdAt__range=[startdate, enddate])
            # get_product_stock = DetailsStock_movement.objects.filter(
            #     storage_depots_id=self.request.query_params.get("storage_depots"),
            #     createdAt__range=[startdate, enddate])
            # get_product_supplies = DetailsSupplies.objects.filter(
            #     storage_depots_id=self.request.query_params.get("storage_depots"),
            #     createdAt__range=[startdate, enddate])
            content = {'content': []}
        return Response(data=content, status=status.HTTP_200_OK)


class InventoryViewSet(viewsets.ModelViewSet):
    queryset = Inventory.objects.filter(deleted=False)
    serializer_class = InventorySerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = InventoryFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Inventory.objects.filter(deleted=False, hospital=user_hospital)
        return Inventory.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        inventory_form = InventoryForm(request.data)
        if inventory_form.is_valid():
            # inventory = inventory_form.save()
            # inventory.hospital = self.request.user.hospital
            # # inventory_update(user=user, inventory=inventory, request=request)
            # inventory.save()
            get_inventory = Inventory.objects.filter(hospital = self.request.user.hospital, user=self.request.user, id=request.data['inventory']).last()
            get_inventory.storage_depots_id = request.data['storage_depots']
            get_inventory.save()
            get_details_inv = DetailsInventory.objects.filter(hospital = self.request.user.hospital, inventory_id=request.data[
                    'inventory']).all()

            for inv in get_details_inv:
                get_details_stock = Stock.objects.filter(hospital = self.request.user.hospital, ingredient_id=inv.ingredient.id,
                                                                storage_depots_id=request.data[
                                                                    'storage_depots']).last()
                get_details_stock.quantity = inv.quantity_adjusted
                get_details_stock.save()
            # get_hospital = Hospital.objects.filter(id=self.request.user.hospital.id).last()
            # get_hospital.is_inventory = False
            # get_hospital.save()
            serializer = self.get_serializer(get_inventory, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**inventory_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        inventory = self.get_object()
        inventory_form = InventoryForm(request.data, instance=inventory)
        if inventory_form.is_valid():
            inventory = inventory_form.save()
            inventory.save()
            serializer = self.get_serializer(inventory, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**inventory_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        inventory = self.get_object()
        inventory.deleted = True
        inventory.save()
        return Response(status=status.HTTP_204_NO_CONTENT)
       
    @action(detail=True, methods=['DELETE'], url_path='delete-empty')
    def delete_empty(self, request, pk=None):
        supply = self.get_object()

        has_lines = DetailsInventory.objects.filter(Inventory=supply, hospital=self.request.user.hospital).exists()

        if has_lines:
            return Response(
                {"detail": "Cannot delete, lines exist."},
                status=status.HTTP_400_BAD_REQUEST
            )

        supply.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='create-empty')
    def create_empty(self, request):
        lines = Inventory.objects.annotate(
            line_count=Count('inventories'),
        ).filter(line_count=0, hospital=self.request.user.hospital, user_id=self.request.user.id)
        lines.delete()

        supply = Inventory.objects.create(
            hospital=self.request.user.hospital,
            user_id = request.user.id
        )
        return Response(data={"id": supply.id}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='close')
    def close(self, request, *args, **kwargs):
        user = self.request.user
        get_cash = Cash.objects.filter(hospital = self.request.user.hospital,user_id=user.id, is_active=True).last()
        get_cash.is_active = False
        get_cash.save()
        return Response(status=status.HTTP_200_OK)



class DetailsInventoryViewSet(viewsets.ModelViewSet):
    queryset = DetailsInventory.objects.all()
    serializer_class = DetailsInventorySerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DetailsInventoryFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user = self.request.user

        qs = (
            DetailsInventory.objects
            .select_related('hospital')   # très important
        )

        if user.hospital_id:
            qs = qs.filter(hospital_id=user.hospital_id)

        return qs
        
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        user = self.request.user
        details_inventory_form = DetailsInventoryForm(data=request.data)
        if details_inventory_form.is_valid():
            get_details_inventory = DetailsInventory.objects.filter(hospital = self.request.user.hospital,
                ingredient_id=request.data['ingredient'],
                inventory=request.data[
                    'inventory'], user_id=user.id).last()
            if get_details_inventory:
                get_details_inventory.amount = request.data['amount']
                get_details_inventory.amount_adjusted = request.data['amount_adjusted']
                get_details_inventory.quantity_adjusted = request.data['quantity_adjusted']
                get_details_inventory.save()

                return Response(status=status.HTTP_201_CREATED)
            else:
                details_inventory = details_inventory_form.save()
                details_inventory.hospital = user.hospital
                details_inventory.user_id = user.id
                details_inventory.save()
                return Response(status=status.HTTP_201_CREATED)
        errors = {**details_inventory_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        path = request.path
        end_path = path.rsplit('/', 1)[-1]
        get_details = DetailsSupplies.objects.filter(hospital = self.request.user.hospital,id=end_path).last()
        if get_details:
            get_details.supplies_id = request.data['supplies']
            get_details.product_id = get_details.product
            get_details.quantity = request.data['quantity']
            get_details.total_amount = request.data['total_amount']
            get_details.product_code = request.data['product_code']
            get_details.product_name = request.data['product_name']
            get_details.arrival_price = request.data['arrival_price']
            get_details.save()
            get_sum_details = DetailsSupplies.objects.filter(hospital = self.request.user.hospital,supplies_id=request.data['supplies']).aggregate(
                Sum('total_amount'))
            get_supplies = Supplies.objects.filter(hospital = self.request.user.hospital,id=get_details.supplies.id).last()
            get_supplies.supply_amount = get_sum_details['total_amount__sum']
            get_supplies.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            errors = {"errors": ["Already exist"]}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        stock_movement = self.get_object()
        stock_movement.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='noneInv')
    def noneInv(self, request, *args, **kwargs):
        user = self.request.user
        DetailsInventory.objects.filter(hospital = self.request.user.hospital,user_id=user.id, inventory_id=None).all().delete()
        return Response(status=status.HTTP_200_OK)

    #
    # @action(detail=False, methods=['get'], url_path='stock_available')
    # def stock_available(self, request):
    #     get_detail_depot = DetailsStock.objects.filter(storage_depots_id=request.query_params.get("id"))
    #     serializer = DetailsStockSerializer(get_detail_depot, many=True)
    #     # get_product = DetailsSupplies.objects.filter(supplies__storage_depots=request.query_params.get("id")).order_by(
    #     #     'product_id').distinct("product_id")
    #     # productList = []
    #     # for product in get_product:
    #     #     get_prod = Product.objects.filter(id=product.product_id).last()
    #     #     productList.append(get_prod)
    #     # serializer = DetailsSuppliesSerializer(get_product, many=True)
    #     content = {'content': serializer.data}
    #     return Response(data=content, status=status.HTTP_200_OK)
    #


class Stock_movementViewSet(viewsets.ModelViewSet):
    queryset = Stock_movement.objects.filter(deleted=False)
    serializer_class = Stock_movementSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = Stock_movementFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return Stock_movement.objects.filter(deleted=False, hospital=user_hospital)
        return Stock_movement.objects.filter(deleted=False)

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        user=self.request.user
        stock_movement_form = Stock_movementForm(request.data)
        if stock_movement_form.is_valid():
            
            get_details_stock_movement = DetailsStock_movement.objects.filter(hospital = user.hospital, stock_movement=request.data['stock_movement']).all()
            for stock_mov in get_details_stock_movement:

                get_details_stock = Stock.objects.filter(hospital = user.hospital,
                    ingredient_id=stock_mov.ingredient.id,
                    storage_depots_id=stock_mov.storage_depots).last()
                stock_mov.stock_initial = get_details_stock.quantity
                stock_mov.save()
                if stock_mov.type_movement == 'EXIT':
                    get_details_stock.quantity -= int(
                        stock_mov.quantity)
                    get_details_stock.save()
                elif stock_mov.type_movement == 'TRANSFER':
                    get_details_stock.quantity -= int(
                        stock_mov.quantity)

                    get_details_stock_dest = Stock.objects.filter(hospital = user.hospital,
                        ingredient_id=stock_mov.ingredient.id,
                        storage_depots_id=stock_mov.storage_depots_dest).last()
                    if get_details_stock_dest:
                        get_details_stock_dest.quantity += int(
                            stock_mov.quantity)
                        get_details_stock.save()
                        get_details_stock_dest.save()
                    else:
                        get_details_stock.save()
                        get_details_stock_dest = Stock.objects.create(hospital = user.hospital,ingredient_id=get_details_stock.ingredient.id, quantity = int(stock_mov.quantity), storage_depots_id=stock_mov.storage_depots_dest.id)
                    
            get_stock_movement = Stock_movement.objects.filter(id=request.data['stock_movement'],hospital = user.hospital,user=user).last()
            get_stock_movement.storage_depots_id = request.data['storage_depots']
            get_stock_movement.type_movement = request.data['type_movement']
            get_stock_movement.reason_movement = request.data['reason_movement']
            get_stock_movement.movement_value = request.data['movement_value']
            get_stock_movement.save()
            
            serializer = self.get_serializer(get_stock_movement, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**stock_movement_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = Stock_movementForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            if obj.type_movement == 'TRANSFER' and request.data['is_valid'] == True:
                get_details_stock_movement = DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,
                storage_depots_id=request.data['storage_depots'], stock_movement_id=obj.id).all()
                get_details_stock_movement_first = DetailsStock_movement.objects.filter(hospital = self.request.user.hospital,
                storage_depots_id=request.data['storage_depots'], stock_movement_id=obj.id).first()
                message= f'Effectue lors de transfert des produits du document de stock {obj.code}'
                save_mvt_entry = Supplies.objects.create(hospital = self.request.user.hospital,storage_depots_id=get_details_stock_movement_first.storage_depots_dest.id, additional_info=message)
                for stock_mov in get_details_stock_movement:
                    get_details_stock = DetailsStock.objects.filter(hospital = self.request.user.hospital,
                    product_id=stock_mov.ingredient_id,
                    storage_depots_id=request.data['storage_depots']).last()
                    get_details_stock_dest = DetailsStock.objects.filter(hospital = self.request.user.hospital,
                        product_id=stock_mov.details_stock.product_id,
                        storage_depots_id=stock_mov.storage_depots_dest).last()
                    if not get_details_stock_dest:
                        DetailsStock.objects.create(hospital = self.request.user.hospital,product_id=stock_mov.details_stock.product_id,product_name=stock_mov.details_stock.product_id.name,
                        storage_depots_id=stock_mov.storage_depots_dest)
                        get_details_stock_dest = DetailsStock.objects.filter(hospital = self.request.user.hospital,
                        product_id=stock_mov.details_stock.product_id,
                        storage_depots_id=stock_mov.storage_depots_dest).last()
                        get_details_stock.qte_stock = get_details_stock.qte_stock - int(
                                stock_mov.quantity)
                        get_details_stock.save()
                        get_details_stock_dest.qte_stock = get_details_stock_dest.qte_stock + int(
                            stock_mov.quantity)
                        get_details_stock_dest.save()
                    else:
                        get_details_stock.qte_stock = get_details_stock.qte_stock - int(
                                stock_mov.quantity)
                        get_details_stock.save()
                        get_details_stock_dest.qte_stock = get_details_stock_dest.qte_stock + int(
                            stock_mov.quantity)
                        get_details_stock_dest.save()
                    DetailsSupplies.objects.create(hospital = self.request.user.hospital,supplies_id=save_mvt_entry.id, storage_depots_id=stock_mov.storage_depots_dest.id, product_id=stock_mov.details_stock.product_id, quantity=stock_mov.quantity)

            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        stock_movement = self.get_object()
        stock_movement.deleted = True
        stock_movement.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    #
    # @action(detail=False, methods=['get'], url_path='get_stock_movement')
    # def get_stock_movement(self, request, *args, **kwargs):
    #     stock_movement = Stock_movement.objects.filter(
    #         storage_depots_id=request.query_params.get("id"))
    #     serializer = Stock_movementSerializer(stock_movement, many=True)
    #     content = {'content': serializer.data}
    #     return Response(data=content, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='close')
    def close(self, request, *args, **kwargs):
        user = self.request.user
        get_cash = Cash.objects.filter(user_id=user.id, is_active=True).last()
        get_cash.is_active = False
        get_cash.save()
        return Response(status=status.HTTP_200_OK)
    
    
    @action(detail=True, methods=['DELETE'], url_path='delete-empty')
    def delete_empty(self, request, pk=None):
        supply = self.get_object()

        has_lines = DetailsStock_movement.objects.filter(stock_movement=supply, hospital=request.user.hospital).exists()

        if has_lines:
            return Response(
                {"detail": "Cannot delete, lines exist."},
                status=status.HTTP_400_BAD_REQUEST
            )

        supply.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='create-empty')
    def create_empty(self, request):
        lines = Stock_movement.objects.annotate(
            line_count=Count('stock_movements'),
        ).filter(line_count=0, hospital=self.request.user.hospital, user_id=self.request.user.id)
        lines.delete()

        supply = Stock_movement.objects.create(
            hospital=self.request.user.hospital,
            movement_value=0,
            user_id = request.user.id
        )
        return Response(data={"id": supply.id}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='print')
    def print(self, request, *args, **kwargs):
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = Stock_movementSerializer(queryset, many=True)

        html_render = get_template('export_stock_movement.html')
        html_content = html_render.render(
            {'stock_movements': serializer.data, 'hospital': self.request.user.hospital,
             'date': datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result,
                                        link_callback=link_callback)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = 'Export' + datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '.pdf'
            response['Content-Disposition'] = 'inline; filename="' + filename + '"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
            return response
        else:
            errors = {"pdf": ["Error to generate PDF."]}
            return Response(data=errors, status=status.HTTP_500)

    @action(detail=False, methods=['get'], url_path='export')
    def exportMedicalAct(self, request, *args, **kwargs):
        if self.request.query_params == {}:
            queryset = Stock_movement.objects.all()
        else:
            queryset = self.filter_queryset(self.get_queryset())
        serializer = Stock_movementSerializer(queryset, many=True)
        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Mouvements de stock', index=1)
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.freeze_panes = 'I1'

        columns = ['Code', 'Type de mouvement', 'Motifs', 'Valeur', 'Date du mouvement']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        # Iterate through all coins

        for _, act in enumerate(serializer.data, 1):
            row_num += 1

            row = [
                act['code'],
                act['type_movement'],
                act['reason_movement'],
                act['movement_value'],
                act['createdAt'],
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(excelfile)
        response = HttpResponse(content_type='application/vnd.ms-excel')

        # tell the browser what the file is named
        response['Content-Disposition'] = 'attachment;filename="some_file_name.xlsx"'

        # put the spreadsheet data into the response
        response.write(excelfile.getvalue())

        # return the response
        return response


class DeliveryInfoViewSet(viewsets.ModelViewSet):
    queryset = DeliveryInfo.objects.all()
    serializer_class = DeliveryInfoSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DeliveryInfoFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return DeliveryInfo.objects.filter(hospital=user_hospital)
        return DeliveryInfo.objects.all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = DeliveryInfoForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = DeliveryInfoForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['post'], url_path='title/exists', permission_classes=[AllowAny])
    def check_obj(self, request, *args, **kwargs):
        data = request.data
        errors = {"title": ["This field already exists."]}
        if 'title' in data:
            obj = DeliveryInfo.objects.filter(title__icontains=data['title'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

class EventInfoViewSet(viewsets.ModelViewSet):
    queryset = EventInfo.objects.all()
    serializer_class = EventInfoSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = EventInfoFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return EventInfo.objects.filter(hospital=user_hospital)
        return EventInfo.objects.all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = EventInfoForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = EventInfoForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['post'], url_path='title/exists', permission_classes=[AllowAny])
    def check_obj(self, request, *args, **kwargs):
        data = request.data
        errors = {"title": ["This field already exists."]}
        if 'title' in data:
            obj = EventInfo.objects.filter(title__icontains=data['title'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

class CateringInfoViewSet(viewsets.ModelViewSet):
    queryset = CateringInfo.objects.all()
    serializer_class = CateringInfoSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = CateringInfoFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return CateringInfo.objects.filter(hospital=user_hospital)
        return CateringInfo.objects.all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = CateringInfoForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = CateringInfoForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['post'], url_path='title/exists', permission_classes=[AllowAny])
    def check_obj(self, request, *args, **kwargs):
        data = request.data
        errors = {"title": ["This field already exists."]}
        if 'title' in data:
            obj = CateringInfo.objects.filter(title__icontains=data['title'])
            if obj:
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(errors, status=status.HTTP_400_BAD_REQUEST)



class DetailsBillsIngredientViewSet(viewsets.ModelViewSet):
    queryset = DetailsBillsIngredient.objects.all()
    serializer_class = DetailsBillsIngredientSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = DetailsBillsIngredientFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return DetailsBillsIngredient.objects.filter(hospital=user_hospital)
        return DetailsBillsIngredient.objects.all()

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        
        user = self.request.user
        obj_form = DetailsBillsIngredientForm(request.data)
        if obj_form.is_valid():
            parts = request.data['item_uid'].split('-')
            # filters = {
            #     "details_bills": request.data['details_bills'],
            #     "ingredient": int(parts[1]),
            #     "hospital": user_hospital,
            # }
            # get_details = DetailsBillsIngredient.objects.filter(**filters).last()
            # if get_details:

            #     get_details.quantity = request.data["quantity"]
            #     get_details.total_amount = float(request.data["quantity"]) * float(request.data['impact_price'])
            #     get_details.action = request.data["action"]
            #     get_details.save()
            #     return Response(status=status.HTTP_201_CREATED)
            # obj = obj_form.save()
            # obj.hospital = self.request.user.hospital
            # obj.save()
            # serializer = self.get_serializer(obj, many=False)
            # return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            queryset = Ingredient.objects.filter(hospital=user.hospital,translations__name__icontains=request.data['name']).last()
            querysetCompose = ComposeIngredient.objects.filter(hospital=user.hospital,translations__name__icontains=request.data['name']).last()
            

            if queryset:
                filters = {
                    "details_bills": request.data['details_bills'],
                    "ingredient": int(parts[1]),
                    "hospital": user.hospital,
                }
                get_details = DetailsBillsIngredient.objects.filter(**filters).last()
                if get_details:

                    get_details.quantity = request.data["quantity"]
                    get_details.total_amount = float(request.data["quantity"]) * float(request.data['impact_price'])
                    get_details.action = request.data["action"]
                    get_details.save()
                    return Response(status=status.HTTP_201_CREATED)

                else:
                    obj = obj_form.save()
                    obj.hospital = self.request.user.hospital
                    obj.save()
                    
                    serializer = self.get_serializer(obj, many=False)
                    return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            elif querysetCompose:
                filters = {
                    "details_bills": request.data['details_bills'],
                    "compose_ingredient": int(parts[1]),
                    "hospital": user.hospital,
                }
                get_details = DetailsBillsIngredient.objects.filter(**filters).last()
                if get_details:

                    get_details.quantity = request.data["quantity"]
                    get_details.total_amount = float(request.data["quantity"]) * float(request.data['impact_price'])
                    get_details.action = request.data["action"]
                    get_details.save()
                    return Response(status=status.HTTP_201_CREATED)
                obj = obj_form.save()
                obj.hospital = self.request.user.hospital
                obj.save()
                serializer = self.get_serializer(obj, many=False)
                return Response(data=serializer.data, status=status.HTTP_201_CREATED)
            errors = {**obj_form.errors}
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = DetailsBillsIngredientForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.total_amount = float(request.data['quantity']) * request.data['impact_price']
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
class MovementStockViewSet(viewsets.ModelViewSet):
    queryset = MovementStock.objects.all()
    serializer_class = MovementStockSerializer
    renderer_classes = [JSONRenderer]
    permission_classes = (IsAuthenticated, DjangoModelPermissions)
    pagination_class = CustomPagination
    filterset_class = MovementStockFilter
    filter_backends = (filters.DjangoFilterBackend,)

    def get_queryset(self):
        user_hospital = self.request.user.hospital
        if user_hospital:
            return MovementStock.objects.filter(hospital=user_hospital)
        return MovementStock.objects.all()


    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action == 'create':
            user = self.request.user
            if isinstance(user, User):
                permission_classes = [IsAuthenticated]
            else:
                permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        obj_form = DetailsBillsIngredientForm(request.data)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.hospital = self.request.user.hospital
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_form = DetailsBillsIngredientForm(request.data, instance=obj)
        if obj_form.is_valid():
            obj = obj_form.save()
            obj.total_amount= self.request.data['quantity'] * self.request.data['impact_price']
            obj.save()
            serializer = self.get_serializer(obj, many=False)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        errors = {**obj_form.errors}
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)


    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
